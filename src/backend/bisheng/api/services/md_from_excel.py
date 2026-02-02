import math
import os
from typing import List
from uuid import uuid4

import openpyxl
import pandas as pd
from loguru import logger


def xls_to_xlsx(xls_path):
    if not xls_path.lower().endswith(".xls"):
        return None

    if not os.path.exists(xls_path):
        return None

    try:
        xls_file = pd.ExcelFile(xls_path)
        sheets_to_write = {}

        # 2. Iterate through all worksheets, check if empty, and save non-empty content to the dictionary
        for sheet_name in xls_file.sheet_names:
            df = xls_file.parse(sheet_name)
            # df.empty will judge DataFrame No data (the number of rows is0）
            if not df.empty:
                sheets_to_write[sheet_name] = df
            else:
                #  Discard Blank Sheet
                pass

        # 3. Do not create a new file if there are any non-empty worksheets
        if not sheets_to_write:
            return None

        # 4. Write a new file if a non-empty worksheet exists
        xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets_to_write.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return xlsx_path

    except Exception as e:
        logger.exception(f'xls_to_xlsx error: ')
        return None


def remove_characters(s, chars_to_remove=["\n", "\r"]):
    """
    Removes the specified character from the string.
    """
    if not isinstance(s, str):
        return s
    for char in chars_to_remove:
        s = s.replace(char, "")
    return s.strip()


def unmerge_and_read_sheet(sheet_obj):
    """
    read out openpyxl Sheet object, unmerge cells by populating the top-left value of the merge range into all cells in the range.
    and returns the data as a list of lists.
    """
    if sheet_obj.max_row == 0 or sheet_obj.max_column == 0:
        return []
    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column
    data_grid = [
        [None for _ in range(max_column)] for _ in range(max_row)
    ]

    # Berturut-turut50Row Blank Row Stop Reading Content
    empty_row_num = 0
    max_empty_rows = 50
    empty_row_end = 0
    for r_idx, row in enumerate(sheet_obj.iter_rows()):
        if empty_row_num > max_empty_rows:
            break
        row_empty = True
        for c_idx, cell in enumerate(row):
            data_grid[r_idx][c_idx] = cell.value
            if cell.value:
                row_empty = False
        if row_empty:
            empty_row_num += 1
        else:
            empty_row_num = 0
    if empty_row_num > 0:
        data_grid = data_grid[:-empty_row_num]

    merged_cell_ranges = list(sheet_obj.merged_cells.ranges)
    for merged_range in merged_cell_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_cell_value = sheet_obj.cell(row=min_row, column=min_col).value
        # ignore empty rows
        if min_row > len(data_grid):
            continue
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                data_grid[r - 1][c - 1] = top_left_cell_value

    return data_grid


def generate_markdown_table_string(
        header_rows_list_of_lists,
        data_rows_list_of_lists,
        num_columns,
        separator_placement_index=1,
):
    """
    Generate from new rulesMarkdownTable String
    Automatically close purchase order afterheader_rows_list_of_listsIf empty, no headers and delimiters are generated.
    """
    md_lines = []

    # Handle headers and delimiters only if a header row is provided
    if header_rows_list_of_lists:
        pre_separator_header = header_rows_list_of_lists[:separator_placement_index]
        for row_values in pre_separator_header:
            md_lines.append(
                "| "
                + " | ".join(
                    remove_characters(str(v)) if v is not None else ""
                    for v in row_values
                )
                + " |"
            )

        # Insert below the header in the first rowMarkdownSeparator
        if num_columns > 0:
            md_lines.append("|" + "---|" * num_columns)

        post_separator_header = header_rows_list_of_lists[separator_placement_index:]
        for row_values in post_separator_header:
            md_lines.append(
                "| "
                + " | ".join(
                    remove_characters(str(v)) if v is not None else ""
                    for v in row_values
                )
                + " |"
            )

    # Always process data rows
    for row_values in data_rows_list_of_lists:
        md_lines.append(
            "| "
            + " | ".join(
                remove_characters(str(v)) if v is not None else "" for v in row_values
            )
            + " |"
        )

    return "\n".join(md_lines)


def process_dataframe_to_markdown_files(
        df,
        sheet_index: str,
        num_header_rows,
        rows_per_markdown,
        output_dir,
        append_header=True,
):
    """
    - append_header=True: Tekan num_header_rows Separate the header and data.
    - append_header=False: All content is treated as data, table header is empty, ignored num_header_rows。
    """
    if df.empty:
        logger.warning(f"  feed '{sheet_index}' DataDataFrameEmpty, skippingMarkdownBuat")
        return

    num_columns = df.shape[1]
    rows = df.shape[0]

    if rows == 0 or num_columns == 0:
        return

    header_block_df = pd.DataFrame()
    start_header_idx, end_header_idx = num_header_rows[0], num_header_rows[1]
    if start_header_idx >= rows:
        append_header = False

    # --- Core Logic Modified: According to append_header Decide how to split the data ---
    if append_header:
        # Handle header index outliers based on user rules
        if start_header_idx >= rows:
            logger.warning(
                f"Table Header Start Row {start_header_idx} Total lines exceeded {rows}. The first row will be used as the table header.")
            start_header_idx, end_header_idx = 0, 0
        elif end_header_idx >= rows:
            logger.warning(
                f"Table Header End Row {end_header_idx} Total lines exceeded {rows}. will be truncated to the last line.")
            end_header_idx = rows - 1

        # Make sure the index is legitimate
        if start_header_idx < 0: start_header_idx = 0
        if end_header_idx < start_header_idx: end_header_idx = start_header_idx

        try:
            header_slice = slice(start_header_idx, end_header_idx + 1)
            header_block_df = df.iloc[header_slice]
            data_block_df = df.drop(df.index[header_slice]).reset_index(drop=True)
            header_rows_as_lists = header_block_df.values.tolist()
        except Exception as e:
            logger.error(
                f"  At Source '{sheet_index}' Index by header in [{start_header_idx}, {end_header_idx}] Error Splitting Data: {e}Skip")
            return
    else:
        # when append_header are False , everything is treated as data and the header list is empty
        header_rows_as_lists = []
        data_block_df = df.reset_index(drop=True)

    # --- Subsequent pagination logic ---
    if data_block_df.empty:
        if append_header and not header_block_df.empty:
            markdown_content = generate_markdown_table_string(
                header_rows_as_lists, [], num_columns
            )
            # BUG FIX: Use zfill for proper padding. This is file '000' for the sheet.
            file_name = f"{str(sheet_index).zfill(2)}000.md"
            file_path = os.path.join(output_dir, file_name)
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(markdown_content)
                logger.debug(f"  Header-only files saved:'{file_path}'")
            except Exception as e:
                logger.debug(f"  Save file '{file_path}' Error during: {e}")
        return

    num_data_rows_total = len(data_block_df)
    num_files_to_create = math.ceil(num_data_rows_total / rows_per_markdown) if rows_per_markdown > 0 else (
        1 if num_data_rows_total > 0 else 0)

    for i in range(num_files_to_create):
        start_idx = i * rows_per_markdown
        end_idx = min(start_idx + rows_per_markdown, num_data_rows_total)
        current_data_chunk_as_lists = data_block_df.iloc[start_idx:end_idx].values.tolist()

        final_header_for_chunk = header_rows_as_lists
        final_data_for_chunk = current_data_chunk_as_lists

        # If no real header is attached and the current data block is not empty, the first row of data is used as the “pseudo header” to generate the delimiter
        if not append_header and current_data_chunk_as_lists:
            final_header_for_chunk = [current_data_chunk_as_lists[0]]
            final_data_for_chunk = current_data_chunk_as_lists[1:]

        markdown_content = generate_markdown_table_string(
            final_header_for_chunk, final_data_for_chunk, num_columns
        )

        # BUG FIX: Use zfill for proper 2-digit sheet and 3-digit file padding.
        file_name = f"{str(sheet_index).zfill(2)}{str(i).zfill(3)}.md"
        file_path = os.path.join(output_dir, file_name)

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(markdown_content)
            logger.debug(
                f"  Sudah disimpan'{file_path}' (incl.  {len(current_data_chunk_as_lists)} Row Raw Data)"
            )
        except Exception as e:
            logger.debug(f"  Save file '{file_path}' Error during: {e}")


def is_list_of_lists_empty(data_list):
    """
    Determine if a 2D list is empty or contains only empty values (None, '')。
    """
    if not data_list:
        return True
    # Use any() and generator expressions for efficient judgment
    # any(row) Check for non-empty lines
    # any(cell is not None and cell != '' for cell in row) Check if there are non-empty cells in the row
    return not any(any(cell is not None and str(cell).strip() != '' for cell in row) for row in data_list)


def excel_file_to_markdown(
        excel_path, num_header_rows, rows_per_markdown, output_dir, append_header=True
):
    logger.debug(f"\nStart ProcessingExcelDocumentation:'{excel_path}'")
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    except Exception as e:
        logger.debug(f"Error: Unable to loadExcelDoc. '{excel_path}'Reason: {e}")
        return

    sheet_index = 0
    for sheet_name in workbook.sheetnames:
        logger.debug(f"\n  (In work)ExcelWorksheet'{sheet_name}'...")
        sheet_obj = workbook[sheet_name]
        unmerged_data_list_of_lists = unmerge_and_read_sheet(sheet_obj)
        logger.debug(f"\n  <read all data>Excel<UNK>'{sheet_name}'...{len(unmerged_data_list_of_lists)}")

        # Using the new decision function
        if is_list_of_lists_empty(unmerged_data_list_of_lists):
            logger.debug(f"  Worksheet '{sheet_name}' Empty or no valid data, skipping.")
            continue

        df = pd.DataFrame(unmerged_data_list_of_lists)
        df.fillna("", inplace=True)
        if df.empty:
            logger.debug(f"  Worksheet '{sheet_name}' Empty after processingDataFrameSkip")
            continue

        process_dataframe_to_markdown_files(
            df,
            str(sheet_index),
            num_header_rows,
            rows_per_markdown,
            output_dir,
            append_header=append_header,
        )
        sheet_index += 1

    if workbook:
        workbook.close()
    logger.debug(f"\nExcelDoc. '{excel_path}' Process Completed.")


def csv_file_to_markdown(
        csv_path,
        num_header_rows,
        rows_per_markdown,
        output_dir,
        csv_encoding="utf-8",
        csv_delimiter=",",
        append_header=True,
):
    logger.debug(f"\nStart ProcessingCSVDocumentation:'{csv_path}'")
    try:
        df = pd.read_csv(
            csv_path,
            header=None,
            dtype=str,
            encoding=csv_encoding,
            sep=csv_delimiter,
            keep_default_na=False,
        )
        df.fillna("", inplace=True)

    except pd.errors.EmptyDataError:
        logger.debug(f"Error: CSVDoc. '{csv_path}' Empty")
        return
    except FileNotFoundError:
        logger.debug(f"Error: CSVDoc. '{csv_path}' Nothing found.")
        return
    except Exception as e:
        logger.debug(f"Error: UnreadableCSVDoc. '{csv_path}'Reason: {e}")
        return

    if df.empty:
        logger.debug(f"CSVDoc. '{csv_path}' Empty or empty after processing, skipping.")
        return

    process_dataframe_to_markdown_files(
        df,
        "0",
        num_header_rows,
        rows_per_markdown,
        output_dir,
        append_header,
    )
    logger.debug(f"\nCSVDoc. '{csv_path}' Process Completed.")


def convert_file_to_markdown(
        input_file_path,
        num_header_rows,
        rows_per_markdown,
        base_output_dir="output_markdown_files",
        csv_encoding="utf-8",
        csv_delimiter=",",
        append_header=True,
):
    """
    will be Excel OR CSV Convert files to multiple Markdown files.
    """
    if not os.path.exists(input_file_path):
        logger.debug(f"Error: Input file '{input_file_path}' Nothing found.")
        return

    if not os.path.exists(base_output_dir):
        os.makedirs(base_output_dir)
        logger.debug(f"To create an output directory:'{base_output_dir}'")

    _, file_extension = os.path.splitext(input_file_path)
    file_extension = file_extension.lower()
    if file_extension == ".xls":
        input_file_path = xls_to_xlsx(input_file_path)

    if file_extension in [".xlsx", ".xls"]:
        excel_file_to_markdown(
            input_file_path,
            num_header_rows,
            rows_per_markdown,
            base_output_dir,
            append_header,
        )
    elif file_extension == ".csv":
        csv_file_to_markdown(
            input_file_path,
            num_header_rows,
            rows_per_markdown,
            base_output_dir,
            csv_encoding,
            csv_delimiter,
            append_header,
        )
    else:
        logger.debug(
            f"Error: Unsupported file type '{file_extension}'Please provide user. Excel (.xlsx, .xls) OR CSV (.csv) files."
        )


def handler(
        cache_dir,
        file_name: str,
        header_rows: List[int] = [0, 1],
        data_rows: int = 12,
        append_header=True,
):
    """
    The main function that handles file conversions.
    """
    doc_id = uuid4()
    md_file_name = f"{cache_dir}/{doc_id}"

    convert_file_to_markdown(
        input_file_path=file_name,
        base_output_dir=md_file_name,
        num_header_rows=header_rows,
        rows_per_markdown=data_rows,
        append_header=append_header,
    )
    return md_file_name, None, doc_id


if __name__ == "__main__":
    # Define test parameters
    test_cache_dir = "/Users/zhangguoqing/Downloads/tmp"
    test_file_name = "/Users/zhangguoqing/Downloads/124327.xlsx"
    # Test append_header=True and the index is out of bounds
    test_header_rows = [0, 0]  # start_header_index Out of Scope
    test_data_rows = 2
    test_append_header = True

    # Recall handler Function
    print("--- Test Scenarios: append_header=True, Table header index out of bounds ---")
    handler(
        cache_dir=test_cache_dir,
        file_name=test_file_name,
        header_rows=test_header_rows,
        data_rows=test_data_rows,
        append_header=test_append_header,
    )
