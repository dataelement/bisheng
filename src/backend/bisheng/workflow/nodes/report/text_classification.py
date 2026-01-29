# Convert the string to the format required by the docx replacement tool.
import os
import re
import tempfile
from dataclasses import dataclass
from enum import Enum
from typing import List, Dict, Tuple, Any
from urllib.parse import urlparse, unquote
from uuid import uuid4

import pandas as pd
import requests
from charset_normalizer import detect
from docx.enum.text import WD_ALIGN_PARAGRAPH
from loguru import logger
from openpyxl import load_workbook

from bisheng.core.storage.minio.minio_manager import get_minio_storage_sync
from bisheng.utils import md5_hash


class ResourceType(Enum):
    """Resource Type Enumeration"""
    IMAGE = "image"
    TABLE = "table"
    TEXT = "text"
    HEADING = "heading"
    BOLD_TEXT = "bold_text"


@dataclass
class MatchPattern:
    """Match Pattern Definition"""

    name: str
    resource_type: ResourceType
    pattern: str
    flags: int
    handler_method: str


# priority by order in the list
Patterns = [
    MatchPattern(
        name="markdown_table",
        resource_type=ResourceType.TABLE,
        pattern=r"(\s*\|[^\r\n]*\|[^\r\n]*(?:\r?\n\s*\|[^\r\n]*\|[^\r\n]*)+)",
        flags=re.MULTILINE,
        handler_method="_handle_markdown_table",
    ),
    MatchPattern(
        name="http_excel_csv",
        resource_type=ResourceType.TABLE,
        pattern=r"(https?://[^\s\u4e00-\u9fff]*\.(?:xlsx?|csv)(?:\?[^\s\u4e00-\u9fff]*)?)",
        flags=re.IGNORECASE,
        handler_method="_handle_http_excel_csv",
    ),
    MatchPattern(
        name="local_excel_csv",
        resource_type=ResourceType.TABLE,
        pattern=r"([^\s]*[/\\][^\s]*\.(?:xlsx?|csv))",
        flags=re.IGNORECASE,
        handler_method="_handle_http_excel_csv",
    ),
    MatchPattern(
        name="markdown_image",
        resource_type=ResourceType.IMAGE,
        pattern=r"!\[([^\]]*)\]\(([^)]+\.(?:png|jpg|jpeg|bmp|gif|webp)(?:\?[^)]*)?)\)",
        flags=re.IGNORECASE,
        handler_method="_handle_http_image",
    ),
    MatchPattern(
        name="http_image",
        resource_type=ResourceType.IMAGE,
        pattern=r"""
        (?:                         # 非捕获组开始
            (?:                     # URL协议
                https?://            # http或https
                |                    # 或
                //                   # 协议相对URL
                |                    # 或
                file://              # 文件协议
            )?                      # 协议可选
        )?                          # 整个协议部分可选

        (?:                         # 域名或路径
            (?:                     # 域名部分
                (?:www\.)?          # 可选的www
                [a-z0-9][a-z0-9-]* # 域名主体
                (?:\.[a-z0-9]+)+    # 域名后缀
            )                       # 或
            |                       # 或
            (?:                     # 本地路径
                (?:[a-z]:)?         # Windows驱动器(可选)
                [\\/]?              # 可选的路径分隔符
                (?:[^<>()\[\]"'\s]+[\\/])*  # 路径部分
            )
        )

        [^<>()\[\]"\'\s]*?          # 文件名(不含扩展名)

        \.                          # 扩展名前的点
        (?:png|jpg|jpeg|gif|bmp|webp|tif)  # 图片扩展名

        (?:                         # 可选的查询字符串
            \?                      # 问号
            [^<>()\[\]"'\s]*        # 查询参数
        )?                          # 查询字符串可选

        (?:                         # 可选的片段标识符
            \#                      # 井号
            [^<>()\[\]"'\s]*        # 片段
        )?                          # 片段可选
    """,
        flags=re.IGNORECASE | re.VERBOSE,
        handler_method="_handle_http_image",
    ),
    MatchPattern(
        name="heading",
        resource_type=ResourceType.HEADING,
        pattern=r'^(#{1,6})\s+(.+?)(?:\s+#+)?\s*$',
        flags=re.IGNORECASE | re.MULTILINE,
        handler_method="_handle_heading",
    ),
    MatchPattern(
        name="bold_text",
        resource_type=ResourceType.BOLD_TEXT,
        # 严格正则表达式
        # (?<!\\)          - 前向否定断言，确保前面没有转义符\
        # (?:\\*\\*|__)   - 匹配**或__
        # (?![\s])        - 后向否定断言，确保后面不是空白字符
        # (.+?)           - 非贪婪匹配内容
        # (?<![\s])       - 前向否定断言，确保前面不是空白字符
        # (?:\\*\\*|__)   - 匹配**或__
        pattern=r'(?<!\\)(?:\*\*|__)(?![\s])(.+?)(?<![\s])(?<!\\)(?:\*\*|__)',
        flags=re.IGNORECASE,
        handler_method="_handle_bold_text",
    )
]


class TextClassificationReport:
    def __init__(self, static_resource_path: str):
        # Ensure the static resource path exists, save some downloaded files in there
        self.static_resource_path = static_resource_path or tempfile.gettempdir()
        assert os.path.exists(self.static_resource_path), f"{self.static_resource_path} dir not exist."
        self.file_cache = {}

    def get_all_classified_data(self, string_data: str) -> List[Dict]:
        string_data = [{"type": "text", "content": string_data}]
        for pattern in Patterns:
            handle_method = pattern.handler_method
            handler = getattr(self, handle_method, None)
            if not handler:
                logger.warning(f"Handler method {handle_method} not found.")
                continue
            string_data = handler(pattern, string_data)
        return string_data

    def _download_file(self, file_url: str) -> str:
        # Set request headers to simulate browser access
        file_url_md5 = md5_hash(file_url)
        if self.file_cache.get(file_url_md5):
            return self.file_cache.get(file_url_md5)
        url_obj = urlparse(file_url)
        filename = unquote(url_obj.path.split('/')[-1])
        if file_url.startswith(("https", "http")):
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }

            response = requests.get(file_url, headers=headers, timeout=30, verify=False)
            response.raise_for_status()
            # Get filename
            # accordingContent-TypeInferred extension
            content_type = response.headers.get("Content-Type", "").lower()
            if "image/png" in content_type:
                filename = f"{uuid4().hex}.png"
            elif "image/jpeg" in content_type or "image/jpg" in content_type:
                filename = f"{uuid4().hex}.jpg"
            elif "image/bmp" in content_type:
                filename = f"{uuid4().hex}.bmp"
            else:
                content_disposition = response.headers.get('Content-Disposition')
                if content_disposition:
                    filename = unquote(content_disposition).split('filename=')[-1].strip("\"'")
                file_ext = os.path.splitext(filename)[1].lower()
                filename = f"{uuid4().hex}{file_ext}"
            file_content = response.content
        else:
            # maybe minio file path
            minio_client = get_minio_storage_sync()
            if not file_url.lstrip("/").startswith((minio_client.bucket, minio_client.tmp_bucket)):
                raise ValueError(f"Invalid minio path {file_url}")
            path_parts = url_obj.path.lstrip("/").split('/', 1)
            if len(path_parts) != 2:
                raise ValueError(f"Invalid minio path {file_url}")

            bucket_name, object_name = path_parts
            # Call Synchronized minio Method download
            object_name = unquote(object_name)
            file_content = minio_client.get_object_sync(bucket_name, object_name)

        # Creating temp file
        temp_file = os.path.join(self.static_resource_path, filename)
        with open(temp_file, "wb") as f:
            f.write(file_content)
        self.file_cache[file_url_md5] = temp_file
        return temp_file

    def _handle_bold_text(self, pattern: MatchPattern, string_data: List[Dict]):
        return self._base_handle_string_data(pattern, string_data, self._parse_bold_text_data)

    def _parse_bold_text_data(self, bold_text_str: str) -> Dict:
        # Remove the ** or __ markers
        content = re.sub(r'(?<!\\)(\*\*|__)(?![\s])', '', bold_text_str)
        return {"type": "text", "content": content, "bold": True}

    def _handle_heading(self, pattern: MatchPattern, string_data: List[Dict]):
        return self._base_handle_string_data(pattern, string_data, self._parse_heading_data)

    @staticmethod
    def _parse_heading_data(heading_str: str) -> Dict:
        heading_level, heading_text = heading_str.split(" ", 1)
        level = min(len(heading_level.strip()), 6)
        return {"type": "heading", "content": heading_text, "level": level}

    def _handle_http_image(self, pattern: MatchPattern, string_data: List[Dict]):
        # Placeholder for actual implementation
        return self._base_handle_string_data(pattern, string_data, self._parse_http_image_data)

    def _parse_http_image_data(self, image_str: str) -> Dict:
        # image_str: ![xxx](http://example.com/image.png)  or http://example.com/image.png or /local/path/image.png
        try:
            if image_str.startswith("!["):
                new_image_str = image_str.split("](", 1)[1].rstrip(")")
                local_file_path = self._download_file(new_image_str)
            else:
                local_file_path = self._download_file(image_str)
            return {
                "type": "image",
                "content": local_file_path
            }
        except Exception as e:
            logger.error(f"Failed to parse http image data: {str(e)}")
            return {"type": "text", "content": image_str}

    def _handle_http_excel_csv(self, pattern: MatchPattern, string_data: List[Dict]):
        # Placeholder for actual implementation
        return self._base_handle_string_data(pattern, string_data, self._parse_http_table_data)

    def _parse_http_table_data(self, table_url: str) -> Dict:
        # Placeholder for actual implementation
        try:
            if table_url.startswith(("http", "https")):
                local_file_path = self._download_file(table_url)
            else:
                local_file_path = table_url
            if local_file_path.endswith(".xlsx"):
                table_data, table_alignment = self._parse_excel_data(local_file_path)
            elif local_file_path.endswith(".csv"):
                table_data, table_alignment = self._parse_csv_data(local_file_path)
            else:
                raise ValueError(f"Unsupported file format for table parsing: {local_file_path}")
            return self._handle_table_data_to_dict(table_data, table_alignment)
        except Exception as e:
            logger.error(f"Failed to parse http table data: {str(e)}")
            return {"type": "text", "content": table_url}

    @staticmethod
    def _convert_docx_alignment(alignment: str) -> Any:
        if alignment == "left":
            return WD_ALIGN_PARAGRAPH.LEFT
        elif alignment == "center":
            return WD_ALIGN_PARAGRAPH.CENTER
        elif alignment == "right":
            return WD_ALIGN_PARAGRAPH.RIGHT
        else:
            return WD_ALIGN_PARAGRAPH.LEFT

    def _handle_table_data_to_dict(self, table_data: List[List[str]], table_alignment: List[str]) -> Dict:
        # Convert table data and alignment into a dictionary format
        content = []
        for row in table_data:
            one_row_data = []
            for index, cell in enumerate(row):
                # convert each cell to text type with image alignment
                markdown_image_pattern = Patterns[3]  # markdown_image pattern
                http_image_pattern = Patterns[4]  # http_image pattern
                bold_text_pattern = Patterns[6]  # bold_text pattern
                cell_data = getattr(self, markdown_image_pattern.handler_method)(markdown_image_pattern,
                                                                                 [{"type": "text", "content": cell}])
                cell_data = getattr(self, http_image_pattern.handler_method)(http_image_pattern, cell_data)
                cell_data = getattr(self, bold_text_pattern.handler_method)(bold_text_pattern, cell_data)
                cell_data[0]["alignment"] = self._convert_docx_alignment(
                    table_alignment[index] if index < len(table_alignment) else "left")
                one_row_data.append(cell_data)
            content.append(one_row_data)
        return {"type": "table", "content": content}

    @staticmethod
    def _parse_excel_data(file_path: str) -> Tuple[List[List[str]], List[str]]:
        # Use openpyxl read docx
        workbook = load_workbook(file_path, data_only=True)  # data_only=TrueGet Calculated Value

        # Using the first worksheet
        worksheet = workbook.active

        table_data = []
        max_col = 0

        # Read all rows
        for row in worksheet.iter_rows(values_only=True):
            # Skip completely blank rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_data = [str(cell) if cell is not None else "" for cell in row]
            table_data.append(row_data)
            max_col = max(max_col, len(row_data))

        # Make sure all rows have the same number of columns
        for row in table_data:
            while len(row) < max_col:
                row.append("")

        # Generate alignment information (default left alignment)
        alignments = ["left"] * max_col

        logger.debug(f"ExcelFile parsed successfully: {len(table_data)}Parade x {max_col}column")

        return table_data, alignments

    @staticmethod
    def _parse_csv_data(file_path: str) -> Tuple[List[List[str]], List[str]]:
        # Auto-Detect Encoding
        with open(file_path, 'rb') as f:
            raw_data = f.read()
            encoding_info = detect(raw_data)
            encoding = encoding_info['encoding'] or 'utf-8'

        # Use pandas read outCSVto better handle various formats
        df = pd.read_csv(file_path, encoding=encoding)

        # Convert to Tabular Data Format
        table_data = []

        # Add header
        headers = [str(col) for col in df.columns]
        table_data.append(headers)

        # Add Data Row
        for _, row in df.iterrows():
            row_data = [str(cell) if pd.notna(cell) else "" for cell in row]
            table_data.append(row_data)

        # Generate alignment information (default left alignment)
        alignments = ["left"] * len(headers)

        logger.debug(f"CSVFile parsed successfully: {len(table_data)}Parade x {len(headers)}column")
        return table_data, alignments

    def _handle_markdown_table(self, pattern: MatchPattern, string_data: List[Dict]):
        return self._base_handle_string_data(pattern, string_data, self._parse_table_data)

    @staticmethod
    def _base_handle_string_data(pattern: MatchPattern, string_data: List[Dict], parse_method) -> List[Dict]:
        new_string_data = []
        for one in string_data:
            if one["type"] != "text" or not one["content"]:
                new_string_data.append(one)
                continue
            content = one["content"]
            match_flag = False
            for one_match in re.finditer(pattern.pattern, one["content"], flags=pattern.flags):
                match_flag = True
                raw_text = one_match.group(0)
                before, end = content.split(raw_text, 1)
                new_string_data.append({
                    "type": "text",
                    "content": before
                })
                new_string_data.append(parse_method(raw_text))
                content = end
            if not match_flag:
                new_string_data.append(one)
                continue
            if content:
                new_string_data.append({
                    "type": "text",
                    "content": content
                })
        return new_string_data

    def _parse_table_data(self, table_content: str) -> Dict:
        """Parse the table data while working with the in the table string"""
        try:
            # Analyze the table structure first
            table_data, table_alignment = self._parse_markdown_table_from_content(table_content)
            return self._handle_table_data_to_dict(table_data, table_alignment)
        except Exception as e:
            logger.error(f"Failed to parse table data: {str(e)}")
            return {"type": "text", "content": table_content}

    def _parse_markdown_table_from_content(self, table_content: str) -> Tuple[List[List[str]], List[str]]:
        """
        Parsing from file contentsMarkdownTable Filter

        Args:
            table_content: table content in File contents

        Returns:
            tuple: (Table Data, Alignment Info List)
        """
        # Keep all rows, including empty rows - Fully preserving the original table structure
        lines = [line.strip() for line in table_content.strip().split("\n")]

        table_rows = []
        table_alignments = []
        separator_found = False

        for line in lines:
            # Skip completely empty rows, but leave empty table rows with only vertical lines
            if not line:
                continue

            # Check if it is a delimiter line
            if self._is_table_separator_line(line):
                table_alignments = self._parse_table_alignments(line)
                separator_found = True
                continue

            # Parse Data Rows - Keep all table rows, including empty rows
            cells = self._parse_table_row(line)
            # Remove if cells Condition, leave empty table rows
            table_rows.append(cells)

        # Use default alignment if no separator found
        if not separator_found and table_rows:
            table_alignments = ["left"] * len(table_rows[0])

        # Make sure all rows have the same number of columns
        if table_rows:
            max_cols = max(len(row) for row in table_rows)
            for row in table_rows:
                while len(row) < max_cols:
                    row.append("")

            # Make sure the number of aligned messages matches the number of columns
            while len(table_alignments) < max_cols:
                table_alignments.append("left")
            table_alignments = table_alignments[:max_cols]

        logger.debug(
            f"Successfully parsed table in content, size: {len(table_rows)}Parade x {len(table_alignments)}column")
        return table_rows, table_alignments

    @staticmethod
    def _is_table_separator_line(line: str) -> bool:
        """Check if yesMarkdownTable Separator Row"""
        content = line.strip().strip("|").strip()
        if not content:
            return False

        cells = [cell.strip() for cell in content.split("|")]

        # At least one cell must contain a delimiter character (-OR:）
        has_separator_chars = False
        for cell in cells:
            if not cell:
                continue
            # Check for delimiter characters
            if '-' in cell or ':' in cell:
                has_separator_chars = True
            # Check for anything else after removing the separator character
            clean_cell = cell.replace("-", "").replace(":", "").strip()
            if clean_cell:
                return False

        # Only lines containing delimiter characters can be considered delimiter lines
        return has_separator_chars

    @staticmethod
    def _parse_table_alignments(separator_line: str) -> list:
        """Parse column alignment from delimiter rows"""
        alignments = []
        content = separator_line.strip().strip("|").strip()
        cells = [cell.strip() for cell in content.split("|")]
        for cell in cells:
            if not cell:
                alignments.append("left")
                continue
            if cell.startswith(":") and cell.endswith(":"):
                alignments.append("center")
            elif cell.endswith(":"):
                alignments.append("right")
            else:
                alignments.append("left")

        return alignments

    @staticmethod
    def _parse_table_row(line: str) -> list:
        """Parse Table Rows"""
        content = line.strip()
        if content.startswith("|"):
            content = content[1:]
        if content.endswith("|"):
            content = content[:-1]

        cells = []
        current_cell = ""
        escaped = False

        for char in content:
            if escaped:
                current_cell += char
                escaped = False
            elif char == "\\":
                escaped = True
                current_cell += char
            elif char == "|":
                cells.append(current_cell.strip())
                current_cell = ""
            else:
                current_cell += char

        # Always add the last cell to ensure empty table rows are also parsed correctly
        cells.append(current_cell.strip())
        return cells
