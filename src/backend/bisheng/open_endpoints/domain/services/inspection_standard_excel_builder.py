from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from bisheng.common.errcode.inspection_standard_sync import InspectionStandardSyncExcelBuildError
from bisheng.open_endpoints.domain.schemas.inspection_standard_sync import (
    InspectionStandardItemRecord,
    InspectionStandardRecord,
)

SHEET_STANDARD_NAME = "点检标准"
SHEET_ITEM_NAME = "标准项次 "

TEXT_NUMBER_FORMAT = "@"

CHECK_STANDARD_ZH_HEADERS = [
    "设备所属单位",
    "点检标准编号（12）*",
    "分部设备中文名称（100）*",
    "标准类别（小代码）*",
    "油脂料号（20）",
    "点检项目名称（50）*",
    "设备状态（小代码）*",
    "实施方（小代码）*",
    "安全挂牌（小代码）*",
    "实施周期*",
    "周期单位（小代码）*",
    "系统接口",
    "下次排程日期（10）*\n文本格式YYYY-MM-DD",
    "维护原因（50）*",
    "点检员岗号（10）*",
    "点检员工号（10）*",
    "点检员姓名（10）*",
]

CHECK_STANDARD_EN_COLUMNS = [
    "CREATE_DEPT_ID",
    "CHECK_STANDARD_ID",
    "DEVICE_NAME",
    "STANDARD_TYPE",
    "OIL_PART_NO",
    "CHECK_ITEM_NAME",
    "DEVICE_STATUS",
    "ENFORCE_CODE",
    "SAFETY_BOARD",
    "CHECK_PERIOD",
    "PERIOD_UNIT",
    "INTERFACE_SYSTEM",
    "NEXT_SCHE_DATE",
    "MAINTAIN_REASON",
    "DEVICE_MAINTAIN_JOB_ID",
    "REC_CREATOR",
    "REC_CREATOR_NAME",
]

CHECK_STANDARD_ITEM_ZH_HEADERS = [
    "点检标准编号（12）*",
    "点检标准项次（3）*",
    "内容（50）*",
    "点检方法（小代码）*",
    "润滑方式（小代码）*",
    "润滑点数（3）整数",
    "管理控别*\n（小代码）",
    "管理类别*\n（小代码）",
    "数据类别*\n（小代码）",
    "标准（100）*",
    "计量单位",
    "上限（5,2）*",
    "下限（5,2）*",
    "报警设置（100）",
    "法定要求\n（小代码）*",
    "装置名称",
    "润滑部位",
    "分配器\n编号",
    "入机点\n编号",
    "润滑点\n标识",
    "油嘴\n规格",
    "加油工具",
    "油脂牌号",
    "单点\n注入量",
    "合计\n注入量",
    "润滑效果\n判断基准",
    "技术专业负责人",
    "责任班组",
    "润滑负责人",
    "油脂属性\n（小代码）",
]

CHECK_STANDARD_ITEM_EN_COLUMNS = [
    "CHECK_STANDARD_ID",
    "CHECK_STANDARD_SEQ_NO",
    "CONTENT",
    "CHECK_WAY",
    "LUBRIC_WAY",
    "LUBRIC_POINT",
    "MANAGE_CONTROL_MODE",
    "MANAGE_TYPE",
    "DATA_TYPE",
    "CRITERI",
    "UOM",
    "QLTY_TOP",
    "QLTY_BOTTOM",
    "ALARM_SETTINGS",
    "STATUTORY_REQ",
    "EQUIPMENT_NAME",
    "LUBRIC_PART",
    "DISTRIBUTOR_NO",
    "ENTRY_OINT_NO",
    "LUBRIC_POINT_MARK",
    "NOZZLE_SPECIFICATION",
    "FUELING_TOOLS",
    "OIL_NO",
    "SINGLE_INJECTION_VOLUME",
    "TOTAL_INJECTION_VOLUME",
    "LUBRIC_EFFECT_JUDGE_CRITERIA",
    "TECH_MAJOR_PIC",
    "RESPONSIBILITY_TEAM",
    "LUBRIC_PIC",
    "OIL_PROPERTY",
]


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _write_sheet_header(
    sheet: Worksheet,
    *,
    title: str,
    zh_headers: list[str],
    en_headers: list[str],
) -> None:
    sheet.cell(row=1, column=1, value=title).number_format = TEXT_NUMBER_FORMAT
    for col, header in enumerate(zh_headers, start=1):
        sheet.cell(row=2, column=col, value=header).number_format = TEXT_NUMBER_FORMAT
    for col, header in enumerate(en_headers, start=1):
        sheet.cell(row=3, column=col, value=header).number_format = TEXT_NUMBER_FORMAT


def _write_rows(
    sheet: Worksheet,
    *,
    columns: list[str],
    rows: Iterable[dict[str, Any]],
    text_columns: set[str] | None = None,
    start_row: int = 4,
) -> None:
    text_columns = text_columns or set()
    for row_index, row in enumerate(rows, start=start_row):
        for col_index, column in enumerate(columns, start=1):
            value = _normalize_cell(row.get(column))
            cell = sheet.cell(row=row_index, column=col_index)
            if column in text_columns or isinstance(value, str):
                cell.value = str(value)
                cell.number_format = TEXT_NUMBER_FORMAT
            else:
                cell.value = value


def build_inspection_standard_xlsx_bytes(
    *,
    check_standards: list[InspectionStandardRecord],
    check_standard_items: list[InspectionStandardItemRecord],
) -> bytes:
    try:
        workbook = Workbook()
        standard_sheet = workbook.active
        standard_sheet.title = SHEET_STANDARD_NAME
        item_sheet = workbook.create_sheet(SHEET_ITEM_NAME)

        _write_sheet_header(
            standard_sheet,
            title="点检标准",
            zh_headers=CHECK_STANDARD_ZH_HEADERS,
            en_headers=CHECK_STANDARD_EN_COLUMNS,
        )
        _write_sheet_header(
            item_sheet,
            title="点检标准项次",
            zh_headers=CHECK_STANDARD_ITEM_ZH_HEADERS,
            en_headers=CHECK_STANDARD_ITEM_EN_COLUMNS,
        )

        standard_rows = [record.model_dump() for record in check_standards]
        item_rows = [record.model_dump() for record in check_standard_items]

        _write_rows(
            standard_sheet,
            columns=CHECK_STANDARD_EN_COLUMNS,
            rows=standard_rows,
            text_columns={"NEXT_SCHE_DATE"},
        )
        _write_rows(
            item_sheet,
            columns=CHECK_STANDARD_ITEM_EN_COLUMNS,
            rows=item_rows,
            text_columns={"CHECK_STANDARD_SEQ_NO"},
        )

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
    except Exception as exc:
        raise InspectionStandardSyncExcelBuildError(msg="failed to build inspection standard xlsx") from exc
