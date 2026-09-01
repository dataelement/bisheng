from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest

from bisheng.common.errcode.portal_course import (
    PortalCourseCatalogInUseError,
    PortalCourseCatalogNameDuplicateError,
    PortalCourseCatalogNotFoundError,
    PortalCourseCatalogParentInvalidError,
)
from bisheng.shougang_portal_course.domain.schemas.portal_course_schema import (
    CatalogCreate,
    CatalogUpdate,
    CourseCreate,
    CourseUpdate,
)
from bisheng.shougang_portal_course.domain.services.catalog_service import (
    EXCEL_HEADERS,
    PortalCourseCatalogService,
)
from bisheng.shougang_portal_course.domain.services.course_service import (
    PortalCourseService,
)


async def test_create_catalog_builds_yunxuetang_style_paths(course_session):
    service = PortalCourseCatalogService(course_session)
    root = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产", description="公司级", order_index=1),
    )
    child = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(
            name="消防安全",
            parent_id=root.id,
            description="消防专题",
            order_index=1,
        ),
    )

    assert root.routing_path == "001"
    assert root.catalog_id_path == root.id
    assert root.catalog_name_path == "安全生产"
    assert child.routing_path == "001.001"
    assert child.catalog_id_path == f"{root.id},{child.id}"
    assert child.catalog_name_path == "安全生产->消防安全"


async def test_sibling_name_must_be_unique(course_session):
    service = PortalCourseCatalogService(course_session)
    await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    with pytest.raises(PortalCourseCatalogNameDuplicateError):
        await service.create_catalog(
            tenant_id=1,
            user_id=7,
            payload=CatalogCreate(name="安全生产"),
        )


async def test_cannot_move_catalog_under_its_descendant(course_session):
    service = PortalCourseCatalogService(course_session)
    root = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    child = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="消防安全", parent_id=root.id),
    )
    with pytest.raises(PortalCourseCatalogParentInvalidError):
        await service.update_catalog(
            tenant_id=1,
            user_id=7,
            catalog_id=root.id,
            payload=CatalogUpdate(parent_id=child.id),
        )


async def test_rename_updates_descendant_name_path(course_session):
    service = PortalCourseCatalogService(course_session)
    root = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    child = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="消防安全", parent_id=root.id),
    )
    await service.update_catalog(
        tenant_id=1,
        user_id=7,
        catalog_id=root.id,
        payload=CatalogUpdate(name="安全培训"),
    )
    updated_child = await service.repository.get_catalog(
        tenant_id=1,
        catalog_id=child.id,
    )
    assert updated_child is not None
    assert updated_child.catalog_name_path == "安全培训->消防安全"


async def test_delete_catalog_rejects_children_and_assigned_courses(course_session):
    catalog_service = PortalCourseCatalogService(course_session)
    course_service = PortalCourseService(course_session)
    root = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="消防安全", parent_id=root.id),
    )
    with pytest.raises(PortalCourseCatalogInUseError):
        await catalog_service.delete_catalog(tenant_id=1, catalog_id=root.id)

    leaf = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="单独目录"),
    )
    await course_service.create_course(
        tenant_id=1,
        user_id=7,
        payload=CourseCreate(name="高炉安全", catalog_id=leaf.id),
    )
    with pytest.raises(PortalCourseCatalogInUseError):
        await catalog_service.delete_catalog(tenant_id=1, catalog_id=leaf.id)


async def test_course_update_can_clear_catalog(course_session):
    catalog_service = PortalCourseCatalogService(course_session)
    course_service = PortalCourseService(course_session)
    catalog = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    course = await course_service.create_course(
        tenant_id=1,
        user_id=7,
        payload=CourseCreate(name="高炉安全", catalog_id=catalog.id),
    )
    assert course.catalog_id == catalog.id

    updated = await course_service.update_course(
        tenant_id=1,
        course_id=course.id,
        payload=CourseUpdate.model_validate({"catalog_id": None}),
    )
    assert updated.catalog_id is None


async def test_course_rejects_missing_catalog(course_session):
    course_service = PortalCourseService(course_session)
    with pytest.raises(PortalCourseCatalogNotFoundError):
        await course_service.create_course(
            tenant_id=1,
            user_id=7,
            payload=CourseCreate(name="高炉安全", catalog_id="a" * 32),
        )


async def test_public_tree_hides_children_of_closed_parent(course_session):
    service = PortalCourseCatalogService(course_session)
    root = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产", opened=False),
    )
    await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="消防安全", parent_id=root.id, opened=True),
    )
    items = await service.list_read_models(tenant_id=1, public_only=True, as_tree=True)
    assert items == []
    with pytest.raises(PortalCourseCatalogNotFoundError):
        await service.resolve_catalog_ids(
            tenant_id=1,
            catalog_id=root.id,
            include_descendants=True,
            public_only=True,
        )
    ids = await service.resolve_catalog_ids(
        tenant_id=1,
        catalog_id=root.id,
        include_descendants=True,
        public_only=False,
    )
    assert ids[0] == root.id
    assert len(ids) == 2


async def test_course_list_nav_includes_closed_catalogs_with_enabled_courses(course_session):
    catalog_service = PortalCourseCatalogService(course_session)
    course_service = PortalCourseService(course_session)
    hidden = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="内部资料", opened=False),
    )
    await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="空目录", opened=False),
    )
    opened = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产", opened=True),
    )
    hidden_course = await course_service.create_course(
        tenant_id=1,
        user_id=7,
        payload=CourseCreate(name="高炉安全", catalog_id=hidden.id),
    )
    hidden_course.enabled = True
    await course_service.repository.add(hidden_course)
    await course_service.create_course(
        tenant_id=1,
        user_id=7,
        payload=CourseCreate(name="草稿课", catalog_id=opened.id),
    )

    tree = await catalog_service.list_course_list_nav(tenant_id=1)
    names = {item.name for item in tree}
    assert names == {"内部资料", "安全生产"}
    by_name = {item.name: item for item in tree}
    assert by_name["内部资料"].course_count == 1
    assert by_name["安全生产"].course_count == 0


async def test_course_list_nav_keyword_keeps_matching_ancestors(course_session):
    service = PortalCourseCatalogService(course_session)
    root = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    child = await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="消防安全", parent_id=root.id),
    )
    await service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="设备操作"),
    )

    tree = await service.list_course_list_nav(tenant_id=1, keyword="消防")
    assert [item.name for item in tree] == ["安全生产"]
    assert tree[0].children is not None
    assert [item.name for item in tree[0].children] == ["消防安全"]
    assert child.catalog_name_path == "安全生产->消防安全"


def _catalog_workbook(*rows: tuple) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        sheet.cell(row=1, column=col, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def test_excel_import_empty_parent_goes_to_root_and_upserts(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            ("", "", "", "安全生产", "公司级", 1, "是"),
            ("", "", "", "消防安全", "消防专题", 2, "是"),
        ),
    )
    assert result.failed == 0
    assert result.success == 2

    tree = await service.list_read_models(tenant_id=1, public_only=False, as_tree=True)
    assert [item.name for item in tree] == ["安全生产", "消防安全"]
    assert tree[0].description == "公司级"
    assert tree[1].description == "消防专题"

    updated = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(("", "", "", "安全生产", "更新后的简介", 3, "否")),
    )
    assert updated.failed == 0
    tree = await service.list_read_models(tenant_id=1, public_only=False, as_tree=True)
    by_name = {item.name: item for item in tree}
    assert by_name["安全生产"].description == "更新后的简介"
    assert by_name["安全生产"].order_index == 3
    assert by_name["安全生产"].opened is False


async def test_excel_import_uses_parent_name_not_path(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            ("", "", "安全生产", "消防安全", "消防专题", 1, "是"),
            ("", "", "", "安全生产", "公司级", 1, "是"),
        ),
    )
    assert result.failed == 0
    assert result.success == 2
    tree = await service.list_read_models(tenant_id=1, public_only=False, as_tree=True)
    assert len(tree) == 1
    assert tree[0].name == "安全生产"
    assert tree[0].children is not None
    assert tree[0].children[0].name == "消防安全"
    assert tree[0].children[0].description == "消防专题"


async def test_excel_import_missing_parent_fails_row(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(("", "", "不存在", "消防安全", "消防专题", 1, "是")),
    )
    assert result.success == 0
    assert result.failed == 1
    assert any("上级目录「不存在」不存在" in item for item in result.errors)


async def test_excel_preview_flags_missing_parent_as_recoverable(course_session):
    service = PortalCourseCatalogService(course_session)
    preview = await service.preview_excel(
        tenant_id=1,
        content=_catalog_workbook(
            ("", "", "不存在", "消防安全", "消防专题", 1, "是"),
            ("", "", "", "安全生产", "公司级", 1, "是"),
            ("", "", "安全生产", "应急管理", "应急", 2, "是"),
        ),
    )
    assert preview.total == 3
    assert preview.valid == 2
    assert len(preview.issues) == 1
    assert preview.issues[0].code == "missing_parent"
    assert preview.issues[0].recoverable is True
    assert preview.issues[0].row == 2


async def test_excel_force_import_puts_missing_parent_under_root(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        force=True,
        content=_catalog_workbook(
            ("", "", "不存在", "消防安全", "消防专题", 1, "是"),
            ("", "", "消防安全", "灭火器使用", "实操", 1, "是"),
        ),
    )
    assert result.failed == 0
    assert result.success == 2
    tree = await service.list_read_models(tenant_id=1, public_only=False, as_tree=True)
    assert [item.name for item in tree] == ["消防安全"]
    assert tree[0].children is not None
    assert tree[0].children[0].name == "灭火器使用"


async def test_template_headers_match_import_format():
    service = PortalCourseCatalogService(session=None)
    workbook = openpyxl.load_workbook(BytesIO(service.build_template()))
    sheet = workbook["课程目录"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == EXCEL_HEADERS
    assert headers[0] == "目录ID"
    assert headers[1] == "上级目录ID"
    assert "目录路径" not in headers


async def test_excel_import_uses_catalog_id_to_create_and_update(course_session):
    service = PortalCourseCatalogService(course_session)
    catalog_id = "CAT-1001"
    created = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            (catalog_id, "", "", "安全生产", "公司级", 1, "是"),
        ),
    )
    assert created.failed == 0
    assert created.success == 1
    item = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id=catalog_id,
    )
    assert item is not None
    assert item.id != catalog_id
    assert len(item.id) == 32
    assert item.external_id == catalog_id
    assert item.name == "安全生产"

    updated = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            (catalog_id, "", "", "安全培训", "更新简介", 4, "否"),
        ),
    )
    assert updated.failed == 0
    item = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id=catalog_id,
    )
    assert item is not None
    assert item.name == "安全培训"
    assert item.description == "更新简介"
    assert item.order_index == 4
    assert item.opened is False


async def test_excel_import_uses_parent_id_over_parent_name(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            ("CAT-ROOT", "", "", "安全生产", "公司级", 1, "是"),
            ("CAT-OTHER", "", "", "设备操作", "其他根目录", 2, "是"),
            ("CAT-CHILD", "CAT-ROOT", "设备操作", "消防安全", "消防专题", 1, "是"),
        ),
    )
    assert result.failed == 0
    child = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id="CAT-CHILD",
    )
    root = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id="CAT-ROOT",
    )
    assert child is not None
    assert root is not None
    assert child.parent_id == root.id
    assert child.catalog_name_path == "安全生产->消防安全"


async def test_excel_import_parent_id_from_later_row(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            ("CAT-CHILD", "CAT-ROOT", "", "消防安全", "消防专题", 1, "是"),
            ("CAT-ROOT", "", "", "安全生产", "公司级", 1, "是"),
        ),
    )
    assert result.failed == 0
    assert result.success == 2
    child = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id="CAT-CHILD",
    )
    root = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id="CAT-ROOT",
    )
    assert child is not None
    assert root is not None
    assert child.parent_id == root.id


async def test_excel_import_missing_parent_id_fails_row(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook(
            ("", "MISSING-ID", "", "消防安全", "消防专题", 1, "是"),
        ),
    )
    assert result.success == 0
    assert result.failed == 1
    assert any("上级目录ID「MISSING-ID」不存在" in item for item in result.errors)


async def test_excel_preview_flags_missing_parent_id_as_recoverable(course_session):
    service = PortalCourseCatalogService(course_session)
    preview = await service.preview_excel(
        tenant_id=1,
        content=_catalog_workbook(
            ("", "MISSING-ID", "", "消防安全", "消防专题", 1, "是"),
        ),
    )
    assert preview.total == 1
    assert preview.valid == 0
    assert preview.issues[0].code == "missing_parent"
    assert preview.issues[0].recoverable is True
    assert "上级目录ID「MISSING-ID」不存在" in preview.issues[0].message


async def test_excel_force_import_puts_missing_parent_id_under_root(course_session):
    service = PortalCourseCatalogService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        force=True,
        content=_catalog_workbook(
            ("CAT-CHILD", "MISSING-ID", "", "消防安全", "消防专题", 1, "是"),
        ),
    )
    assert result.failed == 0
    child = await service.repository.get_catalog_by_external_id(
        tenant_id=1,
        external_id="CAT-CHILD",
    )
    assert child is not None
    assert child.parent_id is None


async def test_excel_import_rejects_self_as_parent():
    service = PortalCourseCatalogService(session=None)
    with pytest.raises(ValueError, match="上级目录ID不能与目录ID相同"):
        service._parse_import_row(2, ("CAT-001", "CAT-001", "", "安全生产", "", 1, "是"))
