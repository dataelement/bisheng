"""Dictionary domain service - 系统字典业务逻辑层"""

import re
from io import BytesIO

import openpyxl
import pandas as pd
from fastapi import UploadFile
from loguru import logger

from bisheng.common.dependencies.user_deps import UserPayload
from bisheng.common.errcode.dictionary import (
    DictionaryDuplicateError,
    DictionaryExportEmptyError,
    DictionaryImportFileEmptyError,
    DictionaryImportFormatError,
    DictionaryImportHeaderError,
    DictionaryImportParseError,
    DictionaryImportTypeInvalidError,
    DictionaryKeyFormatError,
    DictionaryKeyInUseError,
    DictionaryNotFoundError,
    DictionaryPermissionDeniedError,
)
from bisheng.common.schemas.api import PageData
from bisheng.dictionary.domain.models.system_dictionary import SystemDictionary
from bisheng.dictionary.domain.repositories.interfaces.system_dictionary_repository import (
    SystemDictionaryRepository,
)
from bisheng.dictionary.domain.schemas.dictionary_schema import (
    DICT_KEY_PATTERN,
    DICTIONARY_EXPORT_HEADERS,
    DICTIONARY_TYPE_CODE_TO_LABEL,
    DICTIONARY_TYPE_LABEL_TO_CODE,
    DictionaryCreateRequest,
    DictionaryImportResult,
    DictionaryResponse,
    DictionaryTypeEnum,
    DictionaryTypeResponse,
    DictionaryUpdateRequest,
)
from bisheng.dictionary.infrastructure.qa_expert_reference_adapter import is_dict_key_in_use


class DictionaryService:
    """系统字典业务服务"""

    def __init__(self, repository: SystemDictionaryRepository):
        self.repository = repository

    @staticmethod
    def _ensure_admin(user: UserPayload) -> None:
        """校验当前用户是否具备运营写资格(超管或平台管理员)."""
        from bisheng.user.domain.services.platform_operator import can_platform_operate

        if not can_platform_operate(user):
            raise DictionaryPermissionDeniedError()

    def _validate_dict_key_format(self, dict_key: str) -> None:
        """校验 dict_key 格式:仅允许字母、数字、下划线,且以字母或数字开头"""
        if not dict_key or not re.match(DICT_KEY_PATTERN, dict_key):
            raise DictionaryKeyFormatError()

    async def _ensure_dict_key_not_in_use(self, dict_type: str, dict_key: str) -> None:
        """校验 dict_key 未被 qa_expert 相关字段使用"""
        if await is_dict_key_in_use(dict_type, dict_key):
            raise DictionaryKeyInUseError()

    async def create(
        self,
        request: DictionaryCreateRequest,
        user: UserPayload,
    ) -> DictionaryResponse:
        """创建字典条目(管理员)"""
        self._ensure_admin(user)

        self._validate_dict_key_format(request.dict_key)

        existing = await self.repository.find_by_type_and_key(request.type, request.dict_key)
        if existing:
            raise DictionaryDuplicateError()

        entity = SystemDictionary(
            type=request.type,
            dict_key=request.dict_key,
            dict_value=request.dict_value,
            sort_order=request.sort_order,
            is_enabled=request.is_enabled,
        )
        saved = await self.repository.save(entity)
        return DictionaryResponse.model_validate(saved)

    async def update(
        self,
        dictionary_id: int,
        request: DictionaryUpdateRequest,
        user: UserPayload,
    ) -> DictionaryResponse:
        """更新字典条目(管理员)"""
        self._ensure_admin(user)

        entity = await self.repository.find_by_id(dictionary_id)
        if not entity:
            raise DictionaryNotFoundError()

        if request.dict_key is not None:
            self._validate_dict_key_format(request.dict_key)
            # 若 dict_key 发生变化,需确保原 dict_key 未被 qa_expert 使用,
            # 且新的 dict_key 在当前类型下不重复
            if entity.dict_key != request.dict_key:
                await self._ensure_dict_key_not_in_use(entity.type, entity.dict_key)
                existing = await self.repository.find_by_type_and_key(entity.type, request.dict_key)
                if existing and existing.id != dictionary_id:
                    raise DictionaryDuplicateError()
            entity.dict_key = request.dict_key

        if request.dict_value is not None:
            entity.dict_value = request.dict_value

        if request.sort_order is not None:
            entity.sort_order = request.sort_order

        if request.is_enabled is not None:
            entity.is_enabled = request.is_enabled

        updated = await self.repository.update(entity)
        return DictionaryResponse.model_validate(updated)

    async def delete(
        self,
        dictionary_id: int,
        user: UserPayload,
    ) -> bool:
        """删除字典条目(管理员)"""
        self._ensure_admin(user)

        entity = await self.repository.find_by_id(dictionary_id)
        if not entity:
            raise DictionaryNotFoundError()

        await self._ensure_dict_key_not_in_use(entity.type, entity.dict_key)

        return await self.repository.delete(dictionary_id)

    async def get_by_id(self, dictionary_id: int) -> DictionaryResponse:
        """根据 ID 查询字典条目"""
        entity = await self.repository.find_by_id(dictionary_id)
        if not entity:
            raise DictionaryNotFoundError()
        return DictionaryResponse.model_validate(entity)

    async def find_by_type_and_key(self, dict_type: str, dict_key: str) -> DictionaryResponse:
        """根据 dict_type 和 dict_key 查询启用的字典条目"""
        entity = await self.repository.find_by_type_and_key(dict_type, dict_key)
        if not entity:
            raise DictionaryNotFoundError()
        return DictionaryResponse.model_validate(entity)

    async def get_list_by_type(self, dict_type: str, page: int = 1, page_size: int = 20) -> list[DictionaryResponse]:
        """根据 dict_type 查询启用的字典条目列表"""
        entities = await self.repository.find_by_type(dict_type, page, page_size)
        if not entities:
            raise DictionaryNotFoundError()
        return [DictionaryResponse.model_validate(entity) for entity in entities]

    async def list_by_type(self) -> list[DictionaryTypeResponse]:
        """查询所有字典类型"""
        return [DictionaryTypeResponse(name=member.value, type=member.name.lower()) for member in DictionaryTypeEnum]

    async def get_next_sort_order(self, dict_type: str) -> int:
        """获取指定类型下推荐的下一个 sort_order(最大 sort_order + 1)"""
        max_order = await self.repository.get_max_sort_order_by_type(dict_type)
        return max_order + 1

    async def list_page(
        self,
        dict_type: str | None = None,
        keyword: str | None = None,
        page: int = 1,
        page_size: int = 20,
        sort_by: bool | None = None,
        is_enabled: bool | None = None,
    ) -> PageData[DictionaryResponse]:
        """分页查询字典条目"""
        entities, total = await self.repository.find_page(
            dict_type=dict_type,
            keyword=keyword,
            page=page,
            page_size=page_size,
            sort_by=sort_by,
            is_enabled=is_enabled,
        )
        data = [DictionaryResponse.model_validate(entity) for entity in entities]
        return PageData(data=data, total=total)

    async def export(
        self,
        user: UserPayload,
        dict_type: str | None = None,
        keyword: str | None = None,
        sort_by: bool | None = None,
        is_enabled: bool | None = None,
    ) -> bytes:
        """导出字典数据为 Excel 字节流,管理员权限,支持 list_page 同款筛选"""
        self._ensure_admin(user)

        entities = await self.repository.find_all_for_export(
            dict_type=dict_type,
            keyword=keyword,
            sort_by=sort_by,
            is_enabled=is_enabled,
        )
        if not entities:
            raise DictionaryExportEmptyError()

        rows = []
        for entity in entities:
            rows.append(
                {
                    "类型": DICTIONARY_TYPE_CODE_TO_LABEL.get(entity.type, entity.type),
                    "字典键": entity.dict_key,
                    "字典取值": entity.dict_value,
                    "排序权重": entity.sort_order,
                    "是否启用": "是" if entity.is_enabled else "否",
                }
            )

        df = pd.DataFrame(rows, columns=DICTIONARY_EXPORT_HEADERS)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="字典数据", index=False)
        bio.seek(0)
        return bio.getvalue()

    async def import_excel(self, user: UserPayload, file: UploadFile) -> DictionaryImportResult:
        """从 Excel 导入字典数据,管理员权限"""
        self._ensure_admin(user)

        if not file.filename:
            raise DictionaryImportFileEmptyError()

        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise DictionaryImportFormatError()

        contents = await file.read()
        if not contents:
            raise DictionaryImportFileEmptyError()

        try:
            wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        except Exception as exc:
            logger.exception("Failed to parse dictionary import excel")
            raise DictionaryImportParseError() from exc

        sheet = wb.active
        if sheet is None:
            raise DictionaryImportParseError()

        # 读取表头
        header_row = next(sheet.iter_rows(values_only=True), None)
        if header_row is None:
            raise DictionaryImportHeaderError()

        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        if headers != DICTIONARY_EXPORT_HEADERS:
            raise DictionaryImportHeaderError()

        total = 0
        success = 0
        failed = 0
        errors: list[str] = []
        seen_keys: set[tuple[str, str]] = set()

        # 从第二行开始读取数据
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过完全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            total += 1

            try:
                if len(row) < len(DICTIONARY_EXPORT_HEADERS):
                    raise ValueError("insufficient columns")

                type_label = str(row[0]).strip() if row[0] is not None else ""
                dict_key = str(row[1]).strip() if row[1] is not None else ""
                dict_value = str(row[2]).strip() if row[2] is not None else ""
                sort_order_raw = row[3]
                is_enabled_raw = row[4]

                if not type_label:
                    raise ValueError("type is required")
                if not dict_key:
                    raise ValueError("dict_key is required")
                self._validate_dict_key_format(dict_key)
                if not dict_value:
                    raise ValueError("dict_value is required")

                dict_type = DICTIONARY_TYPE_LABEL_TO_CODE.get(type_label)
                if not dict_type:
                    raise DictionaryImportTypeInvalidError()

                unique_key = (dict_type, dict_key)
                if unique_key in seen_keys:
                    raise ValueError(f"duplicate dict_key '{dict_key}' under type '{type_label}' in Excel")
                seen_keys.add(unique_key)

                try:
                    sort_order = int(sort_order_raw) if sort_order_raw is not None else 0
                except (TypeError, ValueError) as exc:
                    raise ValueError(f"invalid sort_order: {sort_order_raw}") from exc

                if isinstance(is_enabled_raw, bool):
                    is_enabled = is_enabled_raw
                else:
                    is_enabled_str = str(is_enabled_raw).strip() if is_enabled_raw is not None else ""
                    is_enabled = is_enabled_str in ("是", "True", "true", "1", "Y", "y")

                existing = await self.repository.find_by_type_and_key(dict_type, dict_key)
                if existing:
                    existing.dict_value = dict_value
                    existing.sort_order = sort_order
                    existing.is_enabled = is_enabled
                    await self.repository.update(existing)
                else:
                    entity = SystemDictionary(
                        type=dict_type,
                        dict_key=dict_key,
                        dict_value=dict_value,
                        sort_order=sort_order,
                        is_enabled=is_enabled,
                    )
                    await self.repository.save(entity)

                success += 1
            except Exception as exc:
                failed += 1
                reason = str(exc)
                errors.append(f"第 {row_idx} 行导入失败: {reason}")

        return DictionaryImportResult(total=total, success=success, failed=failed, errors=errors)
