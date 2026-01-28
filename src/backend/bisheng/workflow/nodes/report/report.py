import io
import os
import re
import tempfile
from dataclasses import dataclass
from enum import Enum
from typing import Dict, Tuple, Any, List, Optional
from urllib.parse import urlparse
from uuid import uuid4

import pandas as pd
import requests
from charset_normalizer import detect
from loguru import logger
from openpyxl import load_workbook

from bisheng.core.storage.minio.minio_manager import get_minio_storage_sync
from bisheng.utils.docx_temp import DocxTemplateRender
from bisheng.workflow.callback.event import OutputMsgData
from bisheng.workflow.nodes.base import BaseNode


class ResourceType(Enum):
    """Resource Type Enumeration"""
    IMAGE = "image"
    TABLE = "table"
    TEXT = "text"


class ImageSource(Enum):
    """Image Source Type"""

    LOCAL_FILE = "local_file"  # Local file path
    HTTP_URL = "http_url"  # HTTP/HTTPSLinks
    MINIO_PATH = "minio_path"  # MinIOPath
    MARKDOWN_REF = "markdown_ref"  # MarkdownQuote Post Format


class TableSource(Enum):
    """Table Source Type"""

    MARKDOWN_TABLE = "markdown_table"  # |---|---|Format
    CSV_CONTENT = "csv_content"  # [file content begin]...[end]Format
    EXCEL_CONTENT = "excel_content"  # ExcelFile contents


@dataclass
class ResourceData:
    """Resource Data Classes"""

    resource_id: int
    resource_type: ResourceType
    placeholder: str
    position: int  # Position in original text
    original_content: str  # Original Match
    pattern_name: str  # Matching Pattern Name

    # Image Specific Fields
    image_source: Optional[ImageSource] = None
    original_path: Optional[str] = None
    local_path: Optional[str] = None
    alt_text: Optional[str] = None

    # Table-specific fields
    table_source: Optional[TableSource] = None
    table_data: Optional[List[List[str]]] = None
    alignments: Optional[List[str]] = None
    file_name: Optional[str] = None

    # Process status
    download_success: bool = False
    error_message: Optional[str] = None


class ResourcePlaceholderManager:
    """Unified Resource Placeholder Manager"""

    def __init__(self):
        self.counter = 0
        self.resources: List[ResourceData] = []
        self.placeholder_map: Dict[str, ResourceData] = {}  # placeholder -> resourceMapping

    def create_placeholder(
            self, resource_type: ResourceType, position: int, original_content: str, pattern_name: str
    ) -> str:
        """Create new placeholder"""
        resource_id = self.counter
        self.counter += 1

        placeholder = f"__RESOURCE_{resource_id:04d}__"  # Use4Digit numbers for easy sorting

        resource = ResourceData(
            resource_id=resource_id,
            resource_type=resource_type,
            placeholder=placeholder,
            position=position,
            original_content=original_content,
            pattern_name=pattern_name,
        )

        self.resources.append(resource)
        self.placeholder_map[placeholder] = resource

        return placeholder

    def get_resource_by_placeholder(self, placeholder: str) -> Optional[ResourceData]:
        """Get resources based on placeholders"""
        return self.placeholder_map.get(placeholder)

    def get_resources_by_type(self, resource_type: ResourceType) -> List[ResourceData]:
        """Get all resources of the specified type"""
        return [r for r in self.resources if r.resource_type == resource_type]

    def get_sorted_resources(self) -> List[ResourceData]:
        """Get all resources sorted by location"""
        return sorted(self.resources, key=lambda x: x.position)


@dataclass
class MatchPattern:
    """Match Pattern Definition"""

    name: str
    resource_type: ResourceType
    pattern: str
    flags: int
    priority: int  # Priority, the smaller the number, the higher the priority
    handler_method: str


class OverlapResolver:
    """Overlay Resource Resolver"""

    @staticmethod
    def resolve_overlapping_resources(resources: List[ResourceData]) -> List[ResourceData]:
        """Resolve overlapping resource issues"""
        if not resources:
            return []

        # Sort by starting position
        sorted_resources = sorted(resources, key=lambda x: x.position)
        resolved_resources = []

        for current in sorted_resources:
            # Check for overlap with resolved resources
            overlapping_existing = None
            for existing in resolved_resources:
                if OverlapResolver._is_overlapping(current, existing):
                    overlapping_existing = existing
                    break

            if overlapping_existing:
                # Handle overlaps: Prefer to keep more precise matches
                if OverlapResolver._should_replace(current, overlapping_existing):
                    resolved_resources.remove(overlapping_existing)
                    resolved_resources.append(current)
                    logger.info(f"Replace overlapping resources: {overlapping_existing.placeholder} -> {current.placeholder}")
                else:
                    logger.info(f"Skip overlapping resources: {current.placeholder}")
            else:
                resolved_resources.append(current)

        return sorted(resolved_resources, key=lambda x: x.position)

    @staticmethod
    def _is_overlapping(res1: ResourceData, res2: ResourceData) -> bool:
        """Determine if the two resources overlap"""
        # Calculate end position
        end1 = res1.position + len(res1.original_content)
        end2 = res2.position + len(res2.original_content)

        # Check for overlapping areas
        return not (end1 <= res2.position or end2 <= res1.position)

    @staticmethod
    def _should_replace(new_res: ResourceData, existing_res: ResourceData) -> bool:
        """Determine if existing resources should be replaced with new resources"""
        # Priority Rules:
        # 1. More specific matching priorities (e.g. Markdown Images > General URL）
        # 2. Match length more precisely first
        # 3. Same type of resource, match first priority

        priority_map = {
            "markdown_table": 1,  # The form has the highest priority and includes additional resources
            "markdown_image": 2,  # Images next
            "minio_image": 3,
            "http_image": 4,
            "minio_excel_csv": 5,  # Excel/CSVDoc.
            "http_excel_csv": 6,
            "local_excel_csv": 7,
            "local_image": 8,
        }

        new_priority = priority_map.get(new_res.pattern_name, 999)
        existing_priority = priority_map.get(existing_res.pattern_name, 999)

        return new_priority < existing_priority


class PatternMatcher:
    """Pattern matchers, which are responsible for identifying various resources in the content"""

    def __init__(self):
        self.patterns = [
            # Priority1: MarkdownImages (The clearest format)
            MatchPattern(
                name="markdown_image",
                resource_type=ResourceType.IMAGE,
                pattern=r"!\[([^\]]*)\]\(([^)]+\.(?:png|jpg|jpeg|bmp|gif|webp)(?:\?[^)]*)?)\)",
                flags=re.IGNORECASE,
                priority=1,
                handler_method="_handle_markdown_image",
            ),
            # Priority2: IndependentMarkdownTable (supports indentation at the beginning of a row)
            MatchPattern(
                name="markdown_table",
                resource_type=ResourceType.TABLE,
                pattern=r"(\s*\|[^\r\n]*\|[^\r\n]*(?:\r?\n\s*\|[^\r\n]*\|[^\r\n]*)+)",
                flags=re.MULTILINE,
                priority=2,
                handler_method="_handle_markdown_table",
            ),
            # Priority3: MinIOImage Path
            MatchPattern(
                name="minio_image",
                resource_type=ResourceType.IMAGE,
                pattern=r"((?:minio://|/bisheng/|/tmp-dir/|/tmp/)[^\s]*\.(?:png|jpg|jpeg|bmp|gif|webp))",
                flags=re.IGNORECASE,
                priority=3,
                handler_method="_handle_minio_image",
            ),
            # Priority4: HTTPHero Image Link
            MatchPattern(
                name="http_image",
                resource_type=ResourceType.IMAGE,
                pattern=r"(https?://[^\s\u4e00-\u9fff]*\.(?:png|jpg|jpeg|bmp|gif|webp)(?:\?[^\s\u4e00-\u9fff]*)?)",
                flags=re.IGNORECASE,
                priority=4,
                handler_method="_handle_http_image",
            ),
            # Priority5: Excel/CSVDoc. (MinIOPath andHTTPLinks)
            MatchPattern(
                name="minio_excel_csv",
                resource_type=ResourceType.TABLE,
                pattern=r"((?:minio://|/bisheng/|/tmp-dir/|/tmp/)[^\s]*\.(?:xlsx?|csv))",
                flags=re.IGNORECASE,
                priority=5,
                handler_method="_handle_minio_excel_csv",
            ),
            # Priority6: HTTP Excel/CSVLink to Files
            MatchPattern(
                name="http_excel_csv",
                resource_type=ResourceType.TABLE,
                pattern=r"(https?://[^\s\u4e00-\u9fff]*\.(?:xlsx?|csv)(?:\?[^\s\u4e00-\u9fff]*)?)",
                flags=re.IGNORECASE,
                priority=6,
                handler_method="_handle_http_excel_csv",
            ),
            # Priority7: Perangkat iniExcel/CSVFilePath
            MatchPattern(
                name="local_excel_csv",
                resource_type=ResourceType.TABLE,
                pattern=r"([^\s]*[/\\][^\s]*\.(?:xlsx?|csv))",
                flags=re.IGNORECASE,
                priority=7,
                handler_method="_handle_local_excel_csv",
            ),
            # Priority8: Local Image Path (Broadest, last match)
            MatchPattern(
                name="local_image",
                resource_type=ResourceType.IMAGE,
                pattern=r"([^\s]*[/\\][^\s]*\.(?:png|jpg|jpeg|bmp|gif|webp))",
                flags=re.IGNORECASE,
                priority=8,
                handler_method="_handle_local_image",
            ),
        ]

    def find_all_matches(self, content: str) -> List[Dict[str, Any]]:
        """Find all matches in content"""
        all_matches = []

        for pattern in self.patterns:
            for match in re.finditer(pattern.pattern, content, pattern.flags):
                match_info = {
                    "pattern_name": pattern.name,
                    "resource_type": pattern.resource_type,
                    "start": match.start(),
                    "end": match.end(),
                    "full_match": match.group(0),
                    "groups": match.groups(),
                    "handler_method": pattern.handler_method,
                    "priority": pattern.priority,
                }
                all_matches.append(match_info)

        # Sort by location only, no deduplication (overlap goes to next steps)
        return sorted(all_matches, key=lambda x: x["start"])


class ContentParser:
    """Content parser, responsible for parsing variable content and generating placeholders"""

    def __init__(self, minio_client):
        self.placeholder_manager = ResourcePlaceholderManager()
        self.pattern_matcher = PatternMatcher()
        self.minio_client = minio_client
        self.logger = logger
        self._table_image_resources = []  # Temporary storage of image resources within the table

    def parse_variable_content(self, var_name: str, content: str) -> tuple[str, List[ResourceData]]:
        """
        Parsing the contents of a single variable

        Args:
            var_name: Variables
            content: Variable Content

        Returns:
            (Post-Processing Content, Resource list)
        """
        if not isinstance(content, str):
            content = str(content)

        self.logger.info(f"Start parsing variables '{var_name}', CL: {len(content)}")

        # Add Detailed Content Preview Log
        if content:
            content_preview = content.replace('\n', '\\n').replace('\r', '\\r').replace('\t', '\\t')
            if len(content_preview) > 300:
                self.logger.info(f"Variables '{var_name}' Content Preview (Before300characters. : {content_preview[:300]}...")
            else:
                self.logger.info(f"Variables '{var_name}' Box Full Text: {content_preview}")
        else:
            self.logger.info(f"Variables '{var_name}' Konten kosong")

        # 1. Find all matches
        matches = self.pattern_matcher.find_all_matches(content)

        self.logger.info(f"Variables '{var_name}' found in {len(matches)} resource matches")
        for i, match in enumerate(matches):
            self.logger.debug(f"Matched items {i + 1}: {match['pattern_name']} at {match['start']}-{match['end']}")

        # 2. Create Resource Object
        resources = []
        for match in matches:
            try:
                # Create Placeholder
                placeholder = self.placeholder_manager.create_placeholder(
                    resource_type=match["resource_type"],
                    position=match["start"],
                    original_content=match["full_match"],
                    pattern_name=match["pattern_name"],
                )

                # Get resource objects and populate details
                resource = self.placeholder_manager.get_resource_by_placeholder(placeholder)

                # Call the corresponding processing method
                handler = getattr(self, match["handler_method"])
                handler(resource, match)

                resources.append(resource)

                self.logger.info(f"Successfully processed resources: {match['pattern_name']} -> {placeholder}")

            except Exception as e:
                self.logger.error(f"Failed to process resource: {match['pattern_name']}, Error-free: {str(e)}")
                # Continue with other resources without disrupting the process

        # 3. Resolve Overlap Issues
        resolved_resources = OverlapResolver.resolve_overlapping_resources(resources)

        # 4. Collect image resources within the form
        table_image_resources = []
        for resource in resolved_resources:
            if resource.resource_type == ResourceType.TABLE and hasattr(self, '_table_image_resources'):
                table_image_resources.extend(self._table_image_resources)
                self._table_image_resources = []  # Empty Temporary List

        # 5. REPLACE CONTENT
        processed_content = self._replace_content_with_placeholders(content, resolved_resources)

        # 6. Merge All Resources (Primary Resource + In-Table Picture Resources)
        all_resources = resolved_resources + table_image_resources

        self.logger.info(
            f"Variables '{var_name}' Parsing complete, generating {len(all_resources)} resources (primary {len(resolved_resources)} , images in table {len(table_image_resources)} words)")

        return processed_content, all_resources

    def _replace_content_with_placeholders(self, content: str, resources: List[ResourceData]) -> str:
        """Correct Content Substitution Logic"""
        # Replace from back to front by position (avoid position offset)
        sorted_resources = sorted(resources, key=lambda x: x.position, reverse=True)

        processed_content = content
        for resource in sorted_resources:
            start_pos = resource.position
            end_pos = start_pos + len(resource.original_content)

            # Verify Content Match
            actual_content = processed_content[start_pos:end_pos]
            if actual_content != resource.original_content:
                self.logger.warning(f"Content mismatch, skipping replacement: Expectations '{resource.original_content}', actual '{actual_content}'")
                continue

            # Execute Replacement
            processed_content = processed_content[:start_pos] + resource.placeholder + processed_content[end_pos:]

            self.logger.debug(f"Replaced successfully: '{resource.original_content}' -> '{resource.placeholder}'")

        return processed_content

    # Various resource handling methods
    def _handle_markdown_image(self, resource: ResourceData, match: Dict):
        """<g id="Bold">Medical Treatment:</g>MarkdownFormat image"""
        alt_text, img_path = match["groups"]

        resource.image_source = ImageSource.MARKDOWN_REF
        resource.original_path = img_path
        resource.alt_text = alt_text or "Images"

        self.logger.debug(f"MarkdownImages: alt='{alt_text}', path='{img_path}'")

    def _handle_markdown_table(self, resource: ResourceData, match: Dict):
        """Handling independentMarkdownTable Filter"""
        table_content = match["full_match"]

        resource.table_source = TableSource.MARKDOWN_TABLE
        resource.file_name = "table_data.csv"  # Default filename

        # Parsing table data
        resource.table_data, resource.alignments = self._parse_table_data(table_content)

        self.logger.debug(f"MarkdownTable Filter: rows={len(resource.table_data) if resource.table_data else 0}")

    def _handle_minio_image(self, resource: ResourceData, match: Dict):
        """<g id="Bold">Medical Treatment:</g>MinIOImage Path"""
        img_path = match["full_match"]

        # Unified smart local downloadpipelineLocal: ->MinIO->Original Path
        resource.image_source = ImageSource.LOCAL_FILE
        resource.original_path = img_path
        resource.alt_text = "Images"

        self.logger.debug(f"MinIOImages: path='{img_path}'")

    def _handle_http_image(self, resource: ResourceData, match: Dict):
        """<g id="Bold">Medical Treatment:</g>HTTPHero Image Link"""
        img_url = match["full_match"]

        resource.image_source = ImageSource.HTTP_URL
        resource.original_path = img_url
        resource.alt_text = "Images"

        self.logger.debug(f"HTTPImages: url='{img_url}'")

    def _handle_local_image(self, resource: ResourceData, match: Dict):
        """Working with local image paths"""
        img_path = match["full_match"]

        # Clean extra characters in the path, such as [''] etc.
        img_path = img_path.strip("[]'\"")

        resource.image_source = ImageSource.LOCAL_FILE
        resource.original_path = img_path
        resource.alt_text = "Images"

        self.logger.debug(f"Location Image: path='{img_path}'")

    def _handle_minio_excel_csv(self, resource: ResourceData, match: Dict):
        """<g id="Bold">Medical Treatment:</g>MinIO Excel/CSVDoc."""
        file_path = match["full_match"]

        resource.table_source = self._determine_table_source_by_extension(file_path)
        resource.original_path = file_path
        resource.file_name = os.path.basename(file_path)

        self.logger.debug(f"MinIO Excel/CSVDoc.: path='{file_path}', type='{resource.table_source.value}'")

    def _handle_http_excel_csv(self, resource: ResourceData, match: Dict):
        """<g id="Bold">Medical Treatment:</g>HTTP Excel/CSVDoc."""
        file_url = match["full_match"]

        resource.table_source = self._determine_table_source_by_extension(file_url)
        resource.original_path = file_url
        resource.file_name = os.path.basename(urlparse(file_url).path) or "spreadsheet_file"

        self.logger.debug(f"HTTP Excel/CSVDoc.: url='{file_url}', type='{resource.table_source.value}'")

    def _handle_local_excel_csv(self, resource: ResourceData, match: Dict):
        """Handling LocalExcel/CSVDoc."""
        file_path = match["full_match"]

        # Clean extra characters in the path
        file_path = file_path.strip("[]'\"")

        resource.table_source = self._determine_table_source_by_extension(file_path)
        resource.original_path = file_path
        resource.file_name = os.path.basename(file_path)

        self.logger.debug(f"Perangkat iniExcel/CSVDoc.: path='{file_path}', type='{resource.table_source.value}'")

    def _determine_table_source_by_extension(self, file_path: str) -> TableSource:
        """Determine table source type based on file extension"""
        file_path_lower = file_path.lower()

        if file_path_lower.endswith('.csv'):
            return TableSource.CSV_CONTENT
        elif file_path_lower.endswith(('.xlsx', '.xls')):
            return TableSource.EXCEL_CONTENT
        else:
            # Defaulted toCSV
            return TableSource.CSV_CONTENT

    def _parse_table_data(self, table_content: str) -> tuple[List[List[str]], List[str]]:
        """Parse the table data while working with the images in the table"""
        try:
            # Analyze the table structure first
            table_data, alignments = self._parse_markdown_table_from_content(table_content)

            # Working with image links in forms
            self._process_images_in_table(table_data)

            return table_data, alignments
        except Exception as e:
            self.logger.error(f"Failed to parse table data: {str(e)}")
            return [["Parse Failure", str(e)]], ["left"]

    def _parse_markdown_table_from_content(self, content: str) -> Tuple[list, list]:
        """
        Parsing from file contentsMarkdownTable Filter

        Args:
            content: File contents

        Returns:
            tuple: (Form Data, Alignment Info List)
        """
        try:
            # Find allMarkdownTable (row indentation supported)
            table_pattern = r"(\s*\|[^\r\n]*\|[^\r\n]*(?:\r?\n\s*\|[^\r\n]*\|[^\r\n]*)+)"
            tables = re.findall(table_pattern, content, re.MULTILINE)

            if not tables:
                self.logger.warning("Not found in contentMarkdownTable Filter")
                return [["Content parsing failed", "Table data not found"]], ["left"]

            # Combine all tables (if there are multiple tables, combine into one large table)
            all_rows = []
            alignments = []

            for i, table_content in enumerate(tables):
                # Keep all rows, including empty rows - Fully preserving the original table structure
                lines = [line.strip() for line in table_content.strip().split("\n")]

                if len(lines) < 2:
                    continue

                table_rows = []
                table_alignments = []
                separator_found = False

                for line in lines:
                    # Skip completely empty rows, but leave empty table rows with only vertical lines
                    if not line:
                        continue

                    # Check if it is a delimiter line
                    if self._is_separator_line(line):
                        table_alignments = self._parse_alignments(line)
                        separator_found = True
                        continue

                    # Parse Data Rows - Keep all table rows, including empty rows
                    cells = self._parse_table_row(line)
                    # Remove if cells Condition, leave empty table rows
                    cleaned_cells = []
                    for cell in cells:
                        cleaned_cell = self._clean_cell_content(cell)
                        cleaned_cells.append(cleaned_cell)
                    table_rows.append(cleaned_cells)

                # Use default alignment if no separator found
                if not separator_found and table_rows:
                    table_alignments = ["left"] * len(table_rows[0])

                # Add table to total list
                if table_rows:
                    if i == 0:
                        alignments = table_alignments
                    all_rows.extend(table_rows)

            # Make sure all rows have the same number of columns
            if all_rows:
                max_cols = max(len(row) for row in all_rows)
                for row in all_rows:
                    while len(row) < max_cols:
                        row.append("")

                # Make sure the number of aligned messages matches the number of columns
                while len(alignments) < max_cols:
                    alignments.append("left")
                alignments = alignments[:max_cols]

            self.logger.info(f"Successfully parsed table in content, size: {len(all_rows)}Parade x {len(alignments)}column")
            return all_rows, alignments

        except Exception as e:
            self.logger.error(f"Failed to parse table in content: {str(e)}")
            return [["Table parsing failed", str(e)]], ["left"]

    def _is_separator_line(self, line: str) -> bool:
        """Check if yesMarkdownTable Separator Row"""
        content = line.strip().strip("|").strip()
        if not content:
            return False

        cells = [cell.strip() for cell in content.split("|")]

        # At least one cell must contain a delimiter character (-OR:）
        has_separator_chars = False
        for cell in cells:
            if not cell:
                continue
            # Check for delimiter characters
            if '-' in cell or ':' in cell:
                has_separator_chars = True
            # Check for anything else after removing the separator character
            clean_cell = cell.replace("-", "").replace(":", "").strip()
            if clean_cell:
                return False

        # Only lines containing delimiter characters can be considered delimiter lines
        return has_separator_chars

    def _parse_alignments(self, separator_line: str) -> list:
        """Parse column alignment from delimiter rows"""
        alignments = []
        content = separator_line.strip().strip("|").strip()
        cells = [cell.strip() for cell in content.split("|")]

        for cell in cells:
            if not cell:
                alignments.append("left")
                continue

            if cell.startswith(":") and cell.endswith(":"):
                alignments.append("center")
            elif cell.endswith(":"):
                alignments.append("right")
            else:
                alignments.append("left")

        return alignments

    def _parse_table_row(self, line: str) -> list:
        """Parse Table Rows"""
        content = line.strip()
        if content.startswith("|"):
            content = content[1:]
        if content.endswith("|"):
            content = content[:-1]

        cells = []
        current_cell = ""
        escaped = False

        for char in content:
            if escaped:
                current_cell += char
                escaped = False
            elif char == "\\":
                escaped = True
                current_cell += char
            elif char == "|":
                cells.append(current_cell.strip())
                current_cell = ""
            else:
                current_cell += char

        # Always add the last cell to ensure empty table rows are also parsed correctly
        cells.append(current_cell.strip())

        return cells

    def _clean_cell_content(self, cell: str) -> str:
        """Clean cell contents"""
        if not cell:
            return ""

        cleaned = cell.strip()
        cleaned = cleaned.replace("**", "")
        cleaned = cleaned.replace("*", "")
        cleaned = cleaned.replace("`", "")

        cleaned = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", cleaned)

        if len(cleaned) > 100:
            cleaned = cleaned[:97] + "..."

        return cleaned

    def _process_images_in_table(self, table_data: List[List[str]]):
        """Process image links within the table to create image assets"""
        if not table_data:
            return

        for row_idx, row in enumerate(table_data):
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue

                # Find and process picture links in cell contents
                updated_cell = self._process_cell_images(cell)
                table_data[row_idx][col_idx] = updated_cell

    def _process_cell_images(self, cell_content: str) -> str:
        """Process images inside cells, using existing pattern matching logic"""
        if not cell_content:
            return cell_content

        # Reuse existing pattern matching logic to find all images
        matches = self.pattern_matcher.find_all_matches(cell_content)

        # Handle image type matches only
        image_matches = [m for m in matches if m["resource_type"] == ResourceType.IMAGE]

        updated_content = cell_content

        # Process from back to front to avoid position offset
        for match in reversed(image_matches):
            # Create image resource
            placeholder = self.placeholder_manager.create_placeholder(
                resource_type=match["resource_type"],
                position=match["start"],
                original_content=match["full_match"],
                pattern_name=match["pattern_name"],
            )

            # Get the resource object and call the corresponding processing method
            resource = self.placeholder_manager.get_resource_by_placeholder(placeholder)
            handler = getattr(self, match["handler_method"])
            handler(resource, match)

            # Add image resources within the table to a temporary list (to avoid being skipped by the overlap resolver)
            self._table_image_resources.append(resource)

            self.logger.info(f"In-Table Images: {match['pattern_name']} {match['full_match']} -> {placeholder}")

            # Replace with a placeholder
            start_pos = match["start"]
            end_pos = match["end"]
            updated_content = updated_content[:start_pos] + placeholder + updated_content[end_pos:]

        return updated_content


class ResourceDownloadManager:
    """Resource Download Manager"""

    def __init__(self, minio_client):
        self.minio_client = minio_client
        self.temp_files: List[str] = []  # Manage all temporary files
        self.logger = logger

    def download_all_resources(self, resources: List[ResourceData]) -> Dict[str, Any]:
        """
        Download all resources that need to be downloaded

        Returns:
            Download stats
        """
        stats = {"total": len(resources), "images_success": 0, "images_failed": 0, "tables_processed": 0, "errors": []}

        image_resources = [r for r in resources if r.resource_type == ResourceType.IMAGE]
        table_resources = [r for r in resources if r.resource_type == ResourceType.TABLE]

        self.logger.info(f"Start downloading resources: Images {len(image_resources)} Pcs, Table Filter {len(table_resources)} Pcs")

        # Download image resources
        for resource in image_resources:
            try:
                self._download_image_resource(resource)
                if resource.download_success:
                    stats["images_success"] += 1
                else:
                    stats["images_failed"] += 1
            except Exception as e:
                stats["images_failed"] += 1
                stats["errors"].append(f"This image failed to load {resource.original_path}: {str(e)}")
                self.logger.error(f"Abnormal image download: {str(e)}")

        # Processing Table Resources
        for resource in table_resources:
            try:
                if resource.table_source == TableSource.MARKDOWN_TABLE:
                    # MarkdownThe form has been processed in the parsing phase, just verify
                    self._validate_table_resource(resource)
                elif resource.table_source in [TableSource.CSV_CONTENT, TableSource.EXCEL_CONTENT]:
                    # Excel/CSVFile needs to be downloaded and parsed
                    self._download_and_parse_table_file(resource)
                    self._validate_table_resource(resource)
                stats["tables_processed"] += 1
            except Exception as e:
                stats["errors"].append(f"Form processing failed {resource.file_name}: {str(e)}")
                self.logger.error(f"Form Processing Exception: {str(e)}")

        self.logger.info(
            f"Resource download complete: Berhasil {stats['images_success']}, Kalah {stats['images_failed']}, Table Filter {stats['tables_processed']}"
        )

        return stats

    def _download_image_resource(self, resource: ResourceData):
        """Download individual image assets"""
        if resource.image_source == ImageSource.LOCAL_FILE:
            self._handle_smart_local_download(resource)
        elif resource.image_source == ImageSource.HTTP_URL:
            self._handle_http_download(resource)
        elif resource.image_source == ImageSource.MARKDOWN_REF:
            # MarkdownReferences need to be further judged based on the path type
            self._handle_markdown_reference(resource)

    def _handle_smart_local_download(self, resource: ResourceData):
        """
        Handle local path files intelligently
        1. Prioritize local downloads
        2. If not locally, resolve the first directory tobucketFROMMinIOMengunduh
        3. If none, return to the original path
        """
        file_path = resource.original_path

        # Buyi1: Try local files first
        if os.path.exists(file_path):
            resource.local_path = file_path
            resource.download_success = True
            self.logger.info(f"Local image file exists: {file_path}")
            return

        # Buyi2: Try fromMinIODownload (parsingbucketAndobjectLast name
        self.logger.info(f"Local file does not exist, try fromMinIOMengunduh: {file_path}")

        # Resolve Path AcquisitionbucketAndobject_name
        bucket_name, object_name = self._parse_path_for_minio(file_path)

        if bucket_name and object_name:
            # attemptMinIOMengunduh
            success = self._try_minio_download(resource, bucket_name, object_name)
            if success:
                return

        # Buyi3: Neither download was successful, returning to the original path
        resource.local_path = file_path
        resource.download_success = True  # Return to original path is also considered successful
        self.logger.warning(f"Image download failed, use original path: {file_path}")

    def _parse_path_for_minio(self, file_path: str) -> tuple[str, str]:
        """
        Resolve file path asMinIOright of privacybucketAndobject_name
        Supports correct mapping of special paths

        rule’:
        - "/bisheng/xxx" -> bucket="bisheng", object_name="xxx"  
        - "/tmp-dir/xxx" -> bucket="tmp-dir", object_name="xxx"
        - "/tmp/xxx" -> Try first bucket="bisheng", object_name="tmp/xxx", then try bucket="tmp-dir", object_name="xxx"
        - "images/photo.jpg" -> bucket="images", object_name="photo.jpg"
        """
        if not file_path:
            return None, None

        # Special Path Processing
        if file_path.startswith("/bisheng/"):
            # /bisheng/object/name -> bucket="bisheng", object_name="object/name"
            object_name = file_path[9:]  # Remove '/bisheng/'
            return "bisheng", object_name if object_name else None

        elif file_path.startswith("/tmp-dir/"):
            # /tmp-dir/object/name -> bucket="tmp-dir", object_name="object/name"  
            object_name = file_path[9:]  # Remove '/tmp-dir/'
            return "tmp-dir", object_name if object_name else None

        elif file_path.startswith("/tmp/"):
            # /tmp/xxx -> Returns multiple possiblebucketoption, the caller needs to try
            # Here first return to the Lordbucketmapping, the caller should implement morebucketTry Logic
            object_name = file_path[5:]  # Remove '/tmp/'
            return "tmp-dir", object_name if object_name else None

        # Generic path resolution (keep the original logic)
        clean_path = file_path.lstrip("/")
        if not clean_path or "/" not in clean_path:
            # Cannot parse without directory delimitation
            return None, None

        # Detach first directory and remaining paths
        parts = clean_path.split("/", 1)
        if len(parts) == 2:
            bucket_name = parts[0]
            object_name = parts[1]

            # VerifybucketFirst name is legal (simple check)
            if bucket_name and object_name and bucket_name.replace("_", "").replace("-", "").isalnum():
                return bucket_name, object_name

        return None, None

    def _try_minio_download(self, resource: ResourceData, bucket_name: str, object_name: str) -> bool:
        """
        Try fromMinIODownload file

        Returns:
            bool: Was the download successful?
        """
        try:
            # Checks to see if file exists.
            if not self.minio_client.object_exists_sync(bucket_name, object_name):
                self.logger.debug(f"MinIOFile don\'t exists: {bucket_name}/{object_name}")
                return False

            # Download File Contents
            file_content = self.minio_client.get_object_sync(bucket_name, object_name)

            # Generate temporary filename
            file_ext = os.path.splitext(object_name)[1] or ".dat"
            filename = f"{uuid4().hex}{file_ext}"
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, filename)

            # Save to Temporary File
            with open(temp_file, "wb") as f:
                f.write(file_content)

            # Update resource information
            resource.local_path = temp_file
            resource.download_success = True
            self.temp_files.append(temp_file)

            self.logger.info(f"MinIODownloaded Successfully: {bucket_name}/{object_name} -> {temp_file}")
            return True

        except Exception as e:
            self.logger.debug(f"FROMMinIODownload failed {bucket_name}/{object_name}: {str(e)}")
            return False

    def _handle_http_download(self, resource: ResourceData):
        """<g id="Bold">Medical Treatment:</g>HTTPMengunduh"""
        try:
            local_path, success = self._download_file_from_url(resource.original_path)
            resource.local_path = local_path
            resource.download_success = success

            if success:
                self.temp_files.append(local_path)
                self.logger.info(f"HTTPImage downloaded successfully: {resource.original_path} -> {local_path}")
            else:
                self.logger.warning(f"HTTPImage download failed, use original path: {resource.original_path}")

        except Exception as e:
            resource.local_path = resource.original_path
            resource.download_success = True  # Return to original path is also considered successful
            self.logger.warning(f"HTTPImage download failed, use original path: {resource.original_path} (Error-free: {str(e)})")

    def _handle_markdown_reference(self, resource: ResourceData):
        """<g id="Bold">Medical Treatment:</g>MarkdownQuotes (specific type needs to be determined)"""
        img_path = resource.original_path

        if self._is_valid_url(img_path):
            resource.image_source = ImageSource.HTTP_URL
            self._handle_http_download(resource)
        else:
            # All non-HTTP URLAll paths to smart local downloadspipeline
            resource.image_source = ImageSource.LOCAL_FILE
            self._handle_smart_local_download(resource)

    def _download_and_parse_table_file(self, resource: ResourceData):
        """Download and parseExcel/CSVDoc."""
        file_path = resource.original_path

        # 1. Try downloading the file first
        local_file_path = self._download_table_file(file_path)

        if not local_file_path:
            raise Exception(f"Unable to download form file: {file_path}")

        # 2. Parse based on file type
        try:
            if resource.table_source == TableSource.CSV_CONTENT:
                resource.table_data, resource.alignments = self._parse_csv_file(local_file_path)
            elif resource.table_source == TableSource.EXCEL_CONTENT:
                resource.table_data, resource.alignments = self._parse_excel_file(local_file_path)

            # Set download path
            resource.local_path = local_file_path
            resource.download_success = True
            self.logger.info(f"Table file parsed successfully: {file_path} -> {len(resource.table_data)}Parade")

        except Exception as e:
            self.logger.error(f"Table file parsing failed: {file_path}, Error-free: {str(e)}")
            # Create error table
            resource.table_data = [["File parsing failed", str(e)]]
            resource.alignments = ["left", "left"]
            resource.local_path = local_file_path if local_file_path else file_path  # Make sure there's a path
            resource.download_success = False

    def _download_table_file(self, file_path: str) -> Optional[str]:
        """Download form files to local temporary files"""
        # 1. Try local files first
        if os.path.exists(file_path):
            self.logger.info(f"Local table file exists: {file_path}")
            return file_path

        # 2. attemptHTTPMengunduh
        if self._is_valid_url(file_path):
            try:
                temp_file, success = self._download_file_from_url(file_path)
                if success and temp_file:
                    self.temp_files.append(temp_file)
                    self.logger.info(f"HTTPForm file downloaded successfully: {file_path} -> {temp_file}")
                    return temp_file
            except Exception as e:
                self.logger.warning(f"HTTPForm file download failed: {file_path}, Error-free: {str(e)}")

        # 3. Try fromMinIOMengunduh
        bucket_name, object_name = self._parse_path_for_minio(file_path)
        if bucket_name and object_name:
            try:
                if self.minio_client.object_exists_sync(bucket_name, object_name):
                    file_content = self.minio_client.get_object_sync(bucket_name, object_name)

                    # Generate Temporary Files
                    file_ext = os.path.splitext(object_name)[1] or ".dat"
                    filename = f"{uuid4().hex}{file_ext}"
                    temp_dir = tempfile.gettempdir()
                    temp_file = os.path.join(temp_dir, filename)

                    with open(temp_file, "wb") as f:
                        f.write(file_content)

                    self.temp_files.append(temp_file)
                    self.logger.info(f"MinIOForm file downloaded successfully: {bucket_name}/{object_name} -> {temp_file}")
                    return temp_file
            except Exception as e:
                self.logger.warning(f"MinIOForm file download failed: {file_path}, Error-free: {str(e)}")

        # 4. None of the downloads were successful
        self.logger.warning(f"Form file download failed: {file_path}")
        return None

    def _parse_csv_file(self, file_path: str) -> Tuple[List[List[str]], List[str]]:
        """analyzingCSVDoc."""
        try:
            # Auto-Detect Encoding
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                encoding_info = detect(raw_data)
                encoding = encoding_info['encoding'] or 'utf-8'

            # Usepandasread outCSVto better handle various formats
            df = pd.read_csv(file_path, encoding=encoding)

            # Convert to Tabular Data Format
            table_data = []

            # Add header
            headers = [str(col) for col in df.columns]
            table_data.append(headers)

            # Add Data Row
            for _, row in df.iterrows():
                row_data = [str(cell) if pd.notna(cell) else "" for cell in row]
                table_data.append(row_data)

            # Generate alignment information (default left alignment)
            alignments = ["left"] * len(headers)

            self.logger.info(f"CSVFile parsed successfully: {len(table_data)}Parade x {len(headers)}column")
            return table_data, alignments

        except Exception as e:
            self.logger.error(f"CSVFile parsing failed: {file_path}, Error-free: {str(e)}")
            raise

    def _parse_excel_file(self, file_path: str) -> Tuple[List[List[str]], List[str]]:
        """analyzingExcelDoc."""
        try:
            # Useopenpyxlread outExcelDoc.
            workbook = load_workbook(file_path, data_only=True)  # data_only=TrueGet Calculated Value

            # Using the first worksheet
            worksheet = workbook.active

            table_data = []
            max_col = 0

            # Read all rows
            for row in worksheet.iter_rows(values_only=True):
                # Skip completely blank rows
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                row_data = [str(cell) if cell is not None else "" for cell in row]
                table_data.append(row_data)
                max_col = max(max_col, len(row_data))

            # Make sure all rows have the same number of columns
            for row in table_data:
                while len(row) < max_col:
                    row.append("")

            # Generate alignment information (default left alignment)
            alignments = ["left"] * max_col

            self.logger.info(f"ExcelFile parsed successfully: {len(table_data)}Parade x {max_col}column")

            return table_data, alignments

        except Exception as e:
            self.logger.error(f"ExcelFile parsing failed: {file_path}, Error-free: {str(e)}")
            raise

    def _validate_table_resource(self, resource: ResourceData):
        """Validate form resources"""
        # Allow empty forms to exist, no more throwing"Table data is empty"Error-free
        if not resource.table_data:
            resource.table_data = []  # Ensuringtable_datais an empty list and notNone

        # Verify table data consistency
        if len(resource.table_data) > 0:
            col_count = len(resource.table_data[0])
            for i, row in enumerate(resource.table_data):
                if len(row) != col_count:
                    self.logger.warning(f"Form pg. {i + 1} The number of rows and columns is inconsistent, the null value will be filled")
                    while len(row) < col_count:
                        row.append("")

        # Verify alignment information
        if resource.alignments and resource.table_data:
            expected_cols = len(resource.table_data[0]) if resource.table_data else 0
            while len(resource.alignments) < expected_cols:
                resource.alignments.append("left")

    def _download_file_from_url(self, url: str) -> Tuple[str, bool]:
        """FROM URL Download file"""
        try:
            # Set request headers to simulate browser access
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }

            response = requests.get(url, headers=headers, timeout=30, verify=False)
            response.raise_for_status()

            # Get filename
            # accordingContent-TypeInferred extension
            content_type = response.headers.get("Content-Type", "").lower()
            if "image/png" in content_type:
                filename = f"{uuid4().hex}.png"
            elif "image/jpeg" in content_type or "image/jpg" in content_type:
                filename = f"{uuid4().hex}.jpg"
            elif "image/bmp" in content_type:
                filename = f"{uuid4().hex}.bmp"
            else:
                file_ext = url.split('.')[-1].lower()[-5:]
                filename = f"{uuid4().hex}.{file_ext}"

            # Creating temp file
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, filename)

            with open(temp_file, "wb") as f:
                f.write(response.content)

            return temp_file, True

        except Exception as e:
            self.logger.error(f"Download failed: {url}, Error-free: {str(e)}")
            return "", False

    def _download_file_from_minio(self, minio_path: str) -> Tuple[str, bool]:
        """FROMMinIODownload file"""
        try:
            # analyzingMinIOPath
            bucket_name = None
            object_name = None

            if minio_path.startswith("minio://"):
                # Format: minio://bucket/object/name
                parts = minio_path[8:].split("/", 1)
                if len(parts) == 2:
                    bucket_name, object_name = parts
                else:
                    object_name = parts[0]
                    bucket_name = self.minio_client.bucket  # Defaultbucket
            elif minio_path.startswith("/bisheng/"):
                # Format: /bisheng/object/name
                bucket_name = self.minio_client.bucket
                object_name = minio_path[9:]  # Remove '/bisheng/'
            elif minio_path.startswith("/tmp-dir/"):
                # Format: /tmp-dir/object/name
                bucket_name = self.minio_client.tmp_bucket
                object_name = minio_path[9:]  # Remove '/tmp-dir/'
            elif minio_path.startswith("/tmp/"):
                # Format: /tmp/object/name -> SmartbucketPilih
                object_name = minio_path[5:]  # Remove '/tmp/'

                # Try the Lord firstbucket
                bucket_name = self.minio_client.bucket
                if self.minio_client.object_exists_sync(bucket_name, object_name):
                    self.logger.debug(f"Schedule an immediate check of all services on this hostbucketFile Found: {bucket_name}/{object_name}")
                else:
                    # MasterbucketNo, trytmp_bucket
                    bucket_name = self.minio_client.tmp_bucket
                    if self.minio_client.object_exists_sync(bucket_name, object_name):
                        self.logger.debug(f"Insidetmp_bucketFile Found: {bucket_name}/{object_name}")
                    else:
                        self.logger.warning(f"TwobucketNo documents found: {object_name}")
            else:
                # Try as fullURL<g id="Bold">Medical Treatment:</g>
                if self._is_valid_url(minio_path):
                    return self._download_file_from_url(minio_path)
                else:
                    self.logger.error(f"Unable to AnalyzeMinIOPath: {minio_path}")
                    return "", False

            # Checks to see if file exists.
            if not self.minio_client.object_exists_sync(bucket_name, object_name):
                self.logger.error(f"MinIOFile don\'t exists: {bucket_name}/{object_name}")
                return "", False

            # Download File Contents
            file_content = self.minio_client.get_object_sync(bucket_name, object_name)

            # Generate temporary filename
            file_ext = os.path.splitext(object_name)[1] or ".dat"
            filename = f"{uuid4().hex}{file_ext}"
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, filename)

            # Save to Temporary File
            with open(temp_file, "wb") as f:
                f.write(file_content)

            return temp_file, True

        except Exception as e:
            self.logger.error(f"FROMMinIODownload failed: {minio_path}, Error-free: {str(e)}")
            return "", False

    def _is_valid_url(self, url: str) -> bool:
        """Check if it is validURL"""
        try:
            result = urlparse(url)
            return all([result.scheme, result.netloc])
        except ValueError:
            return False

    def _is_minio_path(self, path: str) -> bool:
        """Check if it is clearMinIODedicated Path"""
        # Only explicitMinIODedicated paths go straightMinIOMengunduh
        # Other pathways (incl./tmp/) all go smart local downloadspipeline
        return path.startswith(("minio://", "/bisheng/", "/tmp-dir/"))

    def cleanup(self):
        """Clean Up Temp Files"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                    self.logger.debug(f"Clean Up Temp Files: {temp_file}")
            except Exception as e:
                self.logger.warning(f"Failed to clean temporary files {temp_file}: {str(e)}")

        self.temp_files.clear()


class SmartDocumentRenderer:
    """Smart Document Renderer"""

    def __init__(self, template_content: bytes):
        self.template_content = template_content
        self.logger = logger

    def render(self, variables: Dict[str, str], resources: List[ResourceData]) -> bytes:
        """
        Render Final Document

        Args:
            variables: Processed variable dictionary (with placeholders)
            resources: All Resources List

        Returns:
            Rendered document byte stream
        """
        self.logger.info(f"Start rendering documents, variables {len(variables)} , resources {len(resources)} Pcs")

        # 1. InisialisasiDocxTemplateRender
        docx_renderer = DocxTemplateRender(file_content=io.BytesIO(self.template_content))

        # 2. Step 1: Variable replacement (including placeholders)
        template_def = [[f"{{{{{key}}}}}", value] for key, value in variables.items()]

        # 3. Build resource maps (grouped by type)
        resource_map = self._build_resource_map(resources)

        # 4. RecallDocxTemplateRenderPerform Rendering
        output_doc = docx_renderer.render(template_def, resource_map)

        # 5. Convert to byte stream
        output_content = io.BytesIO()
        output_doc.save(output_content)
        output_content.seek(0)

        self.logger.info("Document rendering complete")

        return output_content.read()

    def _build_resource_map(self, resources: List[ResourceData]) -> Dict[str, List[Dict]]:
        """Build resource mappings, adaptingDocxTemplateRenderformat; """
        resource_map = {"images": [], "excel_files": [], "csv_files": [], "markdown_tables": []}

        for resource in resources:
            if resource.resource_type == ResourceType.IMAGE:
                image_info = {
                    "placeholder": resource.placeholder,
                    "local_path": resource.local_path if resource.download_success else resource.original_path,
                    "alt_text": resource.alt_text,
                    "original_path": resource.original_path,
                    "type": "downloaded" if resource.download_success else "failed",
                    "original_text": resource.original_content,
                }
                resource_map["images"].append(image_info)

            elif resource.resource_type == ResourceType.TABLE:
                if resource.table_source == TableSource.MARKDOWN_TABLE:
                    # MarkdownTable incsv_filesMedium Processing
                    table_info = {
                        "placeholder": resource.placeholder,
                        "table_data": resource.table_data,
                        "alignments": resource.alignments,
                        "file_name": resource.file_name,
                        "type": "markdown_table",  # Fix: Use correct type identifier
                    }
                    resource_map["csv_files"].append(table_info)

                elif resource.table_source == TableSource.CSV_CONTENT:
                    # CSVDoc.
                    table_info = {
                        "placeholder": resource.placeholder,
                        "table_data": resource.table_data,
                        "alignments": resource.alignments,
                        "file_name": resource.file_name,
                        "type": "csv",  # Fix: Use correct type identifier
                        # Add compatibility field
                        "local_path": getattr(resource, 'local_path', resource.original_path),  # Use downloaded local path or original path
                    }
                    resource_map["csv_files"].append(table_info)

                elif resource.table_source == TableSource.EXCEL_CONTENT:
                    # ExcelDoc.
                    table_info = {
                        "placeholder": resource.placeholder,
                        "table_data": resource.table_data,
                        "alignments": resource.alignments,
                        "file_name": resource.file_name,
                        "type": "excel",  # Fix: Use correct type identifier
                        # Tambahdocx_temp.pyExpected Fields
                        "local_path": getattr(resource, 'local_path', resource.original_path),  # Use downloaded local path or original path
                    }

                    resource_map["excel_files"].append(table_info)

        # Record Resource Statistics
        self.logger.info(
            f"Resource mapping build complete: Images {len(resource_map['images'])}, "
            f"CSV {len(resource_map['csv_files'])}, Excel {len(resource_map['excel_files'])}"
        )

        return resource_map


class ReportNode(BaseNode):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._report_info = self.node_params["report_info"]
        self._version_key = self._report_info["version_key"].split("_")[0]
        self._object_name = f"workflow/report/{self._version_key}.docx"
        self._file_name = self._report_info["file_name"] if self._report_info["file_name"] else "tmp_report.docx"
        if not self._file_name.endswith(".docx"):
            self._file_name += ".docx"
        self._minio_client = get_minio_storage_sync()

    def _run(self, unique_id: str):
        """Master Execution Process"""
        download_manager = None

        try:
            # 1. Download sample
            logger.info("=== Walking Tongs1: Download report template ===")
            template_content = self._download_template()

            # 2. Resolve Template Variables
            logger.info("=== Walking Tongs2: Resolve Template Variables ===")
            template_variables = self._extract_template_variables(template_content)

            # 3. Get workflow variables
            logger.info("=== Walking Tongs3: Get workflow variables ===")
            workflow_variables = self._get_filtered_workflow_variables(template_variables)

            # 4. Parsing all variable content
            logger.info("=== Walking Tongs4: Parse variable contents ===")
            content_parser = ContentParser(self._minio_client)
            processed_variables = {}
            all_resources = []

            for var_name, var_value in workflow_variables.items():
                # Add Detailed Variable Value Log
                logger.info(f"[Variable parsing] Variables: '{var_name}'")
                logger.info(f"[Variable parsing] Variable type: {type(var_value).__name__}")
                logger.info(f"[Variable parsing] Variable Value Length: {len(str(var_value)) if var_value is not None else 0}")

                # Print variable value contents (before interception500Character Avoidance Log Too Long)
                if var_value is not None:
                    var_value_str = str(var_value)
                    if len(var_value_str) > 500:
                        logger.info(f"[Variable parsing] Variable Value Content (ex.500characters. : {var_value_str[:500]}...")
                    else:
                        logger.info(f"[Variable parsing] Variable Value Content: {var_value_str}")
                else:
                    logger.info(f"[Variable parsing] Variable Value Content: None")

                processed_content, resources = content_parser.parse_variable_content(var_name, var_value)
                processed_variables[var_name] = processed_content
                all_resources.extend(resources)

                # Add Parsing Result Log
                logger.info(f"[Variable parsing] Post-Parse Content Length: {len(processed_content) if processed_content else 0}")
                logger.info(f"[Variable parsing] Identify the number of resources: {len(resources)}")
                if resources:
                    for i, resource in enumerate(resources):
                        logger.info(
                            f"[Variable parsing] reasourse{i + 1}: {resource.resource_type.value}, Placeholder Icon: {resource.placeholder}")
                logger.info(f"[Variable parsing] --- Variables '{var_name}' Parsing complete ---")

            # 5. Download all resources
            logger.info("=== Walking Tongs5: Download Resource File ===")
            download_manager = ResourceDownloadManager(self._minio_client)
            download_stats = download_manager.download_all_resources(all_resources)

            self._log_download_stats(download_stats)

            # 6. Render Document
            logger.info("=== Walking Tongs6: renderedWordDocumentation ===")
            renderer = SmartDocumentRenderer(template_content)
            final_document = renderer.render(processed_variables, all_resources)

            # 7. SAVE AND SHARE
            logger.info("=== Walking Tongs7: Save and share documents ===")
            share_url = self._save_and_share_document(final_document)

            # 8. Send Output Message
            self._send_output_message(unique_id, share_url)

            logger.info("=== Report Generation Complete ===")

        except Exception as e:
            logger.error(f"Report generation failed: {str(e)}")
            self._send_error_message(unique_id, str(e))
            raise

        finally:
            # Temporarily disable cleanup to avoid image file deletionWordDocument display
            # if download_manager:
            #     download_manager.cleanup()
            pass

    def _download_template(self) -> bytes:
        """Download sample"""
        if not self._minio_client.object_exists_sync(self._minio_client.bucket, self._object_name):
            raise Exception(f"Template file does not exists!: {self._object_name}")

        template_content = self._minio_client.get_object_sync(self._minio_client.bucket, self._object_name)
        logger.info(f"Template downloaded successfully, size: {len(template_content)} byte")

        return template_content

    def _get_filtered_workflow_variables(self, template_variables: set) -> Dict[str, Any]:
        """Get filtered workflow variables"""
        all_variables = self.graph_state.get_all_variables()

        # Add Detailed Variable Information Log
        logger.info(f"[Process Variables] Total: {len(all_variables)}")
        logger.info(f"[Process Variables] All Variable Names: {list(all_variables.keys())}")

        # Show basic information for each variable
        for var_name, var_value in all_variables.items():
            var_type = type(var_value).__name__
            var_length = len(str(var_value)) if var_value is not None else 0
            logger.info(f"[Process Variables] '{var_name}': {var_type}, Longitudinal={var_length}")

        # Keep only the variables actually used in the template
        filtered_variables = {k: v for k, v in all_variables.items() if k in template_variables}

        logger.info(f"[Process Variables] Variables required by the template: {list(template_variables)}")
        logger.info(f"[Process Variables] Number of filtered variables: {len(filtered_variables)}")
        logger.info(f"[Process Variables] Filtered Variable Names: {list(filtered_variables.keys())}")

        # Record filtered variables (for debugging)
        excluded_vars = set(all_variables.keys()) - set(filtered_variables.keys())
        if excluded_vars:
            logger.info(f"[Process Variables] Filtered Variables: {list(excluded_vars)}")

        return filtered_variables

    def _log_download_stats(self, stats: Dict[str, Any]):
        """Record download statistics"""
        logger.info("Resource Download Stats:")
        logger.info(f"  - Image Success: {stats['images_success']}")
        logger.info(f"  - Image failed: {stats['images_failed']}")
        logger.info(f"  - Form Processing: {stats['tables_processed']}")

        if stats.get("errors"):
            logger.warning("Download error details:")
            for error in stats["errors"]:
                logger.warning(f"  - {error}")

    def _save_and_share_document(self, document_content: bytes) -> str:
        """Save the document and get a share link"""
        # Generate unique file path
        tmp_object_name = f"workflow/report/{uuid4().hex}/{self._file_name}"

        # Uploaded toMinIO
        self._minio_client.put_object_tmp_sync(tmp_object_name, document_content)

        # Get share link
        share_url = self._minio_client.get_share_link_sync(tmp_object_name, self._minio_client.tmp_bucket)

        logger.info(f"Document saved successfully: {tmp_object_name}")
        logger.info(f"Share Links: {share_url}")

        return share_url

    def _send_output_message(self, unique_id: str, share_url: str):
        """Send Output Message"""
        self.callback_manager.on_output_msg(
            OutputMsgData(
                unique_id=unique_id,
                node_id=self.id,
                name=self.name,
                msg="",
                files=[{"path": share_url, "name": self._file_name}],
                output_key="",
            )
        )

    def _send_error_message(self, unique_id: str, error_msg: str):
        """Send error message"""
        self.callback_manager.on_output_msg(
            OutputMsgData(
                unique_id=unique_id,
                node_id=self.id,
                name=self.name,
                msg=f"Report generation failed: {error_msg}",
                files=[],
                output_key="",
            )
        )

    # Preserve necessary auxiliary methods
    def _get_unique_placeholder_id(self) -> int:
        """Old method retained for compatibility"""
        return 0  # No longer used in the new system

    def _extract_template_variables(self, file_content: bytes) -> set:
        """
        FROMWordExtract all variable placeholders in the template file

        Args:
            file_content: WordThe binary content of the file

        Returns:
            set: Collection of variable names referenced in the template
        """
        import zipfile
        import re

        try:
            # WordThe document is azipfiles, parsing theXMLContents
            template_variables = set()

            with zipfile.ZipFile(io.BytesIO(file_content), "r") as docx_zip:
                # Parse Main Document Section
                if "word/document.xml" in docx_zip.namelist():
                    doc_xml = docx_zip.read("word/document.xml").decode("utf-8")

                    # Find all using regular expressions {{Variables}} Formatted Placeholder
                    pattern = r"\{\{([^}]+)\}\}"
                    matches = re.findall(pattern, doc_xml)

                    # <g id="Bold">Medical Treatment:</g>WordPossible splitting of variables into multipleXMLLabel Situation
                    # Remove all firstXMLTags, then match
                    clean_xml = re.sub(r"<[^>]+>", "", doc_xml)
                    clean_matches = re.findall(pattern, clean_xml)

                    # Merge two matches
                    all_matches = list(set(matches + clean_matches))

                    # Add Detailed Debug Log
                    logger.warning(f"OriginalXMLNo Matching Results: {matches}")
                    logger.warning(f"CleanedXMLPost-Match Results: {clean_matches}")
                    logger.warning(f"Merged Match Result: {all_matches}")

                    for match in all_matches:
                        # Clean up variable names (remove spaces, etc.)
                        var_name = match.strip()
                        if var_name:
                            template_variables.add(var_name)
                            logger.debug(f"Template variables found: {var_name}")

                # Also check the header footer and other sections
                for xml_part in ["word/header1.xml", "word/footer1.xml"]:
                    if xml_part in docx_zip.namelist():
                        xml_content = docx_zip.read(xml_part).decode("utf-8")
                        matches = re.findall(r"\{\{([^}]+)\}\}", xml_content)
                        for match in matches:
                            var_name = match.strip()
                            if var_name:
                                template_variables.add(var_name)
                                logger.debug(f"Template variables found(Header/Pre Footer): {var_name}")

            logger.info(f"[Template resolution] Extract to {len(template_variables)} variables: {list(template_variables)}")
            for i, var_name in enumerate(sorted(template_variables), 1):
                logger.info(f"[Template resolution] Variables{i}: '{var_name}'")
            return template_variables

        except Exception as e:
            logger.error(f"Failed to parse template variable: {e}")
            # If parsing fails, return an empty collection so that no variables are filtered
            return set()
