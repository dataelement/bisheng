from __future__ import annotations

from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from bisheng.common.errcode.portal_course import (
    PortalCourseCatalogDepthExceededError,
    PortalCourseCatalogImportError,
    PortalCourseCatalogInUseError,
    PortalCourseCatalogNameDuplicateError,
    PortalCourseCatalogNotFoundError,
    PortalCourseCatalogParentInvalidError,
)
from bisheng.shougang_portal_course.domain.models.portal_course import PortalCourseCatalog
from bisheng.shougang_portal_course.domain.repositories.catalog_repository import (
    PortalCourseCatalogRepository,
)
from bisheng.shougang_portal_course.domain.schemas.portal_course_schema import (
    CatalogCreate,
    CatalogImportIssue,
    CatalogImportPreview,
    CatalogImportResult,
    CatalogRead,
    CatalogUpdate,
    OrderUpdate,
)

ImportRow = tuple[int, str | None, str | None, str | None, str, str, int, bool]

MAX_CATALOG_DEPTH = 8
NAME_PATH_SEPARATOR = "->"
EXCEL_HEADERS = ["目录ID", "上级目录ID", "上级目录", "目录名称", "描述", "排序", "是否公开"]
MAX_IMPORT_ROWS = 2000
MAX_CATALOG_NAME_LENGTH = 200
MAX_CATALOG_EXTERNAL_ID_LENGTH = 32


class PortalCourseCatalogService:
    """Maintains hierarchical course catalogs and Excel import."""

    def __init__(self, session):
        self.repository = PortalCourseCatalogRepository(session)

    async def create_catalog(
        self,
        *,
        tenant_id: int,
        user_id: int,
        payload: CatalogCreate,
    ) -> PortalCourseCatalog:
        parent = await self._require_parent(
            tenant_id=tenant_id,
            parent_id=payload.parent_id,
        )
        await self._ensure_unique_name(
            tenant_id=tenant_id,
            parent_id=payload.parent_id,
            name=payload.name,
        )
        self._ensure_depth(parent)
        catalog = PortalCourseCatalog(
            tenant_id=tenant_id,
            name=payload.name,
            description=payload.description,
            parent_id=payload.parent_id,
            routing_path="0",
            catalog_id_path="",
            catalog_name_path="",
            order_index=payload.order_index,
            opened=payload.opened,
            create_user=user_id,
            update_user=user_id,
        )
        await self.repository.add(catalog)
        catalog.routing_path = await self._next_routing_path(
            tenant_id=tenant_id,
            parent=parent,
            exclude_id=catalog.id,
        )
        self._apply_paths(catalog, parent)
        await self.repository.add(catalog)
        return catalog

    async def update_catalog(
        self,
        *,
        tenant_id: int,
        user_id: int,
        catalog_id: str,
        payload: CatalogUpdate,
    ) -> PortalCourseCatalog:
        catalog = await self.repository.get_catalog(
            tenant_id=tenant_id,
            catalog_id=catalog_id,
            for_update=True,
        )
        if catalog is None:
            raise PortalCourseCatalogNotFoundError()

        fields = payload.model_fields_set
        new_name = payload.name if "name" in fields else catalog.name
        new_parent_id = payload.parent_id if "parent_id" in fields else catalog.parent_id
        parent_changed = new_parent_id != catalog.parent_id
        name_changed = new_name != catalog.name

        if parent_changed or name_changed:
            await self._ensure_unique_name(
                tenant_id=tenant_id,
                parent_id=new_parent_id,
                name=new_name,
                exclude_id=catalog.id,
            )

        parent_model = None
        if parent_changed or name_changed:
            parent_model = await self._require_parent(
                tenant_id=tenant_id,
                parent_id=new_parent_id,
            )
            if parent_changed:
                descendants = await self.repository.list_descendants(
                    tenant_id=tenant_id,
                    catalog=catalog,
                    include_deleted=True,
                )
                self._ensure_not_descendant(catalog, parent_model)
                self._ensure_depth(
                    parent_model,
                    extra_depth=self._subtree_span(catalog, descendants),
                )

        if "name" in fields:
            catalog.name = payload.name
        if "description" in fields:
            catalog.description = payload.description
        if "order_index" in fields:
            catalog.order_index = payload.order_index
        if "opened" in fields:
            catalog.opened = payload.opened
        if parent_changed:
            catalog.parent_id = new_parent_id
            catalog.routing_path = await self._next_routing_path(
                tenant_id=tenant_id,
                parent=parent_model,
                exclude_id=catalog.id,
            )

        if parent_changed or name_changed:
            await self._rebuild_paths(
                tenant_id=tenant_id,
                catalog=catalog,
                parent=parent_model,
            )

        catalog.update_user = user_id
        catalog.update_time = datetime.now()
        await self.repository.add(catalog)
        return catalog

    async def delete_catalog(self, *, tenant_id: int, catalog_id: str) -> None:
        catalog = await self.repository.get_catalog(
            tenant_id=tenant_id,
            catalog_id=catalog_id,
            for_update=True,
        )
        if catalog is None:
            raise PortalCourseCatalogNotFoundError()
        child_count = await self.repository.count_children(
            tenant_id=tenant_id,
            parent_id=catalog.id,
        )
        course_count = await self.repository.count_courses(
            tenant_id=tenant_id,
            catalog_id=catalog.id,
        )
        if child_count or course_count:
            raise PortalCourseCatalogInUseError()
        catalog.deleted = True
        catalog.update_time = datetime.now()
        await self.repository.add(catalog)

    async def update_catalog_order(self, *, tenant_id: int, payload: OrderUpdate) -> None:
        for item in payload.items:
            catalog = await self.repository.get_catalog(
                tenant_id=tenant_id,
                catalog_id=item.id,
                for_update=True,
            )
            if catalog is None:
                raise PortalCourseCatalogNotFoundError()
            catalog.order_index = item.sort_order
            catalog.update_time = datetime.now()
            await self.repository.add(catalog)

    async def list_read_models(
        self,
        *,
        tenant_id: int,
        public_only: bool,
        include_deleted: bool = False,
        as_tree: bool = True,
    ) -> list[CatalogRead]:
        catalogs = await self.repository.list_catalogs(
            tenant_id=tenant_id,
            include_deleted=include_deleted,
            opened_only=public_only,
        )
        if public_only:
            catalogs = self._visible_public_nodes(catalogs)
        counts = await self.repository.count_courses_by_ids(
            tenant_id=tenant_id,
            catalog_ids=[item.id for item in catalogs],
        )
        reads = [self._to_read(item, course_count=counts.get(item.id, 0)) for item in catalogs]
        if as_tree:
            return self._to_tree(reads)
        return reads

    async def list_course_list_nav(
        self,
        *,
        tenant_id: int,
        keyword: str | None = None,
    ) -> list[CatalogRead]:
        catalogs = await self.repository.list_catalogs(tenant_id=tenant_id)
        enabled_counts = await self.repository.count_courses_by_ids(
            tenant_id=tenant_id,
            catalog_ids=[item.id for item in catalogs],
            enabled_only=True,
        )
        visible = [
            item
            for item in catalogs
            if item.opened or enabled_counts.get(item.id, 0) > 0
        ]
        catalogs = self._with_ancestors(catalogs, {item.id for item in visible})
        needle = (keyword or "").strip()
        if needle:
            matched = {
                item.id
                for item in catalogs
                if self._catalog_matches_keyword(item, needle)
            }
            catalogs = self._with_ancestors(catalogs, matched)
        reads = [
            self._to_read(item, course_count=enabled_counts.get(item.id, 0))
            for item in catalogs
        ]
        return self._to_tree(reads)

    async def get_read_model(
        self,
        *,
        tenant_id: int,
        catalog_id: str,
        public_only: bool = False,
    ) -> CatalogRead:
        catalog = await self.repository.get_catalog(
            tenant_id=tenant_id,
            catalog_id=catalog_id,
        )
        if catalog is None:
            raise PortalCourseCatalogNotFoundError()
        if public_only:
            catalogs = await self.repository.list_catalogs(
                tenant_id=tenant_id,
                opened_only=True,
            )
            visible_ids = {item.id for item in self._visible_public_nodes(catalogs)}
            if catalog.id not in visible_ids:
                raise PortalCourseCatalogNotFoundError()
        course_count = await self.repository.count_courses(
            tenant_id=tenant_id,
            catalog_id=catalog.id,
        )
        return self._to_read(catalog, course_count=course_count)

    async def resolve_catalog_ids(
        self,
        *,
        tenant_id: int,
        catalog_id: str,
        include_descendants: bool,
        public_only: bool,
    ) -> list[str]:
        catalog = await self.repository.get_catalog(
            tenant_id=tenant_id,
            catalog_id=catalog_id,
        )
        if catalog is None:
            raise PortalCourseCatalogNotFoundError()
        if public_only and not catalog.opened:
            raise PortalCourseCatalogNotFoundError()
        ids = [catalog.id]
        if include_descendants:
            descendants = await self.repository.list_descendants(
                tenant_id=tenant_id,
                catalog=catalog,
            )
            if public_only:
                descendants = [item for item in descendants if item.opened]
            ids.extend(item.id for item in descendants)
        return ids

    def build_template(self) -> bytes:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "课程目录"
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        examples = [
            ("CAT-001", "", "", "安全生产", "公司级安全培训", 1, "是"),
            ("CAT-002", "CAT-001", "安全生产", "消防安全", "消防专题", 1, "是"),
            ("CAT-003", "CAT-002", "消防安全", "灭火器使用", "实操课程", 1, "是"),
        ]
        for row_idx, row in enumerate(examples, start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        sheet.column_dimensions["A"].width = 36
        sheet.column_dimensions["B"].width = 36
        sheet.column_dimensions["C"].width = 24
        sheet.column_dimensions["D"].width = 24
        sheet.column_dimensions["E"].width = 30
        sheet.column_dimensions["F"].width = 10
        sheet.column_dimensions["G"].width = 12
        self._write_instruction_sheet(workbook.create_sheet("填写说明"))
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def preview_excel(
        self,
        *,
        tenant_id: int,
        content: bytes,
    ) -> CatalogImportPreview:
        parsed_rows, parse_issues, total = self._parse_import_workbook(content)
        existing = await self.repository.list_catalogs(tenant_id=tenant_id)
        id_index = {item.id: item for item in existing}
        name_index: dict[str, list[PortalCourseCatalog]] = {}
        for item in existing:
            name_index.setdefault(item.name, []).append(item)
        ids_in_file = {catalog_id for _, catalog_id, _, _, _, _, _, _ in parsed_rows if catalog_id}
        names_in_file = {name for _, _, _, _, name, _, _, _ in parsed_rows}
        issues = list(parse_issues)
        valid = 0
        for row_idx, _catalog_id, parent_id, parent_name, _name, _description, _order_index, _opened in parsed_rows:
            row_issues = self._preview_parent_issues(
                row_idx=row_idx,
                parent_id=parent_id,
                parent_name=parent_name,
                id_index=id_index,
                name_index=name_index,
                ids_in_file=ids_in_file,
                names_in_file=names_in_file,
            )
            if row_issues:
                issues.extend(row_issues)
            else:
                valid += 1
        return CatalogImportPreview(total=total, valid=valid, issues=issues)

    async def import_excel(
        self,
        *,
        tenant_id: int,
        user_id: int,
        content: bytes,
        force: bool = False,
    ) -> CatalogImportResult:
        parsed_rows, parse_issues, total = self._parse_import_workbook(content)
        errors = [issue.message for issue in parse_issues]
        existing = await self.repository.list_catalogs(tenant_id=tenant_id)
        id_index = {item.id: item for item in existing}
        name_index: dict[str, list[PortalCourseCatalog]] = {}
        siblings: dict[str | None, dict[str, PortalCourseCatalog]] = {}
        for item in existing:
            name_index.setdefault(item.name, []).append(item)
            siblings.setdefault(item.parent_id, {})[item.name] = item

        success = 0
        failed = len(errors)
        pending: list[ImportRow] = parsed_rows
        while pending:
            next_pending: list[ImportRow] = []
            progressed = False
            for row in pending:
                applied = await self._try_import_row(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    row=row,
                    id_index=id_index,
                    name_index=name_index,
                    siblings=siblings,
                    errors=errors,
                )
                if applied is True:
                    success += 1
                    progressed = True
                elif applied is False:
                    failed += 1
                else:
                    next_pending.append(row)
            if progressed:
                pending = next_pending
                continue
            if not next_pending:
                break
            if force:
                pending, forced_success, forced_failed = await self._force_missing_parents(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    rows=next_pending,
                    id_index=id_index,
                    name_index=name_index,
                    siblings=siblings,
                    errors=errors,
                )
                success += forced_success
                failed += forced_failed
                if pending:
                    continue
                break
            for row_idx, _catalog_id, parent_id, parent_name, *_ in next_pending:
                failed += 1
                if parent_id:
                    errors.append(f"第 {row_idx} 行: 上级目录ID「{parent_id}」不存在")
                else:
                    errors.append(f"第 {row_idx} 行: 上级目录「{parent_name}」不存在")
            break
        return CatalogImportResult(
            total=total,
            success=success,
            failed=failed,
            errors=errors,
        )

    async def _try_import_row(
        self,
        *,
        tenant_id: int,
        user_id: int,
        row: ImportRow,
        id_index: dict[str, PortalCourseCatalog],
        name_index: dict[str, list[PortalCourseCatalog]],
        siblings: dict[str | None, dict[str, PortalCourseCatalog]],
        errors: list[str],
    ) -> bool | None:
        row_idx, catalog_id, parent_id, parent_name, name, description, order_index, opened = row
        try:
            parent = self._resolve_import_parent(
                parent_id=parent_id,
                parent_name=parent_name,
                id_index=id_index,
                name_index=name_index,
            )
        except ValueError as exc:
            errors.append(f"第 {row_idx} 行: {exc}")
            return False
        except LookupError:
            return None
        try:
            created = await self._upsert_child(
                tenant_id=tenant_id,
                user_id=user_id,
                catalog_id=catalog_id,
                parent=parent,
                name=name,
                description=description,
                order_index=order_index,
                opened=opened,
                siblings=siblings,
            )
        except Exception as exc:
            errors.append(f"第 {row_idx} 行: {exc}")
            return False
        if created is not None:
            name_index.setdefault(name, []).append(created)
            id_index[created.id] = created
        return True

    async def _force_missing_parents(
        self,
        *,
        tenant_id: int,
        user_id: int,
        rows: list[ImportRow],
        id_index: dict[str, PortalCourseCatalog],
        name_index: dict[str, list[PortalCourseCatalog]],
        siblings: dict[str | None, dict[str, PortalCourseCatalog]],
        errors: list[str],
    ) -> tuple[list[ImportRow], int, int]:
        pending_ids = {catalog_id for _, catalog_id, _, _, _, _, _, _ in rows if catalog_id}
        pending_names = {name for _, _, _, _, name, _, _, _ in rows}
        leftover: list[ImportRow] = []
        success = 0
        failed = 0
        for row in rows:
            row_idx, catalog_id, parent_id, parent_name, name, description, order_index, opened = row
            waiting_for_parent = (parent_id is not None and parent_id in pending_ids) or (
                parent_id is None and parent_name is not None and parent_name in pending_names
            )
            if waiting_for_parent:
                leftover.append(row)
                continue
            try:
                created = await self._upsert_child(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    catalog_id=catalog_id,
                    parent=None,
                    name=name,
                    description=description,
                    order_index=order_index,
                    opened=opened,
                    siblings=siblings,
                )
            except Exception as exc:
                failed += 1
                errors.append(f"第 {row_idx} 行: {exc}")
                continue
            if created is not None:
                name_index.setdefault(name, []).append(created)
                id_index[created.id] = created
            success += 1
        if success:
            return leftover, success, failed
        for row_idx, catalog_id, _parent_id, _parent_name, name, description, order_index, opened in leftover:
            try:
                created = await self._upsert_child(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    catalog_id=catalog_id,
                    parent=None,
                    name=name,
                    description=description,
                    order_index=order_index,
                    opened=opened,
                    siblings=siblings,
                )
            except Exception as exc:
                failed += 1
                errors.append(f"第 {row_idx} 行: {exc}")
                continue
            if created is not None:
                name_index.setdefault(name, []).append(created)
                id_index[created.id] = created
            success += 1
        return [], success, failed

    def _parse_import_workbook(
        self,
        content: bytes,
    ) -> tuple[list[ImportRow], list[CatalogImportIssue], int]:
        try:
            workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:
            raise PortalCourseCatalogImportError(msg="无法解析 Excel 文件") from exc
        sheet = workbook.active
        if sheet is None:
            raise PortalCourseCatalogImportError(msg="Excel 文件没有工作表")
        header_row = next(sheet.iter_rows(values_only=True), None)
        if header_row is None:
            raise PortalCourseCatalogImportError(msg="Excel 表头为空")
        headers = [str(cell).strip() if cell is not None else "" for cell in header_row[: len(EXCEL_HEADERS)]]
        if headers != EXCEL_HEADERS:
            raise PortalCourseCatalogImportError(msg=f"表头必须为 {' / '.join(EXCEL_HEADERS)}")

        parsed_rows: list[ImportRow] = []
        issues: list[CatalogImportIssue] = []
        total = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None or str(cell).strip() == "" for cell in (row or ())):
                continue
            total += 1
            if total > MAX_IMPORT_ROWS:
                raise PortalCourseCatalogImportError(msg=f"导入行数不能超过 {MAX_IMPORT_ROWS}")
            try:
                parsed_rows.append(self._parse_import_row(row_idx, row))
            except ValueError as exc:
                issues.append(
                    CatalogImportIssue(
                        row=row_idx,
                        code="parse_error",
                        message=str(exc),
                        recoverable=False,
                    )
                )
        return parsed_rows, issues, total

    @staticmethod
    def _preview_parent_issues(
        *,
        row_idx: int,
        parent_id: str | None,
        parent_name: str | None,
        id_index: dict[str, PortalCourseCatalog],
        name_index: dict[str, list[PortalCourseCatalog]],
        ids_in_file: set[str],
        names_in_file: set[str],
    ) -> list[CatalogImportIssue]:
        if parent_id:
            parent = id_index.get(parent_id)
            if parent is None and parent_id not in ids_in_file:
                return [
                    CatalogImportIssue(
                        row=row_idx,
                        code="missing_parent",
                        message=(
                            f"第 {row_idx} 行: 上级目录ID「{parent_id}」不存在，"
                            "强行导入将挂到根目录"
                        ),
                        recoverable=True,
                    )
                ]
            if parent is not None:
                try:
                    PortalCourseCatalogService._ensure_depth(parent)
                except PortalCourseCatalogDepthExceededError:
                    return [
                        CatalogImportIssue(
                            row=row_idx,
                            code="depth_exceeded",
                            message=f"第 {row_idx} 行: 层级不能超过 {MAX_CATALOG_DEPTH} 级",
                            recoverable=False,
                        )
                    ]
            return []
        if parent_name is None:
            return []
        matches = name_index.get(parent_name, [])
        if len(matches) > 1:
            return [
                CatalogImportIssue(
                    row=row_idx,
                    code="ambiguous_parent",
                    message=f"第 {row_idx} 行: 上级目录「{parent_name}」不唯一",
                    recoverable=False,
                )
            ]
        if not matches and parent_name not in names_in_file:
            return [
                CatalogImportIssue(
                    row=row_idx,
                    code="missing_parent",
                    message=(
                        f"第 {row_idx} 行: 上级目录「{parent_name}」不存在，"
                        "强行导入将挂到根目录"
                    ),
                    recoverable=True,
                )
            ]
        if len(matches) == 1:
            try:
                PortalCourseCatalogService._ensure_depth(matches[0])
            except PortalCourseCatalogDepthExceededError:
                return [
                    CatalogImportIssue(
                        row=row_idx,
                        code="depth_exceeded",
                        message=f"第 {row_idx} 行: 层级不能超过 {MAX_CATALOG_DEPTH} 级",
                        recoverable=False,
                    )
                ]
        return []

    async def _upsert_child(
        self,
        *,
        tenant_id: int,
        user_id: int,
        catalog_id: str | None,
        parent: PortalCourseCatalog | None,
        name: str,
        description: str,
        order_index: int,
        opened: bool,
        siblings: dict[str | None, dict[str, PortalCourseCatalog]],
    ) -> PortalCourseCatalog | None:
        parent_key = parent.id if parent else None
        existing = None
        if catalog_id:
            existing = await self.repository.get_catalog(
                tenant_id=tenant_id,
                catalog_id=catalog_id,
            )
        else:
            existing = siblings.setdefault(parent_key, {}).get(name)
        if existing is None:
            self._ensure_depth(parent)
            await self._ensure_unique_name(
                tenant_id=tenant_id,
                parent_id=parent_key,
                name=name,
            )
            fields = {
                "tenant_id": tenant_id,
                "name": name,
                "description": description,
                "parent_id": parent_key,
                "routing_path": "0",
                "catalog_id_path": "",
                "catalog_name_path": "",
                "order_index": order_index,
                "opened": opened,
                "create_user": user_id,
                "update_user": user_id,
            }
            catalog = (
                PortalCourseCatalog(id=catalog_id, **fields)
                if catalog_id
                else PortalCourseCatalog(**fields)
            )
            await self.repository.add(catalog)
            catalog.routing_path = await self._next_routing_path(
                tenant_id=tenant_id,
                parent=parent,
                exclude_id=catalog.id,
            )
            self._apply_paths(catalog, parent)
            await self.repository.add(catalog)
            siblings.setdefault(parent_key, {})[name] = catalog
            return catalog
        old_parent = existing.parent_id
        old_name = existing.name
        name_changed = old_name != name
        parent_changed = old_parent != parent_key
        if name_changed or parent_changed:
            await self._ensure_unique_name(
                tenant_id=tenant_id,
                parent_id=parent_key,
                name=name,
                exclude_id=existing.id,
            )
        if parent_changed:
            descendants = await self.repository.list_descendants(
                tenant_id=tenant_id,
                catalog=existing,
                include_deleted=True,
            )
            self._ensure_not_descendant(existing, parent)
            self._ensure_depth(parent, extra_depth=self._subtree_span(existing, descendants))
            existing.parent_id = parent_key
            existing.routing_path = await self._next_routing_path(
                tenant_id=tenant_id,
                parent=parent,
                exclude_id=existing.id,
            )
        if name_changed:
            existing.name = name
        if parent_changed or name_changed:
            await self._rebuild_paths(
                tenant_id=tenant_id,
                catalog=existing,
                parent=parent,
            )
        existing.description = description
        existing.order_index = order_index
        existing.opened = opened
        existing.update_user = user_id
        existing.update_time = datetime.now()
        await self.repository.add(existing)
        siblings.get(old_parent, {}).pop(old_name, None)
        siblings.setdefault(parent_key, {})[name] = existing
        return None

    @staticmethod
    def _resolve_import_parent(
        *,
        parent_id: str | None,
        parent_name: str | None,
        id_index: dict[str, PortalCourseCatalog],
        name_index: dict[str, list[PortalCourseCatalog]],
    ) -> PortalCourseCatalog | None:
        if parent_id:
            parent = id_index.get(parent_id)
            if parent is None:
                raise LookupError(parent_id)
            return parent
        if parent_name is None:
            return None
        matches = name_index.get(parent_name, [])
        if len(matches) > 1:
            raise ValueError(f"上级目录「{parent_name}」不唯一")
        if not matches:
            raise LookupError(parent_name)
        return matches[0]

    @staticmethod
    def _parse_import_row(
        row_idx: int, row: tuple
    ) -> ImportRow:
        values = list(row or ())
        while len(values) < 7:
            values.append(None)
        catalog_id = PortalCourseCatalogService._parse_optional_catalog_id(
            row_idx, values[0], field="目录ID"
        )
        parent_id = PortalCourseCatalogService._parse_optional_catalog_id(
            row_idx, values[1], field="上级目录ID"
        )
        parent_name = PortalCourseCatalogService._normalize_import_name(
            row_idx, values[2], field="上级目录", required=False
        )
        name = PortalCourseCatalogService._normalize_import_name(
            row_idx, values[3], field="目录名称", required=True
        )
        assert name is not None
        description = str(values[4]).strip() if values[4] is not None else ""
        if len(description) > 200:
            raise ValueError(f"第 {row_idx} 行: 描述不能超过 200 字")
        try:
            order_index = int(values[5]) if values[5] is not None and str(values[5]).strip() != "" else 0
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第 {row_idx} 行: 排序必须是整数") from exc
        opened = PortalCourseCatalogService._parse_opened(values[6])
        if catalog_id and parent_id and catalog_id == parent_id:
            raise ValueError(f"第 {row_idx} 行: 上级目录ID不能与目录ID相同")
        return row_idx, catalog_id, parent_id, parent_name, name, description, order_index, opened

    @staticmethod
    def _parse_optional_catalog_id(
        row_idx: int,
        value: object,
        *,
        field: str = "目录ID",
    ) -> str | None:
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            raise ValueError(f"第 {row_idx} 行: {field}格式无效")
        if isinstance(value, int):
            text = str(value)
        elif isinstance(value, float):
            if not value.is_integer():
                raise ValueError(f"第 {row_idx} 行: {field}格式无效")
            text = str(int(value))
        else:
            text = str(value).strip()
        if not text:
            return None
        if len(text) > MAX_CATALOG_EXTERNAL_ID_LENGTH:
            raise ValueError(
                f"第 {row_idx} 行: {field}不能超过 {MAX_CATALOG_EXTERNAL_ID_LENGTH} 个字符"
            )
        return text

    @staticmethod
    def _normalize_import_name(
        row_idx: int,
        value: object,
        *,
        field: str,
        required: bool,
    ) -> str | None:
        name = str(value).strip() if value is not None else ""
        if not name:
            if required:
                raise ValueError(f"第 {row_idx} 行: {field}不能为空")
            return None
        if len(name) > MAX_CATALOG_NAME_LENGTH:
            raise ValueError(f"第 {row_idx} 行: {field}不能超过 {MAX_CATALOG_NAME_LENGTH} 字")
        if "/" in name or NAME_PATH_SEPARATOR in name:
            raise ValueError(f"第 {row_idx} 行: {field}不能包含路径分隔符")
        return name

    @staticmethod
    def _parse_opened(value: object) -> bool:
        if isinstance(value, bool):
            return value
        text = str(value).strip() if value is not None else "是"
        if text == "":
            return True
        return text in {"是", "1", "true", "True", "Y", "y"}

    @staticmethod
    def _write_instruction_sheet(sheet: Worksheet) -> None:
        lines = [
            "1. 请使用「课程目录」工作表导入. 表头不可修改.",
            "2. 「目录ID」选填，填写外部系统目录编号. 填写后重复导入会按该 ID 更新；留空则按「同一上级 + 目录名称」更新或新建.",
            "3. 「上级目录ID」选填，填写直接上级的外部系统目录编号，优先于「上级目录」名称.",
            "4. 「上级目录」填写直接上级的目录名称. 上级目录ID 和上级目录都留空则导入到根目录.",
            "5. 不要填写路径. 例如上级填 安全生产, 目录名称填 消防安全.",
            "6. 上级目录必须已存在, 或在本文件中有对应的目录行. 预检失败时可选择强行导入, 不存在的上级将改为根目录.",
            "7. 同一上级下同名目录重复导入时会更新描述、排序和是否公开.",
            "8. 是否公开填写 是 / 否.",
            "9. 目录名称和上级目录不能包含 / 或 ->.",
            f"10. 最多 {MAX_CATALOG_DEPTH} 级. 单次最多导入 {MAX_IMPORT_ROWS} 行.",
        ]
        sheet.column_dimensions["A"].width = 80
        for idx, line in enumerate(lines, start=1):
            sheet.cell(row=idx, column=1, value=line)

    async def _require_parent(
        self,
        *,
        tenant_id: int,
        parent_id: str | None,
    ) -> PortalCourseCatalog | None:
        if parent_id is None:
            return None
        parent = await self.repository.get_catalog(
            tenant_id=tenant_id,
            catalog_id=parent_id,
        )
        if parent is None:
            raise PortalCourseCatalogParentInvalidError()
        return parent

    async def _ensure_unique_name(
        self,
        *,
        tenant_id: int,
        parent_id: str | None,
        name: str,
        exclude_id: str | None = None,
    ) -> None:
        existing = await self.repository.find_sibling_by_name(
            tenant_id=tenant_id,
            parent_id=parent_id,
            name=name,
            exclude_id=exclude_id,
        )
        if existing is not None:
            raise PortalCourseCatalogNameDuplicateError()

    @staticmethod
    def _ensure_depth(
        parent: PortalCourseCatalog | None,
        extra_depth: int = 1,
    ) -> None:
        parent_depth = 0 if parent is None else parent.routing_path.count(".") + 1
        if parent_depth + extra_depth > MAX_CATALOG_DEPTH:
            raise PortalCourseCatalogDepthExceededError()

    @staticmethod
    def _subtree_span(
        catalog: PortalCourseCatalog,
        descendants: list[PortalCourseCatalog],
    ) -> int:
        catalog_depth = catalog.routing_path.count(".") + 1
        max_depth = catalog_depth
        for child in descendants:
            max_depth = max(max_depth, child.routing_path.count(".") + 1)
        return max_depth - catalog_depth + 1

    @staticmethod
    def _ensure_not_descendant(
        catalog: PortalCourseCatalog,
        parent: PortalCourseCatalog | None,
    ) -> None:
        if parent is None:
            return
        if parent.id == catalog.id:
            raise PortalCourseCatalogParentInvalidError()
        prefix = f"{catalog.catalog_id_path},"
        if parent.catalog_id_path == catalog.catalog_id_path or parent.catalog_id_path.startswith(prefix):
            raise PortalCourseCatalogParentInvalidError()

    async def _next_routing_path(
        self,
        *,
        tenant_id: int,
        parent: PortalCourseCatalog | None,
        exclude_id: str | None = None,
    ) -> str:
        siblings = await self.repository.list_children(
            tenant_id=tenant_id,
            parent_id=parent.id if parent else None,
            include_deleted=True,
        )
        max_segment = 0
        for sibling in siblings:
            if exclude_id and sibling.id == exclude_id:
                continue
            last = sibling.routing_path.rsplit(".", 1)[-1]
            try:
                max_segment = max(max_segment, int(last))
            except ValueError:
                continue
        segment = f"{max_segment + 1:03d}"
        if parent is None:
            return segment
        return f"{parent.routing_path}.{segment}"

    @staticmethod
    def _apply_paths(
        catalog: PortalCourseCatalog,
        parent: PortalCourseCatalog | None,
    ) -> None:
        if parent is None:
            catalog.catalog_id_path = catalog.id
            catalog.catalog_name_path = catalog.name
            return
        catalog.catalog_id_path = f"{parent.catalog_id_path},{catalog.id}"
        catalog.catalog_name_path = f"{parent.catalog_name_path}{NAME_PATH_SEPARATOR}{catalog.name}"

    async def _rebuild_paths(
        self,
        *,
        tenant_id: int,
        catalog: PortalCourseCatalog,
        parent: PortalCourseCatalog | None,
    ) -> None:
        descendants = await self.repository.list_descendants(
            tenant_id=tenant_id,
            catalog=catalog,
            include_deleted=True,
        )
        old_id_path = catalog.catalog_id_path
        old_name_path = catalog.catalog_name_path
        old_routing = catalog.routing_path
        self._apply_paths(catalog, parent)
        for child in descendants:
            child.catalog_id_path = catalog.catalog_id_path + child.catalog_id_path[len(old_id_path) :]
            child.catalog_name_path = catalog.catalog_name_path + child.catalog_name_path[len(old_name_path) :]
            child.routing_path = catalog.routing_path + child.routing_path[len(old_routing) :]
            child.update_time = datetime.now()
            await self.repository.add(child)

    @staticmethod
    def _catalog_matches_keyword(catalog: PortalCourseCatalog, keyword: str) -> bool:
        needle = keyword.casefold()
        return needle in catalog.name.casefold() or needle in catalog.catalog_name_path.casefold()

    @staticmethod
    def _with_ancestors(
        catalogs: list[PortalCourseCatalog],
        seed_ids: set[str],
    ) -> list[PortalCourseCatalog]:
        by_id = {item.id: item for item in catalogs}
        keep: set[str] = set()
        for catalog_id in seed_ids:
            current = by_id.get(catalog_id)
            while current is not None and current.id not in keep:
                keep.add(current.id)
                current = by_id.get(current.parent_id) if current.parent_id else None
        return [item for item in catalogs if item.id in keep]

    @staticmethod
    def _visible_public_nodes(
        catalogs: list[PortalCourseCatalog],
    ) -> list[PortalCourseCatalog]:
        visible = {item.id: item for item in catalogs}
        changed = True
        while changed:
            changed = False
            for catalog_id, catalog in list(visible.items()):
                if catalog.parent_id and catalog.parent_id not in visible:
                    del visible[catalog_id]
                    changed = True
        return [item for item in catalogs if item.id in visible]

    @staticmethod
    def _to_read(catalog: PortalCourseCatalog, *, course_count: int) -> CatalogRead:
        return CatalogRead(
            id=catalog.id,
            name=catalog.name,
            description=catalog.description,
            parent_id=catalog.parent_id,
            routing_path=catalog.routing_path,
            catalog_id_path=catalog.catalog_id_path,
            catalog_name_path=catalog.catalog_name_path,
            order_index=catalog.order_index,
            opened=catalog.opened,
            deleted=catalog.deleted,
            course_count=course_count,
            created_at=catalog.create_time,
            updated_at=catalog.update_time,
            create_user=catalog.create_user,
            update_user=catalog.update_user,
            children=None,
        )

    @staticmethod
    def _to_tree(items: list[CatalogRead]) -> list[CatalogRead]:
        nodes = {item.id: item.model_copy(update={"children": []}) for item in items}
        roots: list[CatalogRead] = []
        for item in items:
            node = nodes[item.id]
            if item.parent_id and item.parent_id in nodes:
                parent = nodes[item.parent_id]
                assert parent.children is not None
                parent.children.append(node)
            else:
                roots.append(node)
        return roots
