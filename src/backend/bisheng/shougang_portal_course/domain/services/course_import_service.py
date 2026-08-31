from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from bisheng.common.errcode.portal_course import PortalCourseImportError
from bisheng.shougang_portal_course.domain.models.portal_course import PortalCourseCatalog
from bisheng.shougang_portal_course.domain.repositories.catalog_repository import (
    PortalCourseCatalogRepository,
)
from bisheng.shougang_portal_course.domain.repositories.portal_course_repository import (
    PortalCourseRepository,
)
from bisheng.shougang_portal_course.domain.schemas.portal_course_schema import (
    CourseCreate,
    CourseImportIssue,
    CourseImportPreview,
    CourseImportResult,
    CourseTag,
    CourseUpdate,
    validate_media_url,
)
from bisheng.shougang_portal_course.domain.services.course_service import PortalCourseService

EXCEL_HEADERS = [
    "课程ID",
    "课程名称",
    "主讲人",
    "课程目录",
    "更新日期",
    "课程简介",
    "课程封面",
    "标签",
    "课程链接",
]
SHEET_NAME = "第三方课程"
MAX_IMPORT_ROWS = 2000
MAX_EXTERNAL_ID_LENGTH = 128
MAX_NAME_LENGTH = 200
MAX_INSTRUCTOR_LENGTH = 100
_TAG_SPLIT = re.compile(r"[,，;；、|]+")
_PATH_SEPARATORS = (". ", "->", "/", ".")

ParsedRow = tuple[
    int,
    str | None,
    str,
    str,
    str | None,
    datetime | None,
    str,
    str,
    list[CourseTag],
    str,
]


class PortalCourseImportService:
    """Excel import for third-party (external) courses."""

    def __init__(self, session):
        self.catalog_repository = PortalCourseCatalogRepository(session)
        self.course_repository = PortalCourseRepository(session)
        self.course_service = PortalCourseService(session)

    @staticmethod
    def build_template() -> bytes:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, header in enumerate(EXCEL_HEADERS, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        example = (
            "YX123456",
            "7分钟带你了解财务报告——助力业财融合",
            "赵晨露",
            "D-通用能力培训类. F-微课件. DF22-微课大赛",
            "2025/4/8",
            "围绕业财融合场景，拆解合并报表与母公司报表差异。",
            "https://example.com/cover.jpg",
            "隐患排查",
            "https://learn.example.com/course/yx123456",
        )
        for col_idx, value in enumerate(example, start=1):
            sheet.cell(row=2, column=col_idx, value=value)
        widths = [18, 42, 14, 42, 14, 48, 36, 18, 42]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + idx)].width = width
        PortalCourseImportService._write_instruction_sheet(workbook.create_sheet("填写说明"))
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def preview_excel(self, *, tenant_id: int, content: bytes) -> CourseImportPreview:
        parsed_rows, parse_issues, total = self._parse_import_workbook(content)
        catalogs = await self.catalog_repository.list_catalogs(tenant_id=tenant_id)
        issues = list(parse_issues)
        valid = 0
        for row in parsed_rows:
            row_issues = self._row_catalog_issues(row[0], row[4], catalogs)
            if row_issues:
                issues.extend(row_issues)
            else:
                valid += 1
        return CourseImportPreview(total=total, valid=valid, issues=issues)

    async def import_excel(
        self,
        *,
        tenant_id: int,
        user_id: int,
        content: bytes,
        force: bool = False,
    ) -> CourseImportResult:
        parsed_rows, parse_issues, total = self._parse_import_workbook(content)
        catalogs = await self.catalog_repository.list_catalogs(tenant_id=tenant_id)
        errors = [issue.message for issue in parse_issues]
        success = 0
        failed = len(parse_issues)
        for row in parsed_rows:
            row_idx, external_id, name, instructor, catalog_path, source_updated_at, description, cover_url, tags, external_url = row
            catalog_issues = self._row_catalog_issues(row_idx, catalog_path, catalogs)
            catalog_id = None
            if catalog_issues:
                if not force or any(not issue.recoverable for issue in catalog_issues):
                    errors.extend(issue.message for issue in catalog_issues)
                    failed += 1
                    continue
            else:
                catalog_id = self._resolve_catalog_id(catalog_path, catalogs)
            try:
                await self._upsert_row(
                    tenant_id=tenant_id,
                    user_id=user_id,
                    external_id=external_id,
                    name=name,
                    instructor=instructor,
                    catalog_id=catalog_id,
                    source_updated_at=source_updated_at,
                    description=description,
                    cover_url=cover_url,
                    tags=tags,
                    external_url=external_url,
                )
            except Exception as exc:
                errors.append(f"第 {row_idx} 行: {exc}")
                failed += 1
                continue
            success += 1
        return CourseImportResult(total=total, success=success, failed=failed, errors=errors)

    async def _upsert_row(
        self,
        *,
        tenant_id: int,
        user_id: int,
        external_id: str | None,
        name: str,
        instructor: str,
        catalog_id: str | None,
        source_updated_at: datetime | None,
        description: str,
        cover_url: str,
        tags: list[CourseTag],
        external_url: str,
    ) -> None:
        enabled = bool(external_url)
        existing = None
        if external_id:
            existing = await self.course_repository.get_course_by_external_id(
                tenant_id=tenant_id,
                external_id=external_id,
                for_update=True,
            )
        if existing is None:
            await self.course_service.create_course(
                tenant_id=tenant_id,
                user_id=user_id,
                payload=CourseCreate(
                    name=name,
                    instructor=instructor,
                    description=description,
                    tags=tags,
                    enabled=enabled,
                    catalog_id=catalog_id,
                    course_type="external",
                    external_url=external_url,
                    external_id=external_id,
                    cover_url=cover_url,
                    source_updated_at=source_updated_at,
                ),
            )
            return
        payload = {
            "name": name,
            "instructor": instructor,
            "description": description,
            "tags": [tag.model_dump() for tag in tags],
            "enabled": enabled,
            "catalog_id": catalog_id,
            "course_type": "external",
            "external_url": external_url,
            "external_id": external_id,
            "cover_url": cover_url,
            "source_updated_at": source_updated_at,
        }
        await self.course_service.update_course(
            tenant_id=tenant_id,
            course_id=existing.id,
            payload=CourseUpdate.model_validate(payload),
        )

    @staticmethod
    def _row_catalog_issues(
        row_idx: int,
        catalog_path: str | None,
        catalogs: list[PortalCourseCatalog],
    ) -> list[CourseImportIssue]:
        if not catalog_path:
            return []
        try:
            PortalCourseImportService._resolve_catalog_id(catalog_path, catalogs)
        except LookupError:
            return [
                CourseImportIssue(
                    row=row_idx,
                    code="missing_catalog",
                    message=f"第 {row_idx} 行: 课程目录「{catalog_path}」不存在",
                    recoverable=True,
                )
            ]
        except ValueError as exc:
            return [
                CourseImportIssue(
                    row=row_idx,
                    code="ambiguous_catalog",
                    message=f"第 {row_idx} 行: {exc}",
                    recoverable=False,
                )
            ]
        return []

    @staticmethod
    def _resolve_catalog_id(
        catalog_path: str | None,
        catalogs: list[PortalCourseCatalog],
    ) -> str | None:
        if not catalog_path:
            return None
        parts = PortalCourseImportService.split_catalog_path(catalog_path)
        parent_id = None
        current = None
        for name in parts:
            matches = [
                item
                for item in catalogs
                if item.parent_id == parent_id and item.name == name
            ]
            if len(matches) > 1:
                raise ValueError(f"课程目录「{name}」不唯一")
            if not matches:
                current = None
                break
            current = matches[0]
            parent_id = current.id
        if current is not None:
            return current.id
        if len(parts) == 1:
            name_matches = [item for item in catalogs if item.name == parts[0]]
            if len(name_matches) == 1:
                return name_matches[0].id
            if len(name_matches) > 1:
                raise ValueError(f"课程目录「{parts[0]}」不唯一")
        normalized = "->".join(parts)
        path_matches = [item for item in catalogs if item.catalog_name_path == normalized]
        if len(path_matches) == 1:
            return path_matches[0].id
        if len(path_matches) > 1:
            raise ValueError(f"课程目录「{catalog_path}」不唯一")
        raise LookupError(catalog_path)

    @staticmethod
    def split_catalog_path(value: str) -> list[str]:
        text = value.strip()
        if not text:
            return []
        for separator in _PATH_SEPARATORS:
            if separator in text:
                parts = [part.strip() for part in text.split(separator) if part.strip()]
                if len(parts) > 1:
                    return parts
        return [text]

    @classmethod
    def _parse_import_workbook(cls, content: bytes) -> tuple[list[ParsedRow], list[CourseImportIssue], int]:
        try:
            workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
        except Exception as exc:
            raise PortalCourseImportError(msg="无法解析 Excel 文件") from exc
        if not workbook.worksheets:
            raise PortalCourseImportError(msg="Excel 文件没有工作表")
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise PortalCourseImportError(msg="Excel 表头为空")
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        if headers[: len(EXCEL_HEADERS)] != EXCEL_HEADERS:
            raise PortalCourseImportError(msg=f"表头必须为 {' / '.join(EXCEL_HEADERS)}")
        data_rows = rows[1:]
        if len(data_rows) > MAX_IMPORT_ROWS:
            raise PortalCourseImportError(msg=f"导入行数不能超过 {MAX_IMPORT_ROWS}")
        parsed: list[ParsedRow] = []
        issues: list[CourseImportIssue] = []
        total = 0
        for offset, raw in enumerate(data_rows, start=2):
            if cls._is_empty_row(raw):
                continue
            total += 1
            try:
                parsed.append(cls._parse_import_row(offset, tuple(raw or ())))
            except ValueError as exc:
                issues.append(
                    CourseImportIssue(
                        row=offset,
                        code="invalid_row",
                        message=str(exc),
                        recoverable=False,
                    )
                )
        return parsed, issues, total

    @staticmethod
    def _is_empty_row(row: tuple | None) -> bool:
        if not row:
            return True
        return all(cell is None or str(cell).strip() == "" for cell in row)

    @classmethod
    def _parse_import_row(cls, row_idx: int, row: tuple) -> ParsedRow:
        values = list(row or ())
        while len(values) < 9:
            values.append(None)
        external_id = cls._optional_text(values[0], field="课程ID", row_idx=row_idx, max_length=MAX_EXTERNAL_ID_LENGTH)
        name = cls._required_text(values[1], field="课程名称", row_idx=row_idx, max_length=MAX_NAME_LENGTH)
        instructor = cls._optional_text(
            values[2],
            field="主讲人",
            row_idx=row_idx,
            max_length=MAX_INSTRUCTOR_LENGTH,
        ) or ""
        catalog_path = cls._optional_text(values[3], field="课程目录", row_idx=row_idx, max_length=1000)
        source_updated_at = cls._parse_date(row_idx, values[4])
        description = str(values[5]).strip() if values[5] is not None else ""
        cover_url = cls._optional_url(row_idx, values[6], field="课程封面")
        tags = cls._parse_tags(values[7])
        external_url = cls._optional_url(row_idx, values[8], field="课程链接")
        return (
            row_idx,
            external_id,
            name,
            instructor,
            catalog_path,
            source_updated_at,
            description,
            cover_url,
            tags,
            external_url,
        )

    @staticmethod
    def _required_text(value: object, *, field: str, row_idx: int, max_length: int) -> str:
        text = str(value).strip() if value is not None else ""
        if not text:
            raise ValueError(f"第 {row_idx} 行: {field}不能为空")
        if len(text) > max_length:
            raise ValueError(f"第 {row_idx} 行: {field}不能超过 {max_length} 字")
        return text

    @staticmethod
    def _optional_text(value: object, *, field: str, row_idx: int, max_length: int) -> str | None:
        text = str(value).strip() if value is not None else ""
        if not text:
            return None
        if len(text) > max_length:
            raise ValueError(f"第 {row_idx} 行: {field}不能超过 {max_length} 字")
        return text

    @staticmethod
    def _optional_url(row_idx: int, value: object, *, field: str) -> str:
        text = str(value).strip() if value is not None else ""
        if not text:
            return ""
        try:
            return validate_media_url(text)
        except ValueError as exc:
            raise ValueError(f"第 {row_idx} 行: {field}必须是有效的 http(s) 链接") from exc

    @staticmethod
    def _parse_date(row_idx: int, value: object) -> datetime | None:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        text = str(value).strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        raise ValueError(f"第 {row_idx} 行: 更新日期格式无效，请使用 2025/4/8")

    @staticmethod
    def _parse_tags(value: object) -> list[CourseTag]:
        text = str(value).strip() if value is not None else ""
        if not text:
            return []
        tags: list[CourseTag] = []
        seen: set[str] = set()
        for chunk in _TAG_SPLIT.split(text):
            label = chunk.strip()
            if not label or label in seen:
                continue
            seen.add(label)
            tags.append(CourseTag(label=label[:50], display_type="gray"))
        return tags

    @staticmethod
    def _write_instruction_sheet(sheet: Worksheet) -> None:
        lines = [
            "1. 请使用「第三方课程」工作表导入. 表头不可修改.",
            "2. 课程ID 为第三方平台中的课程 ID，字符串. 填写后重复导入会按该 ID 更新已有课程.",
            "3. 课程名称为必填. 课程链接填写后课程会自动启用；留空则导入为未启用.",
            "4. 课程目录填写完整路径，层级用「. 」(点+空格)、-> 或 / 分隔. 例如：D-通用能力培训类. F-微课件. DF22-微课大赛.",
            "5. 目录必须已在课程目录中存在. 预检失败时可选择强行导入，找不到的目录会改为未分类.",
            "6. 更新日期示例 2025/4/8. 课程封面填写 http(s) 图片链接，不要嵌入图片.",
            "7. 多个标签用逗号、顿号或分号分隔.",
            f"8. 单次最多导入 {MAX_IMPORT_ROWS} 行.",
        ]
        sheet.column_dimensions["A"].width = 96
        for idx, line in enumerate(lines, start=1):
            sheet.cell(row=idx, column=1, value=line)
