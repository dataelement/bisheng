from __future__ import annotations

from datetime import datetime
from io import BytesIO

import openpyxl
import pytest

from bisheng.shougang_portal_course.domain.schemas.portal_course_schema import CatalogCreate
from bisheng.shougang_portal_course.domain.services.catalog_service import (
    PortalCourseCatalogService,
)
from bisheng.shougang_portal_course.domain.services.course_import_service import (
    EXCEL_HEADERS,
    PortalCourseImportService,
)
from bisheng.shougang_portal_course.domain.services.course_service import PortalCourseService


def _course_workbook(*rows: tuple) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "第三方课程"
    for col, header in enumerate(EXCEL_HEADERS, start=1):
        sheet.cell(row=1, column=col, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _catalog_workbook(*rows: tuple) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "课程目录"
    headers = ["目录ID", "上级目录ID", "上级目录", "目录名称", "描述", "排序", "是否公开"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_template_headers_match_import_format():
    workbook = openpyxl.load_workbook(BytesIO(PortalCourseImportService.build_template()))
    sheet = workbook["第三方课程"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert headers == EXCEL_HEADERS
    assert "填写说明" in workbook.sheetnames


def test_rejects_catalog_path_in_course_import():
    with pytest.raises(ValueError, match="单个目录名称"):
        PortalCourseImportService._optional_catalog_name(
            2,
            "D-通用能力培训类. F-微课件. DF22-微课大赛",
        )


async def test_import_creates_and_upserts_by_external_id(course_session):
    catalog_service = PortalCourseCatalogService(course_session)
    root = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="D-通用能力培训类"),
    )
    child = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="F-微课件", parent_id=root.id),
    )
    leaf = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="DF22-微课大赛", parent_id=child.id),
    )
    service = PortalCourseImportService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_course_workbook(
            (
                "YX123456",
                "7分钟带你了解财务报告",
                "赵晨露",
                "",
                "DF22-微课大赛",
                "2025/4/8",
                "业财融合",
                "https://example.com/cover.jpg",
                "隐患排查",
                "https://learn.example.com/course/1",
            )
        ),
    )
    assert result.failed == 0
    assert result.success == 1
    courses = await PortalCourseService(course_session).repository.list_courses(
        tenant_id=1,
        public_only=False,
    )
    assert len(courses) == 1
    course = courses[0]
    assert course.course_type == "external"
    assert course.external_id == "YX123456"
    assert course.instructor == "赵晨露"
    assert course.catalog_id == leaf.id
    assert course.cover_url == "https://example.com/cover.jpg"
    assert course.enabled is True
    assert course.source_updated_at == datetime(2025, 4, 8)

    updated = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_course_workbook(
            (
                "YX123456",
                "财务报告进阶",
                "赵晨露",
                "",
                "",
                "2025/5/1",
                "更新简介",
                "",
                "报表,财务",
                "https://learn.example.com/course/1",
            )
        ),
    )
    assert updated.failed == 0
    courses = await PortalCourseService(course_session).repository.list_courses(
        tenant_id=1,
        public_only=False,
    )
    assert len(courses) == 1
    assert courses[0].name == "财务报告进阶"
    assert courses[0].catalog_id is None
    assert "财务" in courses[0].tags_json


async def test_preview_flags_missing_catalog_and_force_uncategorizes(course_session):
    service = PortalCourseImportService(course_session)
    content = _course_workbook(
        (
            "A1",
            "安全课",
            "王老师",
            "",
            "不存在目录",
            "",
            "",
            "",
            "",
            "https://learn.example.com/a1",
        )
    )
    preview = await service.preview_excel(tenant_id=1, content=content)
    assert preview.total == 1
    assert preview.valid == 0
    assert preview.issues[0].code == "missing_catalog"
    assert preview.issues[0].recoverable is True

    skipped = await service.import_excel(tenant_id=1, user_id=7, content=content)
    assert skipped.success == 0
    forced = await service.import_excel(tenant_id=1, user_id=7, content=content, force=True)
    assert forced.success == 1
    courses = await PortalCourseService(course_session).repository.list_courses(
        tenant_id=1,
        public_only=False,
    )
    assert courses[0].catalog_id is None
    assert courses[0].external_id == "A1"


async def test_import_rejects_missing_name(course_session):
    service = PortalCourseImportService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_course_workbook(
            ("A1", "", "王老师", "", "", "", "", "", "", "https://learn.example.com/a1")
        ),
    )
    assert result.success == 0
    assert result.failed == 1
    assert any("课程名称不能为空" in item for item in result.errors)


async def test_import_assigns_catalog_by_external_id(course_session):
    catalog_service = PortalCourseCatalogService(course_session)
    external_catalog_id = "c" * 32
    imported = await catalog_service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_catalog_workbook((external_catalog_id, "", "", "安全生产", "公司级", 1, "是")),
    )
    assert imported.failed == 0
    service = PortalCourseImportService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_course_workbook(
            (
                "B1",
                "安全课",
                "王老师",
                external_catalog_id,
                "错误名称",
                "",
                "",
                "",
                "",
                "https://learn.example.com/b1",
            )
        ),
    )
    assert result.failed == 0
    courses = await PortalCourseService(course_session).repository.list_courses(
        tenant_id=1,
        public_only=False,
    )
    assert courses[0].catalog_id == external_catalog_id


async def test_import_assigns_catalog_by_id(course_session):
    catalog_service = PortalCourseCatalogService(course_session)
    catalog = await catalog_service.create_catalog(
        tenant_id=1,
        user_id=7,
        payload=CatalogCreate(name="安全生产"),
    )
    service = PortalCourseImportService(course_session)
    result = await service.import_excel(
        tenant_id=1,
        user_id=7,
        content=_course_workbook(
            (
                "B1",
                "安全课",
                "王老师",
                catalog.id,
                "错误路径",
                "",
                "",
                "",
                "",
                "https://learn.example.com/b1",
            )
        ),
    )
    assert result.failed == 0
    courses = await PortalCourseService(course_session).repository.list_courses(
        tenant_id=1,
        public_only=False,
    )
    assert courses[0].catalog_id == catalog.id
