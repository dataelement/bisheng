"""Pictures embedded in a workbook must be staged for MinIO and attributed to
the sheet that actually owns them."""

import os
import zipfile

import openpyxl
import pytest
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

from bisheng.knowledge.rag.pipeline.loader.base import BaseBishengLoader
from bisheng.knowledge.rag.pipeline.loader.excel import ExcelLoader
from bisheng.knowledge.rag.pipeline.loader.utils.excel_images import extract_excel_images

PNG_BYTES = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06"
    b"\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\rIDATx\xdac\xfc\x0f\x00\x01\x03\x01"
    b"\x00\x18\xdd\x8d\xb0\x00\x00\x00\x00IEND\xaeB`\x82"
)


def _write_synthetic_workbook(path: str) -> None:
    """A package where sheet order, r:id and file numbering all disagree.

    "Data" is sheet7.xml and owns nothing; "Cover" is sheet2.xml, has no cells
    at all and owns the only picture. Pairing sheets with drawings positionally
    would attribute the picture to the wrong sheet.
    """
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Data" sheetId="1" r:id="rId9"/>'
            '<sheet name="Cover" sheetId="2" r:id="rId4"/></sheets></workbook>',
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId9" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/worksheet" Target="worksheets/sheet7.xml"/>'
            '<Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/worksheet" Target="worksheets/sheet2.xml"/></Relationships>',
        )
        zf.writestr("xl/worksheets/sheet7.xml", '<?xml version="1.0"?><worksheet/>')
        zf.writestr("xl/worksheets/sheet2.xml", '<?xml version="1.0"?><worksheet/>')
        zf.writestr(
            "xl/worksheets/_rels/sheet2.xml.rels",
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/drawing" Target="../drawings/drawing3.xml"/></Relationships>',
        )
        # Namespace prefixes deliberately renamed: "xdr:"/"a:" are conventions, not rules.
        zf.writestr(
            "xl/drawings/drawing3.xml",
            '<?xml version="1.0"?>'
            '<d:wsDr xmlns:d="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
            ' xmlns:m="http://schemas.openxmlformats.org/drawingml/2006/main"'
            ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            "<d:oneCellAnchor><d:from><d:col>3</d:col><d:row>7</d:row></d:from>"
            '<d:pic><d:blipFill><m:blip r:embed="rId1"/></d:blipFill></d:pic></d:oneCellAnchor>'
            "</d:wsDr>",
        )
        zf.writestr(
            "xl/drawings/_rels/drawing3.xml.rels",
            '<?xml version="1.0"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/image" Target="../media/image1.png"/></Relationships>',
        )
        zf.writestr("xl/media/image1.png", PNG_BYTES)


def _write_real_workbook(path: str, png_path: str) -> None:
    """A workbook openpyxl itself produces: picture-only first sheet, data second."""
    PILImage.new("RGB", (4, 4), "white").save(png_path)

    workbook = openpyxl.Workbook()
    cover = workbook.active
    cover.title = "Portlet"
    cover.add_image(OpenpyxlImage(png_path), "B2")

    data = workbook.create_sheet("Prices")
    data.append(["date", "close"])
    data.append(["2026-05-08", 15.99])
    workbook.save(path)


def test_picture_is_attributed_to_the_sheet_that_owns_it(tmp_path):
    xlsx = str(tmp_path / "synthetic.xlsx")
    _write_synthetic_workbook(xlsx)

    images = extract_excel_images(xlsx)

    assert [(img.sheet_name, img.media_name) for img in images] == [("Cover", "image1.png")]
    assert images[0].content == PNG_BYTES


def test_non_package_files_are_skipped_quietly(tmp_path):
    csv_path = tmp_path / "plain.csv"
    csv_path.write_text("a,b\n1,2\n", encoding="utf-8")

    assert extract_excel_images(str(csv_path)) == []
    assert extract_excel_images(str(tmp_path / "missing.xlsx")) == []


def test_loader_stages_pictures_and_emits_a_chunk_per_sheet(tmp_path, monkeypatch):
    monkeypatch.setattr(BaseBishengLoader, "_minio_bucket", property(lambda self: "bisheng"))

    xlsx = str(tmp_path / "report.xlsx")
    _write_real_workbook(xlsx, str(tmp_path / "pic.png"))

    loader = ExcelLoader(
        file_path=xlsx,
        file_metadata={"source": "report.xlsx"},
        file_extension="xlsx",
        tmp_dir=str(tmp_path / "work"),
        image_object_dir="knowledge/images/90030/90286",
    )
    documents = loader.load()

    image_docs = [doc for doc in documents if doc.page_content.startswith("## Portlet")]
    assert len(image_docs) == 1, "the picture-only sheet must still produce a chunk"
    assert "![image1.png](/bisheng/knowledge/images/90030/90286/image1.png)" in image_docs[0].page_content

    # Bytes are staged locally; ImageUploadTransformer performs the upload.
    assert os.listdir(loader.local_image_dir) == ["image1.png"]

    # The table sheet is still parsed, and chunk_index stays contiguous.
    assert any("2026-05-08" in doc.page_content for doc in documents)
    assert [doc.metadata["chunk_index"] for doc in documents] == list(range(len(documents)))


def test_loader_leaves_pictureless_workbooks_untouched(tmp_path, monkeypatch):
    monkeypatch.setattr(BaseBishengLoader, "_minio_bucket", property(lambda self: "bisheng"))

    xlsx = str(tmp_path / "plain.xlsx")
    workbook = openpyxl.Workbook()
    workbook.active.append(["date", "close"])
    workbook.active.append(["2026-05-08", 15.99])
    workbook.save(xlsx)

    loader = ExcelLoader(
        file_path=xlsx,
        file_metadata={},
        file_extension="xlsx",
        tmp_dir=str(tmp_path / "work"),
        image_object_dir="knowledge/images/1/2",
    )
    documents = loader.load()

    assert documents, "table content must still be produced"
    assert all(not doc.page_content.startswith("## ") for doc in documents)
    # No staging dir is created when there is nothing to stage.
    assert loader.local_image_dir is None


@pytest.mark.parametrize("extension", ["csv"])
def test_image_extraction_is_limited_to_xlsx(tmp_path, extension, monkeypatch):
    """csv/xls never carry an OPC drawing part; the loader must not even look."""
    calls = []
    monkeypatch.setattr(
        "bisheng.knowledge.rag.pipeline.loader.excel.extract_excel_images",
        lambda path: calls.append(path) or [],
    )

    csv_path = tmp_path / f"data.{extension}"
    csv_path.write_text("date,close\n2026-05-08,15.99\n", encoding="utf-8")

    loader = ExcelLoader(
        file_path=str(csv_path),
        file_metadata={},
        file_extension=extension,
        tmp_dir=str(tmp_path / "work"),
        image_object_dir="knowledge/images/1/2",
    )
    loader.load()

    assert calls == []
