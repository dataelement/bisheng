#!/usr/bin/env python3
"""Read out and health-check an .xlsx, for the BiSheng code interpreter.

    python skills/bisheng-xlsx/scripts/inspect_workbook.py output/x.xlsx
    python skills/bisheng-xlsx/scripts/inspect_workbook.py output/x.xlsx --content-only
    python skills/bisheng-xlsx/scripts/inspect_workbook.py output/x.xlsx --check-only

Two jobs the official skill delegates to things this environment does not have:

* **Content read-out** replaces ``markitdown`` (absent here), and unlike it this
  prints cell coordinates, so edits can be planned from the output.
* **Health check** replaces "the model should remember ~10 negative rules".
  Every rule the official SKILL.md states as prose and never verifies — banned
  spilling functions, the ``_xlfn.`` prefix set, percentages stored as text,
  unquoted sheet names, CJK column width — is checked mechanically here.

Always exits 0: the BiSheng executor discards stdout on a non-zero exit, which
would throw away the entire report.
"""

import argparse
import datetime as dt
import os
import re
import sys
import traceback

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, range_boundaries
except ImportError:  # pragma: no cover - only when the image is cut down
    print("[FATAL] openpyxl 不可用；无法体检。请让运维在后端环境确认依赖。")
    sys.exit(0)

# LibreOffice cannot evaluate these under any prefix, and — worse — on builds
# that do evaluate them, an openpyxl-written file carries no spill metadata, so
# only the top-left cell gets a value and the recalc reports zero errors.
BANNED = ["XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE", "LET", "LAMBDA", "TEXTSPLIT"]

# Post-2007 names Excel stores prefixed; openpyxl writes the string verbatim, so
# a bare name becomes a literal #NAME? in the delivered file.
NEEDS_XLFN = ["TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"]

# A banned name stays banned behind the `_xlfn.` prefix — writing `_xlfn.SORT`
# does not make LibreOffice able to spill. So the prefix must be part of the
# match, not something the leading guard silently swallows: with a bare
# `(?<![A-Z_.])` lookbehind the `.` in `_XLFN.SORT(` blocks the match and the
# whole rule goes quiet.
BANNED_RE = {fn: re.compile(rf"(?<![A-Z0-9_])(?:_XLFN\.)?{fn}\s*\(") for fn in BANNED}
# MAXIFS/MINIFS/COUNTIFS end in IFS, hence the leading `[A-Z0-9_]` guard.
XLFN_RE = {fn: re.compile(rf"(?<![A-Z0-9_.]){fn}\s*\(") for fn in NEEDS_XLFN}

ERROR_LITERALS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

PLACEHOLDERS = ["待填", "待补", "TODO", "XXX", "xxx", "占位", "示例数据", "请填写", "TBD", "lorem"]

# A bare multiplier inside a formula is an assumption nobody can find later.
# Matched by capture-then-filter rather than by a nested negative lookahead:
# `(?!1(?![\d.]))` reads as "not a bare 1" but silently also exempts `*1.15`,
# because the char after `1` is `.`. Exemptions belong in a readable set.
COEFFICIENT_RE = re.compile(r"[*/]\s*(\d+(?:\.\d+)?)")
# Unit conversion (元→万元, 月→年, 天→年) and single digits are not assumptions.
UNIT_CONSTANTS = {
    "0",
    "1",
    "10",
    "12",
    "24",
    "30",
    "36",
    "52",
    "60",
    "90",
    "100",
    "180",
    "360",
    "365",
    "1000",
    "1024",
    "10000",
    "100000",
    "1000000",
    "10000000",
    "100000000",
}

# Group 1 = the name inside quotes, group 2 = a bare name. The alternation must
# be NON-capturing at the outer level, otherwise group(1) always matches and
# every "is it quoted?" test silently answers yes.
SHEET_REF_RE = re.compile(r"(?<![\w!])(?:'([^']+)'|([A-Za-z0-9_一-鿿]+))!")

# Numbers parked in a text cell: they do not sum, do not sort, and the status bar
# shows nothing when the user selects the column.
TEXT_PCT_RE = re.compile(r"^[-+]?\d{1,3}(,\d{3})*(\.\d+)?\s*%$")
TEXT_NUM_RE = re.compile(r"^[-+]?[¥￥$€]?\s?(\d{1,3}(,\d{3})+(\.\d+)?|\d+(\.\d+)?)\s*(元|万元|亿元|美元|人民币)$")
TEXT_GROUPED_RE = re.compile(r"^[-+]?[¥￥$€]?\s?\d{1,3}(,\d{3})+(\.\d+)?$")

# Excel's own default when a column has no explicit width. A column nobody sized
# is exactly where ### shows up, so the check must not skip it.
DEFAULT_COL_WIDTH = 8.43

DECIMALS_RE = re.compile(r"\.(0+)")

# Examples printed per rule+scope before the rest collapse into a count.
MAX_PER_GROUP = 3

findings: list[tuple[str, str, str]] = []
_seen: set[tuple[str, str]] = set()


def add(level: str, message: str, group: str | None = None) -> None:
    """Record a finding, dropping exact repeats and tagging it for collapsing.

    Two things ruin this report as an input to a model. Exact repeats: a formula
    like ``=SUM(区域!A:A)+区域!B1`` used to emit the same sentence twice. And one
    rule firing per cell: a 500-row column with no number format used to print
    500 identical INFO lines and bury every ERROR above them. ``group`` names the
    rule+scope so the printer shows a few examples and a count.
    """
    key = (level, message)
    if key in _seen:
        return
    _seen.add(key)
    findings.append((level, message, group or message))


def cjk_width(text) -> int:
    """Display width in half-widths. Must stay identical to
    ``xlsx_helpers.cjk_width`` — the helper sizes the columns, this measures
    them, and two different definitions make the checker fight the builder."""
    if text is None:
        return 0
    width = 0
    for ch in str(text):
        # CJK ideographs, CJK punctuation, kana, and full-width forms.
        if "一" <= ch <= "鿿" or "　" <= ch <= "〿" or "぀" <= ch <= "ヿ" or "＀" <= ch <= "￯":
            width += 2
        else:
            width += 1
    return width


def rendered_text(cell, cached=None) -> str:
    """What Excel actually draws in the cell, honouring ``number_format``.

    Measuring the raw value under-counts by exactly the characters that cause
    ###: 12500000 is 8 wide as a Python int but ``#,##0`` renders it as
    ``12,500,000``, 10 wide. ``cached`` supplies the recalculated value for a
    formula cell, whose own text (``=SUM(B2:B9)``) says nothing about its width.
    """
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        value = cached
    if value is None:
        return ""
    fmt = (cell.number_format or "General").split(";")[0]
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return "X" * min(20, max(8, len(fmt)))
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return str(value)

    if fmt in ("General", "@") or not any(ch in fmt for ch in "0#"):
        # General has no fixed rendering: Excel rounds the display down to
        # whatever fits, so measuring it as `repr` over-states the need. Cap it
        # at the width beyond which Excel starts shortening on its own.
        return str(value)[:11]

    scaled = value * 100 if "%" in fmt else value
    match = DECIMALS_RE.search(fmt)
    decimals = len(match.group(1)) if match else 0
    grouped = "#,##" in fmt or "0,00" in fmt
    body = f"{abs(scaled):,.{decimals}f}" if grouped else f"{abs(scaled):.{decimals}f}"
    if scaled < 0:
        body = "-" + body
    for symbol in ("¥", "￥", "$", "€"):
        if symbol in fmt:
            body = symbol + body
            break
    if "%" in fmt:
        body += "%"
    return body


def needs_quoting(sheet_name: str) -> bool:
    """Whether a cross-sheet reference to this name *must* be single-quoted.

    Deliberately narrower than the quoting rule in ``xlsx_helpers.sheet_ref``:
    that one quotes liberally because quoting is always safe, while flagging
    something as an ERROR needs certainty. A pure-CJK name like ``利润预测`` is
    perfectly legal unquoted (verified: LibreOffice evaluates ``=假设!$B$2``),
    so treating "non-ASCII" as "must quote" would false-positive on nearly every
    Chinese workbook.
    """
    return any(ch in sheet_name for ch in " -()&'+.") or (sheet_name[:1].isdigit() if sheet_name else False)


def hardcoded_coefficients(formula: str) -> list[str]:
    """Bare multipliers/divisors that are neither unit conversions nor 1-digit."""
    hits = []
    for match in COEFFICIENT_RE.finditer(formula):
        number = match.group(1)
        if formula[match.end() : match.end() + 1] in {"%", "E", "e"}:
            continue
        if number in UNIT_CONSTANTS:
            continue
        if "." not in number and len(number) < 2:
            continue
        hits.append(number)
    return hits


def fmt_value(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", "⏎")
    return text if len(text) <= 60 else text[:57] + "…"


# --------------------------------------------------------------------------- #
# Content read-out
# --------------------------------------------------------------------------- #


def dump_content(path: str, max_rows: int) -> None:
    wb = load_workbook(path, data_only=False)
    wb_values = None
    try:
        wb_values = load_workbook(path, data_only=True)
    except Exception:
        pass  # cached values are a bonus here; the formula pass is the required one

    print("=" * 72)
    print(f"内容 · {os.path.basename(path)} · {len(wb.sheetnames)} 个工作表")
    print("=" * 72)

    for name in wb.sheetnames:
        ws = wb[name]
        vs = wb_values[name] if wb_values is not None and name in wb_values.sheetnames else None
        print(f"\n## 工作表「{name}」  {ws.max_row} 行 × {ws.max_column} 列", end="")
        if ws.freeze_panes:
            print(f"  冻结={ws.freeze_panes}", end="")
        if ws.auto_filter and ws.auto_filter.ref:
            print(f"  筛选={ws.auto_filter.ref}", end="")
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f"  合并={len(merged)}处", end="")
        print()

        shown = 0
        for row in ws.iter_rows():
            if shown >= max_rows:
                print(f"   … 其余 {ws.max_row - shown} 行未显示（--max-rows 可调）")
                break
            cells = [c for c in row if c.value is not None]
            if not cells:
                continue
            parts = []
            for cell in cells:
                text = fmt_value(cell.value)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cached = vs[cell.coordinate].value if vs is not None else None
                    text = f"{text} →{fmt_value(cached) if cached is not None else '未重算'}"
                parts.append(f"{cell.coordinate}={text}")
            print("   " + " | ".join(parts))
            shown += 1
    wb.close()
    if wb_values is not None:
        wb_values.close()


# --------------------------------------------------------------------------- #
# Health checks
# --------------------------------------------------------------------------- #


def check_formulas(ws, name: str, sheet_names: set[str]) -> int:
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            where = f"{name}!{cell.coordinate}"
            stripped = value.strip()

            # Exact match only. `=IF(ISNA(X),"#N/A",X)` is a legitimate formula;
            # a baked error is a cell whose entire content IS the literal.
            if stripped in ERROR_LITERALS:
                add(
                    "ERROR",
                    f"{where} 单元格里是错误值 {stripped}（不是公式，是已经算坏的结果）",
                    f"{name} 错误值字面量",
                )
                continue

            if any(p in value for p in PLACEHOLDERS):
                add("WARN", f"{where} 残留占位文本：{fmt_value(value)}", f"{name} 残留占位文本")

            if not value.startswith("="):
                if TEXT_PCT_RE.match(stripped):
                    add(
                        "ERROR",
                        f"{where} 百分比存成了文本 {stripped!r} → 不参与求和、排序按字符串排。"
                        f"改成写数值 {stripped.rstrip('%').strip()} / 100 并设 number_format='0.0%'",
                        f"{name} 百分比存成文本",
                    )
                elif TEXT_NUM_RE.match(stripped) or TEXT_GROUPED_RE.match(stripped):
                    add(
                        "WARN",
                        f"{where} 数字存成了文本 {stripped!r}（带了单位或千分位）→ 选中整列右下角不显示求和。"
                        f"单位写进列名，单元格只放数值，千分位交给 number_format='#,##0'",
                        f"{name} 数字存成文本",
                    )
                continue

            formula_count += 1
            upper = value.upper()

            for fn, pattern in BANNED_RE.items():
                if pattern.search(upper):
                    add(
                        "ERROR",
                        f"{where} 用了 {fn}()——LibreOffice 重算不了，交付出去在用户 Excel 里可能是 #NAME?，"
                        f"且重算脚本抓不到。改用 INDEX/MATCH，或在 Python 里排序去重后直接写值。",
                        f"{name} 用了 {fn}",
                    )
            for fn, pattern in XLFN_RE.items():
                if pattern.search(upper):
                    add(
                        "ERROR",
                        f"{where} {fn}() 缺 `_xlfn.` 前缀 → 会变成字面量 #NAME?。写成 `_xlfn.{fn}(...)`",
                        f"{name} {fn} 缺前缀",
                    )

            # Drive this off the real sheet names rather than trying to infer a
            # name from the bare match: `=假设 输入!$B$2` tokenises as the bare
            # name `输入`, so the space is simply not visible from the match.
            for sname in sheet_names:
                if not needs_quoting(sname):
                    continue
                if f"{sname}!" in value and f"'{sname}'!" not in value:
                    add(
                        "ERROR",
                        f"{where} 跨表引用「{sname}」名字含空格/特殊字符却没加单引号 → #VALUE!。写成 '{sname}'!",
                        f"{name} 引用「{sname}」缺单引号",
                    )

            for match in SHEET_REF_RE.finditer(value):
                inner, bare = match.group(1), match.group(2)
                ref_name = inner or bare
                if not ref_name or ref_name in sheet_names:
                    continue
                if ref_name.upper() in {"TRUE", "FALSE"}:
                    continue
                # A bare token before `!` is only a sheet reference if it is not
                # the tail of a quoted name we already matched.
                if any(ref_name in s for s in sheet_names):
                    continue
                add(
                    "WARN",
                    f"{where} 引用了工作表「{ref_name}」，但工作簿里没有这个表名",
                    f"{name} 引用不存在的表 {ref_name}",
                )

            coefficients = hardcoded_coefficients(value)
            if coefficients:
                add(
                    "WARN",
                    f"{where} 公式里硬编码了系数 {'、'.join(coefficients)}：{fmt_value(value)} "
                    f"→ 把它挪到有标签的假设单元格并引用它",
                    f"{name} 硬编码系数",
                )

            if "IFERROR" not in upper and re.search(r"/\s*(?:'[^']+'!|[\w一-鿿]+!)?\$?[A-Za-z]{1,3}\$?\d", value):
                add("INFO", f"{where} 除法未加 IFERROR 保护：{fmt_value(value)}", f"{name} 除法未加 IFERROR")
    return formula_count


def check_header(ws, name: str) -> None:
    """First data-ish row must be column names, not the first data record."""
    if ws.max_column < 3 or ws.max_row < 4:
        return  # key-value / assumption sheets legitimately have no header row
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        cells = [c for c in row if c.value is not None]
        if len(cells) < 2:
            continue  # a merged title row shows up as a single cell; skip past it
        bad = [
            c.coordinate
            for c in cells
            if (isinstance(c.value, (int, float, dt.date, dt.datetime)) and not isinstance(c.value, bool))
            or (isinstance(c.value, str) and c.value.startswith("="))
        ]
        if bad:
            add(
                "WARN",
                f"工作表「{name}」第 {cells[0].row} 行是首个数据行，里面 {len(bad)}/{len(cells)} 格是数字或公式"
                f"（{'、'.join(bad[:4])}）→ 这张表缺表头。在它上面插一行中文列名，并 freeze_panes 冻结。",
            )
        return


def check_layout(ws, name: str, vs) -> None:
    if ws.max_row <= 1 and ws.max_column <= 1 and ws["A1"].value is None:
        add("WARN", f"工作表「{name}」是空的——要么填内容，要么删掉它")
        return

    check_header(ws, name)

    # Column width vs content: the ### bug, and it hides Chinese first.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        explicit = dim is not None and dim.width
        width = float(dim.width) if explicit else float(ws.sheet_format.defaultColWidth or DEFAULT_COL_WIDTH)
        need_num, need_text = 0, 0
        sample_num, sample_text = "", ""
        for row_idx in range(1, min(ws.max_row, 500) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue
            alignment = cell.alignment
            if alignment is not None and alignment.wrap_text:
                continue  # wrapped text grows the row, not the column
            cached = vs[cell.coordinate].value if vs is not None else None
            text = rendered_text(cell, cached)
            if not text:
                continue
            numeric = isinstance(cell.value, (int, float, dt.date, dt.datetime)) and not isinstance(cell.value, bool)
            numeric = numeric or (isinstance(cell.value, str) and cell.value.startswith("=") and cached is not None)
            # ### is a *formatted* number that will not fit. A General number
            # never shows ### — Excel rounds the display instead — so it belongs
            # in the softer bucket.
            if numeric and (cell.number_format or "General") not in ("General", "@"):
                if cjk_width(text) > need_num:
                    need_num, sample_num = cjk_width(text), text
            elif cjk_width(text) > need_text:
                need_text, sample_text = cjk_width(text), text
        note = "" if explicit else "（没设列宽，用的是默认 8.43）"
        if need_num and width < need_num:
            add(
                "WARN",
                f"{name} 第 {letter} 列宽 {width:.0f}{note} < 数字渲染后的 {need_num}（如 {sample_num}，中文按双宽算）"
                f"→ 打开就是 ###。调 ws.column_dimensions['{letter}'].width，或收尾时调 autofit_columns(ws)",
            )
        elif need_text and width < need_text:
            add(
                "INFO",
                f"{name} 第 {letter} 列宽 {width:.0f}{note} < 文字需要的 {need_text}（如 {sample_text}）"
                f"→ 右边有内容时会被截断。调 autofit_columns(ws) 或给这列设 wrap_text",
            )

    # Merged ranges whose non-anchor cells carry values (openpyxl silently drops them).
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) == (min_row, min_col):
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    add(
                        "WARN",
                        f"{name}!{rng} 合并区里非左上角单元格 {cell.coordinate} 仍有值，Excel 打开时会丢",
                        f"{name} 合并区非左上角有值",
                    )
                    break

    if ws.max_row > 15 and not ws.freeze_panes:
        add("INFO", f"工作表「{name}」超过 15 行但没冻结表头，滚动后看不到列名")


def check_number_formats(ws, name: str) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fmt = cell.number_format or "General"

            # Percent format holding a value that is plainly not a fraction.
            if "%" in fmt and isinstance(cell.value, (int, float)) and abs(cell.value) > 1.5:
                add(
                    "ERROR",
                    f"{name}!{cell.coordinate} 百分比格式但存了 {cell.value} → 会显示成 "
                    f"{cell.value * 100:.0f}%。百分比要存小数（15% 存 0.15）",
                    f"{name} 百分比格式存了整数",
                )

            # Big bare numbers with no format read as a wall of digits.
            if fmt == "General" and isinstance(cell.value, (int, float)) and abs(cell.value) >= 10000:
                add(
                    "INFO",
                    f"{name}!{cell.coordinate} 金额 {cell.value} 没有数字格式，建议 `#,##0`",
                    f"{name} 第 {cell.column_letter} 列金额无数字格式",
                )

            # Years must not be thousands-separated.
            if isinstance(cell.value, int) and 1900 <= cell.value <= 2100 and "#,##" in fmt:
                add(
                    "WARN",
                    f"{name}!{cell.coordinate} 年份 {cell.value} 用了千分位格式，会显示成 {cell.value:,}",
                    f"{name} 年份用千分位",
                )


def run_checks(path: str) -> None:
    wb = load_workbook(path, data_only=False)
    try:
        wb_values = load_workbook(path, data_only=True)
    except Exception:
        wb_values = None  # a file this broken will surface elsewhere; keep the report alive

    sheet_names = set(wb.sheetnames)
    total_formulas = 0
    uncached = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            continue  # chartsheet
        vs = wb_values[name] if wb_values is not None and name in wb_values.sheetnames else None
        total_formulas += check_formulas(ws, name, sheet_names)
        check_layout(ws, name, vs)
        check_number_formats(ws, name)
        if vs is not None:
            for row in ws.iter_rows():
                for cell in row:
                    if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                        continue
                    cached = vs[cell.coordinate].value
                    if cached is None:
                        uncached += 1
                    elif isinstance(cached, str) and cached.strip() in ERROR_LITERALS:
                        # Recalculated to an error. Invisible in the formula pass
                        # (the cell still holds its formula text), so it must be
                        # read off the cached value or only recalc_check sees it.
                        add(
                            "ERROR",
                            f"{name}!{cell.coordinate} 重算结果是 {cached.strip()}：{fmt_value(cell.value)} "
                            f"→ 公式本身算不出来，改完再跑一次 recalc_check.py",
                            f"{name} 重算结果是错误值",
                        )

    if uncached:
        add(
            "ERROR",
            f"{uncached}/{total_formulas} 个公式没有缓存值——openpyxl 写出的公式不带结果，"
            f"pandas 和多数预览器读到的是空。必须先跑 recalc_check.py 重算。",
        )
    wb.close()
    if wb_values is not None:
        wb_values.close()

    print()
    print("=" * 72)
    print(f"体检 · 公式 {total_formulas} 个")
    print("=" * 72)

    order = {"ERROR": 0, "WARN": 1, "INFO": 2}
    counts = {"ERROR": 0, "WARN": 0, "INFO": 0}
    for level, _, _ in findings:
        counts[level] += 1

    # One rule × one column can legitimately fire 500 times. Print a few examples
    # per rule and count the rest, so the ERRORs stay visible above the noise.
    sizes: dict[tuple[str, str], int] = {}
    for level, _, group in findings:
        sizes[(level, group)] = sizes.get((level, group), 0) + 1
    shown: dict[tuple[str, str], int] = {}
    for level, message, group in sorted(findings, key=lambda f: order[f[0]]):
        key = (level, group)
        shown[key] = shown.get(key, 0) + 1
        if shown[key] <= MAX_PER_GROUP:
            print(f"[{level}] {message}")
        elif shown[key] == MAX_PER_GROUP + 1:
            print(f"[{level}] …「{group}」同类还有 {sizes[key] - MAX_PER_GROUP} 处，按同样的办法一次改完")
    if not findings:
        print("（无发现）")

    print()
    print(f"合计: {counts['ERROR']} ERROR / {counts['WARN']} WARN / {counts['INFO']} INFO")
    # Only ERROR gates delivery, and the pass line must be printable whenever
    # ERROR is zero. An extra "conditional pass" tier for WARN reads as helpful
    # but is a trap: SKILL.md tells the model to iterate until `结论: 通过`
    # appears, while several WARN rules here are heuristics that a legitimate
    # workbook can never clear (a genuinely header-less matrix, the ±2 error of
    # the rendered-width estimate). The model would then loop until it runs out
    # of turns. Same wording as bisheng-pptx / bisheng-docx on purpose.
    if counts["ERROR"]:
        print("结论: 不通过 —— ERROR 必须全部修完再交付。")
    else:
        print("结论: 通过 —— 无必须修复项。WARN 逐条复核后即可交付。")


def main() -> None:
    parser = argparse.ArgumentParser(description="读出并体检一个 .xlsx")
    parser.add_argument("path")
    parser.add_argument("--content-only", action="store_true", help="只读内容，不体检")
    parser.add_argument("--check-only", action="store_true", help="只体检，不打印内容")
    parser.add_argument("--max-rows", type=int, default=40, help="每个表最多打印多少行（默认 40）")
    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f"[FATAL] 文件不存在: {args.path}")
        print("提示：代码执行器 cwd 就是工作区根，用相对路径 `output/x.xlsx`，不要写 `/output/x.xlsx`。")
        return

    try:
        if not args.check_only:
            dump_content(args.path, args.max_rows)
        if not args.content_only:
            run_checks(args.path)
    except Exception as exc:
        # Report and still exit 0 — a traceback on stdout is useful; a non-zero
        # exit would make the executor discard this whole report.
        print("[FATAL] 体检过程本身出错：")
        if "not a zip file" in str(exc).lower():
            print("这个文件不是真的 .xlsx（.xlsx 本质是 zip）。旧版 .xls 或改了扩展名的 csv 会这样。")
            print("改法：.xls/.csv 用 pandas.read_excel / read_csv 读进来，再用 openpyxl 写一份新的 .xlsx。")
        traceback.print_exc(file=sys.stdout)


if __name__ == "__main__":
    main()
    sys.exit(0)
