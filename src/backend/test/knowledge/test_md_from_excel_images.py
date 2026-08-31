"""Regression tests: multi-sheet Excel parsing must retain embedded images.

Issue: a multi-sheet xlsx containing embedded images lost every image during
knowledge parsing — md_from_excel only exported cell values, so image-bearing
sheets lost their visual content and image-only sheets were dropped entirely.
"""

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XLImage

from bisheng.knowledge.rag.pipeline.loader.utils.md_from_excel import (
    convert_file_to_markdown,
    extract_sheet_images,
)


@pytest.fixture
def tiny_png_bytes() -> bytes:
    # 1x1 red PNG
    import struct
    import zlib

    def chunk(typ, data):
        c = struct.pack(">I", len(data)) + typ + data
        return c + struct.pack(">I", zlib.crc32(typ + data) & 0xFFFFFFFF)

    ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
    idat = chunk(b"IDAT", zlib.compress(b"\x00\xff\x00\x00"))
    iend = chunk(b"IEND", b"")
    return b"\x89PNG\r\n\x1a\n" + ihdr + idat + iend


def _build_workbook(path, png_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data1"
    ws1.append(["name", "score"])
    ws1.append(["alice", 95])

    ws2 = wb.create_sheet("Data2")
    ws2.append(["date", "value"])
    ws2.append(["2026-06-08", 15.99])
    ws2.append(["2026-06-09", 16.29])
    img = XLImage(png_path)
    ws2.add_image(img, "B10")

    ws3 = wb.create_sheet("OnlyImage")
    img2 = XLImage(png_path)
    ws3.add_image(img2, "A1")
    wb.save(path)


@pytest.fixture
def multi_sheet_xlsx(tmp_path, tiny_png_bytes):
    png_path = tmp_path / "tiny.png"
    png_path.write_bytes(tiny_png_bytes)
    xlsx_path = tmp_path / "multi.xlsx"
    _build_workbook(str(xlsx_path), str(png_path))
    return str(xlsx_path)


def test_extract_sheet_images_maps_anchor_row(tmp_path, multi_sheet_xlsx):
    image_dir = tmp_path / "images"
    image_dir.mkdir()
    wb = openpyxl.load_workbook(multi_sheet_xlsx)
    result = extract_sheet_images(wb["Data2"], str(image_dir), 1)

    # image anchored at B10 -> 0-based row 9
    assert 9 in result
    (line,) = result[9]
    filename = line.split("(")[1].rstrip(")").rsplit("/", 1)[-1]
    assert (image_dir / filename).exists()
    assert (image_dir / filename).read_bytes().startswith(b"\x89PNG")


def test_convert_keeps_images_in_markdown(tmp_path, multi_sheet_xlsx):
    out_dir = tmp_path / "md"
    image_dir = tmp_path / "images"
    image_dir.mkdir()
    convert_file_to_markdown(multi_sheet_xlsx, [0, 1], 12, str(out_dir), image_dir=str(image_dir))

    files = sorted(f for f in out_dir.iterdir())
    # sheet0 (Data1) + sheet1 (Data2, with image) + sheet2 (image-only)
    assert len(files) == 3

    data2_content = (out_dir / "01000.md").read_text(encoding="utf-8")
    assert "| date | value |" in data2_content
    assert "![" in data2_content
    assert str(image_dir) in data2_content

    # image-only sheet still produces content
    image_only_content = (out_dir / "02000.md").read_text(encoding="utf-8")
    assert "![" in image_only_content

    # extracted image bytes staged for the upload transformer
    staged = list(image_dir.iterdir())
    assert len(staged) == 2
    for f in staged:
        assert f.read_bytes().startswith(b"\x89PNG")


def test_convert_without_image_dir_unchanged(tmp_path, multi_sheet_xlsx):
    out_dir = tmp_path / "md"
    convert_file_to_markdown(multi_sheet_xlsx, [0, 1], 12, str(out_dir))

    # legacy behavior: image-only sheet dropped, no image refs anywhere
    assert not (out_dir / "02000.md").exists()
    for f in out_dir.iterdir():
        assert "![" not in f.read_text(encoding="utf-8")
