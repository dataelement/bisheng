from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from bisheng.open_endpoints.domain.schemas.inspection_standard_sync import (
    InspectionStandardItemRecord,
    InspectionStandardRecord,
)
from bisheng.open_endpoints.domain.services.inspection_standard_excel_builder import (
    SHEET_ITEM_NAME,
    SHEET_STANDARD_NAME,
    build_inspection_standard_xlsx_bytes,
)


def _sample_standard(**overrides) -> InspectionStandardRecord:
    payload = {
        "CREATE_DEPT_ID": "DEPT-A",
        "CHECK_STANDARD_ID": "270101J01D01",
        "DEVICE_NAME": "调度大厅",
        "STANDARD_TYPE": "01-常规点检",
        "CHECK_ITEM_NAME": "调度操作台综合检查",
        "DEVICE_STATUS": "1-运转",
        "ENFORCE_CODE": "1-点检",
        "SAFETY_BOARD": "N-否",
        "CHECK_PERIOD": 1,
        "PERIOD_UNIT": "W-周",
        "INTERFACE_SYSTEM": "1-智能点检系统",
        "NEXT_SCHE_DATE": "2026-05-06",
        "MAINTAIN_REASON": "初始建立",
        "DEVICE_MAINTAIN_JOB_ID": "JOB001",
        "REC_CREATOR": "E001",
        "REC_CREATOR_NAME": "张三",
    }
    payload.update(overrides)
    return InspectionStandardRecord.model_validate(payload)


def _sample_item(**overrides) -> InspectionStandardItemRecord:
    payload = {
        "CHECK_STANDARD_ID": "270101J01D01",
        "CHECK_STANDARD_SEQ_NO": "1",
        "CONTENT": "操作按钮、手柄灵活性及定位",
        "CHECK_WAY": "1-五感",
        "LUBRIC_WAY": "0-无",
        "MANAGE_CONTROL_MODE": "0-无",
        "DATA_TYPE": "10-定性",
        "CRITERI": "灵活方便，定位可靠（GB 20905）",
        "STATUTORY_REQ": "0-无",
    }
    payload.update(overrides)
    return InspectionStandardItemRecord.model_validate(payload)


def test_build_inspection_standard_xlsx_bytes_writes_template_headers_and_data():
    content = build_inspection_standard_xlsx_bytes(
        check_standards=[_sample_standard()],
        check_standard_items=[_sample_item()],
    )

    workbook = load_workbook(filename=BytesIO(content))
    assert workbook.sheetnames[0] == SHEET_STANDARD_NAME
    assert workbook.sheetnames[1] == SHEET_ITEM_NAME

    standard_sheet = workbook[SHEET_STANDARD_NAME]
    assert standard_sheet.cell(row=1, column=1).value == "设备所属单位"
    assert standard_sheet.cell(row=2, column=1).value == "DEPT-A"
    assert standard_sheet.cell(row=2, column=2).value == "270101J01D01"
    assert standard_sheet.cell(row=2, column=13).value == "2026-05-06"

    item_sheet = workbook[SHEET_ITEM_NAME]
    assert item_sheet.cell(row=1, column=1).value == "点检标准编号（12）*"
    assert item_sheet.cell(row=2, column=2).value == "001"
