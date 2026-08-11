#!/usr/bin/env python3
"""Read out and health-check an .xlsx, for the BiSheng code interpreter.

    python skills/bisheng-xlsx/scripts/inspect_workbook.py output/x.xlsx
    python skills/bisheng-xlsx/scripts/inspect_workbook.py output/x.xlsx --content-only
    python skills/bisheng-xlsx/scripts/inspect_workbook.py output/x.xlsx --check-only

Two jobs the official skill delegates to things this environment does not have:

* **Content read-out** replaces ``markitdown`` (absent here), and unlike it this
  prints cell coordinates, so edits can be planned from the output.
* **Health check** replaces "the model should remember ~10 negative rules".
  Every rule the official SKILL.md states as prose and never verifies — banned
  spilling functions, the ``_xlfn.`` prefix set, percentages stored as whole
  numbers, unquoted sheet names — is checked mechanically here.

Always exits 0: the BiSheng executor discards stdout on a non-zero exit, which
would throw away the entire report.
"""

import argparse
import os
import re
import sys
import traceback

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, range_boundaries
except ImportError:  # pragma: no cover - only when the image is cut down
    print("[FATAL] openpyxl 不可用；无法体检。请让运维在后端环境确认依赖。")
    sys.exit(0)

# LibreOffice cannot evaluate these under any prefix, and — worse — on builds
# that do evaluate them, an openpyxl-written file carries no spill metadata, so
# only the top-left cell gets a value and the recalc reports zero errors.
BANNED = ["XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE", "LET", "LAMBDA", "TEXTSPLIT"]

# Post-2007 names Excel stores prefixed; openpyxl writes the string verbatim, so
# a bare name becomes a literal #NAME? in the delivered file.
NEEDS_XLFN = ["TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"]

ERROR_LITERALS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]

PLACEHOLDERS = ["待填", "待补", "TODO", "XXX", "xxx", "占位", "示例数据", "请填写", "TBD", "lorem"]

# A bare multiplier inside a formula is an assumption nobody can find later.
# `*1` and `/1` are exempt; `*1.15` is not, so the guard must reject a bare 1
# only when no digit or decimal point follows it.
HARDCODED_MULT_RE = re.compile(r"[*/]\s*(?!1(?![\d.]))(\d+\.\d+|\d{2,})(?![\d%])")

# Group 1 = the name inside quotes, group 2 = a bare name. Both must be
# NON-capturing at the alternation level, otherwise group(1) always matches and
# every "is it quoted?" test silently answers yes.
SHEET_REF_RE = re.compile(r"(?<![\w!])(?:'([^']+)'|([A-Za-z0-9_一-鿿]+))!")

CJK_RE = re.compile(r"[一-鿿]")

findings: list[tuple[str, str]] = []


def add(level: str, message: str) -> None:
    findings.append((level, message))


def cjk_width(text) -> int:
    if text is None:
        return 0
    return sum(2 if CJK_RE.match(ch) or "＀" <= ch <= "￯" else 1 for ch in str(text))


def needs_quoting(sheet_name: str) -> bool:
    """Whether a cross-sheet reference to this name *must* be single-quoted.

    Deliberately narrower than the quoting rule in ``xlsx_helpers.sheet_ref``:
    that one quotes liberally because quoting is always safe, while flagging
    something as an ERROR needs certainty. A pure-CJK name like ``利润预测`` is
    perfectly legal unquoted, so treating "non-ASCII" as "must quote" would
    false-positive on nearly every Chinese workbook.
    """
    return any(ch in sheet_name for ch in " -()&'+.") or (sheet_name[:1].isdigit() if sheet_name else False)


def fmt_value(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", "⏎")
    return text if len(text) <= 60 else text[:57] + "…"


# --------------------------------------------------------------------------- #
# Content read-out
# --------------------------------------------------------------------------- #


def dump_content(path: str, max_rows: int) -> None:
    wb = load_workbook(path, data_only=False)
    wb_values = None
    try:
        wb_values = load_workbook(path, data_only=True)
    except Exception:
        pass  # cached values are a bonus here; the formula pass is the required one

    print("=" * 72)
    print(f"内容 · {os.path.basename(path)} · {len(wb.sheetnames)} 个工作表")
    print("=" * 72)

    for name in wb.sheetnames:
        ws = wb[name]
        vs = wb_values[name] if wb_values is not None and name in wb_values.sheetnames else None
        print(f"\n## 工作表「{name}」  {ws.max_row} 行 × {ws.max_column} 列", end="")
        if ws.freeze_panes:
            print(f"  冻结={ws.freeze_panes}", end="")
        if ws.auto_filter and ws.auto_filter.ref:
            print(f"  筛选={ws.auto_filter.ref}", end="")
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f"  合并={len(merged)}处", end="")
        print()

        shown = 0
        for row in ws.iter_rows():
            if shown >= max_rows:
                print(f"   … 其余 {ws.max_row - shown} 行未显示（--max-rows 可调）")
                break
            cells = [c for c in row if c.value is not None]
            if not cells:
                continue
            parts = []
            for cell in cells:
                text = fmt_value(cell.value)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cached = vs[cell.coordinate].value if vs is not None else None
                    text = f"{text} →{fmt_value(cached) if cached is not None else '未重算'}"
                parts.append(f"{cell.coordinate}={text}")
            print("   " + " | ".join(parts))
            shown += 1
    wb.close()
    if wb_values is not None:
        wb_values.close()


# --------------------------------------------------------------------------- #
# Health checks
# --------------------------------------------------------------------------- #


def check_formulas(ws, name: str, sheet_names: set[str]) -> int:
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            where = f"{name}!{cell.coordinate}"

            if any(err in value for err in ERROR_LITERALS):
                hit = next(e for e in ERROR_LITERALS if e in value)
                add("ERROR", f"{where} 单元格里是错误值 {hit}（不是公式，是已经算坏的结果）")
                continue

            if any(p in value for p in PLACEHOLDERS):
                add("WARN", f"{where} 残留占位文本：{fmt_value(value)}")

            if not value.startswith("="):
                continue
            formula_count += 1
            upper = value.upper()

            for fn in BANNED:
                if re.search(rf"(?<![A-Z_.]){fn}\s*\(", upper):
                    add(
                        "ERROR",
                        f"{where} 用了 {fn}()——LibreOffice 重算不了，交付出去在用户 Excel 里可能是 #NAME?，"
                        f"且重算脚本抓不到。改用 INDEX/MATCH，或在 Python 里排序去重后直接写值。",
                    )
            for fn in NEEDS_XLFN:
                if re.search(rf"(?<![A-Z_.]){fn}\s*\(", upper) and f"_XLFN.{fn}" not in upper:
                    add("ERROR", f"{where} {fn}() 缺 `_xlfn.` 前缀 → 会变成字面量 #NAME?。写成 `_xlfn.{fn}(...)`")

            # Drive this off the real sheet names rather than trying to infer a
            # name from the bare match: `=假设 输入!$B$2` tokenises as the bare
            # name `输入`, so the space is simply not visible from the match.
            for sname in sheet_names:
                if not needs_quoting(sname):
                    continue
                if f"{sname}!" in value and f"'{sname}'!" not in value:
                    add(
                        "ERROR",
                        f"{where} 跨表引用「{sname}」名字含空格/特殊字符却没加单引号 → #VALUE!。写成 '{sname}'!",
                    )

            for match in SHEET_REF_RE.finditer(value):
                inner, bare = match.group(1), match.group(2)
                ref_name = inner or bare
                if not ref_name or ref_name in sheet_names:
                    continue
                if ref_name.upper() in {"TRUE", "FALSE"}:
                    continue
                # A bare token before `!` is only a sheet reference if it is not
                # the tail of a quoted name we already matched.
                if any(ref_name in s for s in sheet_names):
                    continue
                add("WARN", f"{where} 引用了工作表「{ref_name}」，但工作簿里没有这个表名")

            if HARDCODED_MULT_RE.search(value):
                add("WARN", f"{where} 公式里硬编码了系数：{fmt_value(value)} → 把它挪到有标签的假设单元格并引用它")

            if "/" in value and "IFERROR" not in upper and re.search(r"/\s*[A-Z]+\d", value):
                add("INFO", f"{where} 除法未加 IFERROR 保护：{fmt_value(value)}")
    return formula_count


def check_layout(ws, name: str) -> None:
    if ws.max_row <= 1 and ws.max_column <= 1 and ws["A1"].value is None:
        add("WARN", f"工作表「{name}」是空的——要么填内容，要么删掉它")
        return

    # Column width vs content: the ### bug, and it hides Chinese first.
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(letter)
        width = dim.width if dim is not None and dim.width else None
        if width is None:
            continue
        needed = 0
        for row_idx in range(1, min(ws.max_row, 500) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None or (isinstance(cell.value, str) and cell.value.startswith("=")):
                continue
            alignment = cell.alignment
            if alignment is not None and alignment.wrap_text:
                continue  # wrapped text grows the row, not the column
            needed = max(needed, cjk_width(cell.value))
        if needed and width < needed:
            add(
                "WARN",
                f"{name} 第 {letter} 列宽 {width:.0f} < 内容需要的 {needed}（中文按双宽算）→ 数字会显示成 ###",
            )

    # Merged ranges whose non-anchor cells carry values (openpyxl silently drops them).
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) == (min_row, min_col):
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.value is not None:
                    add("WARN", f"{name}!{rng} 合并区里非左上角单元格 {cell.coordinate} 仍有值，Excel 打开时会丢")
                    break

    if ws.max_row > 15 and not ws.freeze_panes:
        add("INFO", f"工作表「{name}」超过 15 行但没冻结表头，滚动后看不到列名")


def check_number_formats(ws, name: str) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fmt = cell.number_format or "General"

            # Percent format holding a value that is plainly not a fraction.
            if "%" in fmt and isinstance(cell.value, (int, float)) and abs(cell.value) > 1.5:
                add(
                    "ERROR",
                    f"{name}!{cell.coordinate} 百分比格式但存了 {cell.value} → 会显示成 "
                    f"{cell.value * 100:.0f}%。百分比要存小数（15% 存 0.15）",
                )

            # Big bare numbers with no format read as a wall of digits.
            if fmt == "General" and isinstance(cell.value, (int, float)) and abs(cell.value) >= 10000:
                add("INFO", f"{name}!{cell.coordinate} 金额 {cell.value} 没有数字格式，建议 `#,##0`")

            # Years must not be thousands-separated.
            if isinstance(cell.value, int) and 1900 <= cell.value <= 2100 and "#,##" in fmt:
                add("WARN", f"{name}!{cell.coordinate} 年份 {cell.value} 用了千分位格式，会显示成 {cell.value:,}")


def run_checks(path: str) -> None:
    wb = load_workbook(path, data_only=False)
    sheet_names = set(wb.sheetnames)
    total_formulas = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            continue  # chartsheet
        total_formulas += check_formulas(ws, name, sheet_names)
        check_layout(ws, name)
        check_number_formats(ws, name)

    if total_formulas:
        wb_values = load_workbook(path, data_only=True)
        uncached = 0
        for name in wb.sheetnames:
            ws, vs = wb[name], wb_values[name]
            if not hasattr(ws, "iter_rows"):
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        if vs[cell.coordinate].value is None:
                            uncached += 1
        wb_values.close()
        if uncached:
            add(
                "ERROR",
                f"{uncached}/{total_formulas} 个公式没有缓存值——openpyxl 写出的公式不带结果，"
                f"pandas 和多数预览器读到的是空。必须先跑 recalc_check.py 重算。",
            )
    wb.close()

    print()
    print("=" * 72)
    print(f"体检 · 公式 {total_formulas} 个")
    print("=" * 72)

    order = {"ERROR": 0, "WARN": 1, "INFO": 2}
    counts = {"ERROR": 0, "WARN": 0, "INFO": 0}
    for level, _ in findings:
        counts[level] += 1
    for level, message in sorted(findings, key=lambda f: order[f[0]]):
        print(f"[{level}] {message}")
    if not findings:
        print("（无发现）")

    print()
    print(f"ERROR {counts['ERROR']} · WARN {counts['WARN']} · INFO {counts['INFO']}")
    if counts["ERROR"]:
        print("结论: 不通过 —— ERROR 必须全部修完再交付。")
    elif counts["WARN"]:
        print("结论: 有条件通过 —— WARN 逐条复核，确认无误可交付。")
    else:
        print("结论: 通过。")


def main() -> None:
    parser = argparse.ArgumentParser(description="读出并体检一个 .xlsx")
    parser.add_argument("path")
    parser.add_argument("--content-only", action="store_true", help="只读内容，不体检")
    parser.add_argument("--check-only", action="store_true", help="只体检，不打印内容")
    parser.add_argument("--max-rows", type=int, default=40, help="每个表最多打印多少行（默认 40）")
    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f"[FATAL] 文件不存在: {args.path}")
        print("提示：代码执行器 cwd 就是工作区根，用相对路径 `output/x.xlsx`，不要写 `/output/x.xlsx`。")
        return

    try:
        if not args.check_only:
            dump_content(args.path, args.max_rows)
        if not args.content_only:
            run_checks(args.path)
    except Exception:
        # Report and still exit 0 — a traceback on stdout is useful; a non-zero
        # exit would make the executor discard this whole report.
        print("[FATAL] 体检过程本身出错：")
        traceback.print_exc(file=sys.stdout)


if __name__ == "__main__":
    main()
    sys.exit(0)
