from openpyxl import Workbook

from bisheng.knowledge.rag.pipeline.loader.utils.md_from_excel import (
    unmerge_and_read_sheet,
)


def test_merged_cell_value_is_kept_only_at_top_left():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Merged header"
    worksheet.merge_cells("A1:C2")
    worksheet["A3"] = "A"
    worksheet["B3"] = "B"
    worksheet["C3"] = "C"

    data_grid = unmerge_and_read_sheet(worksheet)

    assert data_grid == [
        ["Merged header", None, None],
        [None, None, None],
        ["A", "B", "C"],
    ]


def test_multiple_merged_ranges_do_not_repeat_values_in_covered_cells():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Horizontal"
    worksheet.merge_cells("A1:B1")
    worksheet["C1"] = "Vertical"
    worksheet.merge_cells("C1:C2")
    worksheet["A3"] = "A"
    worksheet["B3"] = "B"

    data_grid = unmerge_and_read_sheet(worksheet)

    assert data_grid == [
        ["Horizontal", None, "Vertical"],
        [None, None, None],
        ["A", "B", None],
    ]
