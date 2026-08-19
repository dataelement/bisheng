"""Bulk-volume tests for inspection-standard sync (10 depts × 10k records)."""

from __future__ import annotations

import time
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from bisheng.developer_token.domain.schemas import DeveloperTokenFileSyncRule
from bisheng.open_endpoints.domain.schemas.filelib_sync import FilelibSyncResponseData
from bisheng.open_endpoints.domain.services.filelib_sync_audit_writer import FilelibSyncAuditWriter
from bisheng.open_endpoints.domain.services.inspection_standard_excel_builder import (
    SHEET_ITEM_NAME,
    SHEET_STANDARD_NAME,
    build_inspection_standard_xlsx_bytes,
)
from bisheng.open_endpoints.domain.services.inspection_standard_sync_service import (
    InspectionStandardSyncService,
)
from test.open_endpoints.inspection_standard_bulk_factory import (
    DEFAULT_DEPT_COUNT,
    DEFAULT_RECORDS_PER_DEPT,
    build_bulk_payload_dict,
    build_bulk_request,
    build_create_dept_id,
)


pytestmark = pytest.mark.slow


def _fixed_rule(**target_overrides) -> DeveloperTokenFileSyncRule:
    target_space = {
        "mode": "fixed",
        "knowledge_id": 118,
        "folder_path": "点检标准",
        "dynamic_source": None,
    }
    target_space.update(target_overrides)
    return DeveloperTokenFileSyncRule.model_validate(
        {
            "category": {"code": "REPORT", "subcategory_code": "INSPECTION_STD"},
            "business_domain": {"mode": "fixed", "code": "MANUFACTURE", "dynamic_source": None},
            "target_space": target_space,
        }
    )


def _elapsed_seconds(start: float) -> float:
    return round(time.perf_counter() - start, 2)


def _build_mock_service() -> InspectionStandardSyncService:
    filelib_sync_service = SimpleNamespace(
        file_sync_rule=_fixed_rule(),
        knowledge_space_service=SimpleNamespace(
            find_or_create_folder_path_for_file_sync=AsyncMock(
                return_value=SimpleNamespace(id=9001),
            ),
            find_or_create_folder_for_file_sync=AsyncMock(
                return_value=SimpleNamespace(id=9002),
            ),
        ),
        repository=SimpleNamespace(
            find_knowledge_by_id=AsyncMock(return_value=SimpleNamespace(name="智能制造室(制造)")),
        ),
        sync_from_staged_file=AsyncMock(
            return_value=FilelibSyncResponseData(
                external_file_id="INSPECTION-STD-BULK",
                file_id=456,
                file_encoding="ENC-BULK",
                knowledge_id=118,
                knowledge_name="智能制造室(制造)",
                status=5,
            )
        ),
        request=None,
        login_user=SimpleNamespace(user_id=1, user_name="admin", tenant_id=1),
        token_id=42,
        token_name="BULK-TOKEN",
    )
    return InspectionStandardSyncService(filelib_sync_service=filelib_sync_service)


def test_bulk_payload_has_expected_volume():
    payload = build_bulk_payload_dict()
    standards = payload["data"]["check_standards"]
    items = payload["data"]["check_standard_items"]

    assert len(standards) == DEFAULT_DEPT_COUNT * DEFAULT_RECORDS_PER_DEPT
    assert len(items) == DEFAULT_DEPT_COUNT * DEFAULT_RECORDS_PER_DEPT
    assert standards[0]["CREATE_DEPT_ID"] == build_create_dept_id(0)
    assert standards[DEFAULT_RECORDS_PER_DEPT]["CREATE_DEPT_ID"] == build_create_dept_id(1)


def test_bulk_request_schema_validation():
    start = time.perf_counter()
    request = build_bulk_request()
    elapsed = _elapsed_seconds(start)

    assert len(request.data.check_standards) == 100_000
    assert len(request.data.check_standard_items) == 100_000
    assert request.data.check_standards[0].CREATE_DEPT_ID == "DEPT-00"
    print(f"\n[bulk] schema validation: 100k+100k records in {elapsed}s")


def test_bulk_build_groups_10_depts_10k_each():
    start = time.perf_counter()
    request = build_bulk_request()
    groups = InspectionStandardSyncService._build_groups(request)
    elapsed = _elapsed_seconds(start)

    assert len(groups) == DEFAULT_DEPT_COUNT
    for dept_idx, group in enumerate(groups):
        expected_dept = build_create_dept_id(dept_idx)
        assert group.create_dept_id == expected_dept
        assert len(group.check_standards) == DEFAULT_RECORDS_PER_DEPT
        assert len(group.check_standard_items) == DEFAULT_RECORDS_PER_DEPT

    print(f"\n[bulk] build_groups: {len(groups)} groups × {DEFAULT_RECORDS_PER_DEPT} in {elapsed}s")


def test_bulk_xlsx_single_dept_10k_rows():
    request = build_bulk_request(dept_indices=[0])
    group = InspectionStandardSyncService._build_groups(request)[0]

    start = time.perf_counter()
    content = build_inspection_standard_xlsx_bytes(
        check_standards=group.check_standards,
        check_standard_items=group.check_standard_items,
    )
    elapsed = _elapsed_seconds(start)

    workbook = load_workbook(filename=BytesIO(content))
    standard_sheet = workbook[SHEET_STANDARD_NAME]
    item_sheet = workbook[SHEET_ITEM_NAME]

    assert standard_sheet.max_row == DEFAULT_RECORDS_PER_DEPT + 1
    assert item_sheet.max_row == DEFAULT_RECORDS_PER_DEPT + 1
    assert len(content) > 1_000_000
    print(
        f"\n[bulk] xlsx single dept: rows={DEFAULT_RECORDS_PER_DEPT}, "
        f"bytes={len(content)}, elapsed={elapsed}s"
    )


def test_bulk_xlsx_all_10_depts_10k_each():
    request = build_bulk_request()
    groups = InspectionStandardSyncService._build_groups(request)

    sizes: list[int] = []
    start = time.perf_counter()
    for group in groups:
        content = build_inspection_standard_xlsx_bytes(
            check_standards=group.check_standards,
            check_standard_items=group.check_standard_items,
        )
        sizes.append(len(content))
        workbook = load_workbook(filename=BytesIO(content))
        assert workbook[SHEET_ITEM_NAME].max_row == DEFAULT_RECORDS_PER_DEPT + 1
    elapsed = _elapsed_seconds(start)

    assert len(sizes) == DEFAULT_DEPT_COUNT
    assert all(size > 1_000_000 for size in sizes)
    print(
        f"\n[bulk] xlsx all depts: groups={len(sizes)}, "
        f"bytes_per_file={[f'{s // 1024 // 1024}MB' for s in sizes]}, elapsed={elapsed}s"
    )


@pytest.mark.asyncio
async def test_bulk_sync_10_depts_mocked_without_real_xlsx():
    """Validate sync orchestration counts; mock xlsx builder to keep runtime reasonable."""
    request = build_bulk_request()
    service = _build_mock_service()
    captured_counts: list[tuple[int, int]] = []

    def _fake_xlsx(*, check_standards, check_standard_items):
        captured_counts.append((len(check_standards), len(check_standard_items)))
        return b"PK\x03\x04fake-xlsx"

    start = time.perf_counter()
    with (
        patch(
            "bisheng.open_endpoints.domain.services.inspection_standard_sync_service.build_inspection_standard_xlsx_bytes",
            side_effect=_fake_xlsx,
        ),
        patch.object(FilelibSyncAuditWriter, "write_inspection_batch_success", new_callable=AsyncMock),
        patch.object(FilelibSyncAuditWriter, "write_inspection_batch_failed", new_callable=AsyncMock),
    ):
        result = await service.sync(request)
    elapsed = _elapsed_seconds(start)

    assert result.group_count == DEFAULT_DEPT_COUNT
    assert len(result.files) == DEFAULT_DEPT_COUNT
    assert service.filelib_sync_service.sync_from_staged_file.await_count == DEFAULT_DEPT_COUNT
    assert len(captured_counts) == DEFAULT_DEPT_COUNT
    assert all(std_count == DEFAULT_RECORDS_PER_DEPT for std_count, _ in captured_counts)
    assert all(item_count == DEFAULT_RECORDS_PER_DEPT for _, item_count in captured_counts)
    assert {item.create_dept_id for item in result.files} == {
        build_create_dept_id(i) for i in range(DEFAULT_DEPT_COUNT)
    }
    print(f"\n[bulk] sync orchestration (mock xlsx): elapsed={elapsed}s")


@pytest.mark.asyncio
async def test_bulk_sync_single_dept_real_xlsx_10k():
    """End-to-end service path for one dept including real openpyxl generation."""
    request = build_bulk_request(dept_indices=[0])
    service = _build_mock_service()

    start = time.perf_counter()
    with (
        patch.object(FilelibSyncAuditWriter, "write_inspection_batch_success", new_callable=AsyncMock),
        patch.object(FilelibSyncAuditWriter, "write_inspection_batch_failed", new_callable=AsyncMock),
    ):
        result = await service.sync(request)
    elapsed = _elapsed_seconds(start)

    assert result.group_count == 1
    assert result.files[0].check_standard_count == DEFAULT_RECORDS_PER_DEPT
    assert result.files[0].check_standard_item_count == DEFAULT_RECORDS_PER_DEPT
    service.filelib_sync_service.sync_from_staged_file.assert_awaited_once()
    print(f"\n[bulk] sync single dept real xlsx: elapsed={elapsed}s")
