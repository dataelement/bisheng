import os
from uuid import uuid4

import openpyxl
import pandas as pd
from loguru import logger

from bisheng.common.errcode.knowledge import KnowledgeFileDamagedError
from bisheng.knowledge.rag.pipeline.loader.utils.xlsx_repair import repair_xlsx_styles


def xls_to_xlsx(xls_path):
    if not xls_path.lower().endswith(".xls"):
        return None

    if not os.path.exists(xls_path):
        return None

    try:
        xls_file = pd.ExcelFile(xls_path)
        sheets_to_write = {}

        # 2. Iterate through all worksheets, check if empty, and save non-empty content to the dictionary
        for sheet_name in xls_file.sheet_names:
            df = xls_file.parse(sheet_name)
            # df.empty will judge DataFrame No data (the number of rows is0）
            if not df.empty:
                sheets_to_write[sheet_name] = df
            else:
                #  Discard Blank Sheet
                pass

        # 3. Do not create a new file if there are any non-empty worksheets
        if not sheets_to_write:
            return None

        # 4. Write a new file if a non-empty worksheet exists
        xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            for sheet_name, df in sheets_to_write.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        return xlsx_path

    except Exception:
        logger.exception("xls_to_xlsx error: ")
        return None


def remove_characters(s, chars_to_remove=["\n", "\r"]):
    """
    Removes the specified character from the string.
    """
    if not isinstance(s, str):
        return s
    for char in chars_to_remove:
        s = s.replace(char, "")
    return s.strip()


def unmerge_and_read_sheet(sheet_obj):
    if sheet_obj.max_row == 0 or sheet_obj.max_column == 0:
        return []

    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column

    # 处理 merged cells
    merged_map = {}
    for merged_range in sheet_obj.merged_cells.ranges:
        min_col, min_row, max_col, max_row_r = merged_range.bounds
        val = sheet_obj.cell(row=min_row, column=min_col).value
        for r in range(min_row, max_row_r + 1):
            for c in range(min_col, max_col + 1):
                merged_map[(r, c)] = val

    data_grid = []
    empty_row_num = 0
    max_empty_rows = 50

    for r_idx, row in enumerate(sheet_obj.iter_rows()):
        row_data = []
        row_empty = True

        for c_idx, cell in enumerate(row):
            value = merged_map.get((r_idx + 1, c_idx + 1), cell.value)
            row_data.append(value)

            if value is not None and str(value).strip() != "":
                row_empty = False

        data_grid.append(row_data)

        if row_empty:
            empty_row_num += 1
            if empty_row_num > max_empty_rows:
                data_grid = data_grid[:-max_empty_rows]
                break
        else:
            empty_row_num = 0

    # 裁剪空列
    max_non_empty_col = 0
    for row in data_grid:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None and str(row[i]).strip() != "":
                max_non_empty_col = max(max_non_empty_col, i + 1)
                break

    if max_non_empty_col > 0:
        data_grid = [row[:max_non_empty_col] for row in data_grid]

    return data_grid


def render_markdown_row(row_values) -> str:
    """Render one table row. Single source of truth for row rendering.

    The char-budget planner measures rows with this exact function, so the
    budget can never drift from what actually gets written.
    """
    return "| " + " | ".join(remove_characters(str(v)) if v is not None else "" for v in row_values) + " |"


def render_markdown_separator(num_columns: int) -> str:
    return "|" + "---|" * num_columns


def generate_markdown_table_string(
    header_rows_list_of_lists,
    data_rows_list_of_lists,
    num_columns,
    separator_placement_index=1,
):
    """
    Generate from new rulesMarkdownTable String
    Automatically close purchase order afterheader_rows_list_of_listsIf empty, no headers and delimiters are generated.
    """
    md_lines = []

    # Handle headers and delimiters only if a header row is provided
    if header_rows_list_of_lists:
        pre_separator_header = header_rows_list_of_lists[:separator_placement_index]
        for row_values in pre_separator_header:
            md_lines.append(render_markdown_row(row_values))

        # Insert below the header in the first rowMarkdownSeparator
        if num_columns > 0:
            md_lines.append(render_markdown_separator(num_columns))

        post_separator_header = header_rows_list_of_lists[separator_placement_index:]
        for row_values in post_separator_header:
            md_lines.append(render_markdown_row(row_values))

    # Always process data rows
    for row_values in data_rows_list_of_lists:
        md_lines.append(render_markdown_row(row_values))

    return "\n".join(md_lines)


class ExcelRowTooLongError(ValueError):
    """A single table row does not fit in one chunk even on its own.

    Raised only when no ``long_row_splitter`` was supplied to degrade the row to
    plain text. Kept as a plain ValueError so this module stays free of the
    errcode layer (which pulls in fastapi); the loader translates it.
    """

    def __init__(self, sheet_index: str, row_number: int, row_chars: int, max_chars: int):
        self.sheet_index = sheet_index
        self.row_number = row_number
        self.row_chars = row_chars
        self.max_chars = max_chars
        super().__init__(
            f"sheet {sheet_index} row {row_number} renders to {row_chars} chars, over the {max_chars} chars budget"
        )


def _chunk_fixed_prefix_len(header_rows_as_lists, num_columns: int, append_header: bool) -> int:
    """Chunk length that does not depend on how many data rows the chunk holds.

    ``"\\n".join`` means ``len(chunk) = sum(len(line)) + (n_lines - 1)``, which
    factors into this fixed part plus ``len(render_markdown_row(row)) + 1`` per
    data row -- see ``_per_row_cost``.
    """
    if not append_header:
        # Pseudo-header mode: the group's own first row doubles as the header
        # line, so only the separator is fixed overhead.
        return len(render_markdown_separator(num_columns))
    if not header_rows_as_lists:
        # No header block at all: len = sum(len(line) + 1) - 1.
        return -1
    return (
        sum(len(render_markdown_row(one)) for one in header_rows_as_lists)
        + len(render_markdown_separator(num_columns))
        + len(header_rows_as_lists)
    )


def _per_row_cost(row_values) -> int:
    return len(render_markdown_row(row_values)) + 1


def _plan_chunks(indexed_data_rows, rows_per_markdown, fixed_prefix: int, max_chars: int | None):
    """Group data rows into chunk plans.

    Greedy accumulation *within* each ``rows_per_markdown`` slice, so
    ``rows_per_markdown`` becomes an upper bound instead of an exact count and
    every slice that already fits keeps byte-identical boundaries.

    Returns a list of ``("table", [(row_no, values), ...])`` /
    ``("long_row", (row_no, values))`` plans, in output order.
    """
    if rows_per_markdown and rows_per_markdown > 0:
        slices = [
            indexed_data_rows[i : i + rows_per_markdown] for i in range(0, len(indexed_data_rows), rows_per_markdown)
        ]
    else:
        slices = [indexed_data_rows]

    plans = []
    for one_slice in slices:
        if max_chars is None:
            plans.append(("table", one_slice))
            continue

        group, cost = [], fixed_prefix
        for indexed_row in one_slice:
            add = _per_row_cost(indexed_row[1])
            if group and cost + add > max_chars:
                plans.append(("table", group))
                group, cost = [], fixed_prefix
            if not group and fixed_prefix + add > max_chars:
                # Even alone this row blows the budget -- degrade it separately.
                plans.append(("long_row", indexed_row))
                continue
            group.append(indexed_row)
            cost += add
        if group:
            plans.append(("table", group))
    return plans


def _column_names(header_rows_as_lists, num_columns: int) -> list[str]:
    """Column labels for the plain-text degradation of an over-long row."""
    first_header = header_rows_as_lists[0] if header_rows_as_lists else []
    names = []
    for i in range(num_columns):
        raw = first_header[i] if i < len(first_header) else None
        name = remove_characters(str(raw)) if raw is not None else ""
        names.append(name or f"列{i + 1}")
    return names


def _render_long_row_chunks(
    row_values,
    header_rows_as_lists,
    num_columns: int,
    append_header: bool,
    long_row_splitter,
    max_chars: int,
) -> list[str]:
    """Degrade one over-long row to plain-text chunks.

    Rendered as ``column: value`` lines rather than character-cut markdown so the
    column/value association survives and no chunk ends in a dangling ``| cell``.
    Newlines inside cells are deliberately kept here (unlike table rendering) so
    the text splitter has real separators to cut on. Every fragment carries the
    table header as a prefix.
    """
    column_names = _column_names(header_rows_as_lists if append_header else [], num_columns)
    body_lines = []
    for i, value in enumerate(row_values):
        text = "" if value is None else str(value).strip()
        if not text:
            continue
        name = column_names[i] if i < len(column_names) else f"列{i + 1}"
        body_lines.append(f"{name}: {text}")
    body = "\n".join(body_lines)

    prefix = ""
    if append_header and header_rows_as_lists:
        prefix = generate_markdown_table_string(header_rows_as_lists, [], num_columns) + "\n"

    # The header prefix eats into the budget, so the body gets what is left.
    body_budget = max(1, max_chars - len(prefix))
    chunks = []
    for fragment in long_row_splitter(body, body_budget):
        if not fragment.strip():
            continue
        # Last resort: a splitter cannot break text that holds no separator at
        # all. Slicing plain text here is safe -- this row is no longer a table.
        for i in range(0, len(fragment), body_budget):
            chunks.append(prefix + fragment[i : i + body_budget])
    return chunks


def _write_chunk_file(output_dir, sheet_index: str, chunk_index: int, markdown_content: str) -> None:
    """Write one chunk file.

    The name must sort lexicographically in logical order -- ``ExcelLoader``
    derives chunk order purely from ``sorted(os.listdir(...))``. The old
    ``{sheet:02}{i:03}`` scheme silently reordered chunks past 999 ("001000" <
    "00999"), which a large sheet reaches on its own.
    """
    file_name = f"{str(sheet_index).zfill(3)}_{str(chunk_index).zfill(6)}.md"
    file_path = os.path.join(output_dir, file_name)
    # Not swallowed: a dropped chunk is silent data loss in the knowledge base.
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(markdown_content)
    logger.debug("chunk file saved: {}", file_path)


def process_dataframe_to_markdown_files(
    df,
    sheet_index: str,
    num_header_rows,
    rows_per_markdown,
    output_dir,
    append_header=True,
    max_chars: int | None = None,
    long_row_splitter=None,
):
    """
    - append_header=True: Tekan num_header_rows Separate the header and data.
    - append_header=False: All content is treated as data, table header is empty, ignored num_header_rows。
    - max_chars: per-chunk character budget. ``None`` keeps the legacy fixed-row
      behaviour byte for byte; otherwise ``rows_per_markdown`` becomes an upper
      bound and oversized groups are subdivided.
    - long_row_splitter: ``Callable[[str], list[str]]`` used to degrade a row that
      does not fit even alone. ``None`` makes that case raise
      ``ExcelRowTooLongError``.
    """
    if df.empty:
        logger.warning(f"  feed '{sheet_index}' DataDataFrameEmpty, skippingMarkdownBuat")
        return

    num_columns = df.shape[1]
    rows = df.shape[0]

    if rows == 0 or num_columns == 0:
        return

    header_block_df = pd.DataFrame()
    start_header_idx, end_header_idx = num_header_rows[0], num_header_rows[1]
    if start_header_idx >= rows:
        append_header = False

    # --- Core Logic Modified: According to append_header Decide how to split the data ---
    if append_header:
        # Handle header index outliers based on user rules
        if start_header_idx >= rows:
            logger.warning(
                f"Table Header Start Row {start_header_idx} Total lines exceeded {rows}. The first row will be used as the table header."
            )
            start_header_idx, end_header_idx = 0, 0
        elif end_header_idx >= rows:
            logger.warning(
                f"Table Header End Row {end_header_idx} Total lines exceeded {rows}. will be truncated to the last line."
            )
            end_header_idx = rows - 1

        # Make sure the index is legitimate
        if start_header_idx < 0:
            start_header_idx = 0
        if end_header_idx < start_header_idx:
            end_header_idx = start_header_idx

        try:
            header_slice = slice(start_header_idx, end_header_idx + 1)
            header_block_df = df.iloc[header_slice]
            data_block_df = df.drop(df.index[header_slice]).reset_index(drop=True)
            header_rows_as_lists = header_block_df.values.tolist()
            header_positions = set(range(rows)[header_slice])
            data_row_numbers = [one + 1 for one in range(rows) if one not in header_positions]
        except Exception as e:
            logger.error(
                f"  At Source '{sheet_index}' Index by header in [{start_header_idx}, {end_header_idx}] Error Splitting Data: {e}Skip"
            )
            return
    else:
        # when append_header are False , everything is treated as data and the header list is empty
        header_rows_as_lists = []
        data_block_df = df.reset_index(drop=True)
        data_row_numbers = [one + 1 for one in range(rows)]

    # --- Subsequent pagination logic ---
    if data_block_df.empty:
        if append_header and not header_block_df.empty:
            markdown_content = generate_markdown_table_string(header_rows_as_lists, [], num_columns)
            _write_chunk_file(output_dir, sheet_index, 0, markdown_content)
        return

    # Row numbers are the original 1-based sheet positions, so an over-long row is
    # reported by the row number the user sees in Excel.
    indexed_data_rows = list(zip(data_row_numbers, data_block_df.values.tolist(), strict=True))

    fixed_prefix = _chunk_fixed_prefix_len(header_rows_as_lists, num_columns, append_header)
    plans = _plan_chunks(indexed_data_rows, rows_per_markdown, fixed_prefix, max_chars)

    chunk_contents = []
    for kind, payload in plans:
        if kind == "table":
            group_rows = [values for _, values in payload]
            final_header_for_chunk = header_rows_as_lists
            final_data_for_chunk = group_rows

            # If no real header is attached and the current data block is not empty, the first row of data is used as the "pseudo header" to generate the delimiter
            if not append_header and group_rows:
                final_header_for_chunk = [group_rows[0]]
                final_data_for_chunk = group_rows[1:]

            chunk_contents.append(
                generate_markdown_table_string(final_header_for_chunk, final_data_for_chunk, num_columns)
            )
            continue

        row_number, row_values = payload
        row_chars = _per_row_cost(row_values) - 1
        if long_row_splitter is None:
            raise ExcelRowTooLongError(str(sheet_index), row_number, row_chars, max_chars)
        logger.warning(
            "sheet {} row {} is {} chars, over the {} chars budget; degrading it to plain-text chunks",
            sheet_index,
            row_number,
            row_chars,
            max_chars,
        )
        chunk_contents.extend(
            _render_long_row_chunks(
                row_values, header_rows_as_lists, num_columns, append_header, long_row_splitter, max_chars
            )
        )

    for i, markdown_content in enumerate(chunk_contents):
        _write_chunk_file(output_dir, sheet_index, i, markdown_content)


def is_list_of_lists_empty(data_list):
    """
    Determine if a 2D list is empty or contains only empty values (None, '')。
    """
    if not data_list:
        return True
    # Use any() and generator expressions for efficient judgment
    # any(row) Check for non-empty lines
    # any(cell is not None and cell != '' for cell in row) Check if there are non-empty cells in the row
    return not any(any(cell is not None and str(cell).strip() != "" for cell in row) for row in data_list)


def _load_workbook_with_style_repair(excel_path: str):
    """Open a workbook, retrying once on a repaired copy when styles are the blocker.

    openpyxl rejects a handful of style constructs Excel and WPS write happily
    (see ``xlsx_repair``), and refuses the whole workbook over them — a 28 MB
    customer file with 18 populated sheets was unreadable for three empty
    ``<fill/>`` elements. The retry only ever runs after a real failure, and only
    when the repair actually changed something, so a workbook that opens today
    takes exactly the path it takes today.
    """
    try:
        return openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    except Exception as first_error:
        repaired_path = repair_xlsx_styles(excel_path)
        if repaired_path is None:
            raise
        logger.warning(
            "excel load failed ({}), retrying on a style-repaired copy of {}",
            first_error,
            os.path.basename(excel_path),
        )
        try:
            return openpyxl.load_workbook(repaired_path, data_only=True, read_only=False)
        except Exception:
            # Report the ORIGINAL failure: the repaired copy is our artefact, and
            # its error would send whoever reads the log after the wrong file.
            logger.opt(exception=True).warning("style-repaired copy still unreadable")
            raise first_error from None


def excel_file_to_markdown(
    excel_path,
    num_header_rows,
    rows_per_markdown,
    output_dir,
    append_header=True,
    max_chars: int | None = None,
    long_row_splitter=None,
):
    logger.debug(f"\nStart ProcessingExcelDocumentation:'{excel_path}'")
    try:
        workbook = _load_workbook_with_style_repair(excel_path)
    except Exception as e:
        # Was `logger.debug(...); return`, which made an unopenable workbook look
        # like a workbook with no content: the loader returned zero documents, the
        # caller marked the file parsed successfully, and nothing was ever
        # indexed. The debug level meant the reason never even reached the log
        # file. A file we cannot open is a parse failure and must say so.
        logger.exception(f"Unable to load Excel doc '{excel_path}'")
        raise KnowledgeFileDamagedError(exception=e)

    # Workbook sheet order, with the markdown file prefix each sheet got (None when
    # the sheet had no cell data and therefore produced no markdown). The loader uses
    # this to place a sheet's embedded pictures next to that sheet's table chunks —
    # the numbered file names alone cannot say which sheet a chunk came from, because
    # empty sheets are skipped and do not consume a number.
    sheet_order: list[tuple[str, int | None]] = []
    sheet_index = 0
    for sheet_name in workbook.sheetnames:
        logger.debug(f"\n  (In work)ExcelWorksheet'{sheet_name}'...")
        sheet_obj = workbook[sheet_name]
        unmerged_data_list_of_lists = unmerge_and_read_sheet(sheet_obj)
        logger.debug(f"\n  <read all data>Excel<UNK>'{sheet_name}'...{len(unmerged_data_list_of_lists)}")

        # Using the new decision function
        if is_list_of_lists_empty(unmerged_data_list_of_lists):
            logger.debug(f"  Worksheet '{sheet_name}' Empty or no valid data, skipping.")
            sheet_order.append((sheet_name, None))
            continue

        df = pd.DataFrame(unmerged_data_list_of_lists)
        df.fillna("", inplace=True)
        if df.empty:
            logger.debug(f"  Worksheet '{sheet_name}' Empty after processingDataFrameSkip")
            sheet_order.append((sheet_name, None))
            continue

        process_dataframe_to_markdown_files(
            df,
            str(sheet_index),
            num_header_rows,
            rows_per_markdown,
            output_dir,
            append_header=append_header,
            max_chars=max_chars,
            long_row_splitter=long_row_splitter,
        )
        sheet_order.append((sheet_name, sheet_index))
        sheet_index += 1

    if workbook:
        workbook.close()
    logger.debug(f"\nExcelDoc. '{excel_path}' Process Completed.")
    return sheet_order


def csv_file_to_markdown(
    csv_path,
    num_header_rows,
    rows_per_markdown,
    output_dir,
    csv_encoding="utf-8",
    csv_delimiter=",",
    append_header=True,
    max_chars: int | None = None,
    long_row_splitter=None,
):
    logger.debug(f"\nStart ProcessingCSVDocumentation:'{csv_path}'")
    try:
        df = pd.read_csv(
            csv_path,
            header=None,
            dtype=str,
            encoding=csv_encoding,
            sep=csv_delimiter,
            keep_default_na=False,
        )
        df.fillna("", inplace=True)

    except pd.errors.EmptyDataError:
        logger.debug(f"Error: CSVDoc. '{csv_path}' Empty")
        return
    except FileNotFoundError:
        logger.debug(f"Error: CSVDoc. '{csv_path}' Nothing found.")
        return
    except Exception as e:
        logger.debug(f"Error: UnreadableCSVDoc. '{csv_path}'Reason: {e}")
        return

    if df.empty:
        logger.debug(f"CSVDoc. '{csv_path}' Empty or empty after processing, skipping.")
        return

    process_dataframe_to_markdown_files(
        df,
        "0",
        num_header_rows,
        rows_per_markdown,
        output_dir,
        append_header,
        max_chars=max_chars,
        long_row_splitter=long_row_splitter,
    )
    logger.debug(f"\nCSVDoc. '{csv_path}' Process Completed.")


def convert_file_to_markdown(
    input_file_path,
    num_header_rows,
    rows_per_markdown,
    base_output_dir="output_markdown_files",
    csv_encoding="utf-8",
    csv_delimiter=",",
    append_header=True,
    max_chars: int | None = None,
    long_row_splitter=None,
):
    """
    will be Excel OR CSV Convert files to multiple Markdown files.

    ``max_chars`` turns ``rows_per_markdown`` into an upper bound: groups that
    would exceed the budget are subdivided. ``None`` keeps the legacy fixed-row
    behaviour byte for byte.

    Returns the workbook's sheet order as ``[(sheet_name, markdown prefix or None)]``
    for Excel input; an empty list for CSV (a single, nameless sheet) and on failure.
    """
    if not os.path.exists(input_file_path):
        logger.debug(f"Error: Input file '{input_file_path}' Nothing found.")
        return []

    if not os.path.exists(base_output_dir):
        os.makedirs(base_output_dir)
        logger.debug(f"To create an output directory:'{base_output_dir}'")

    _, file_extension = os.path.splitext(input_file_path)
    file_extension = file_extension.lower()
    if file_extension == ".xls":
        input_file_path = xls_to_xlsx(input_file_path)

    if file_extension in [".xlsx", ".xls"]:
        return excel_file_to_markdown(
            input_file_path,
            num_header_rows,
            rows_per_markdown,
            base_output_dir,
            append_header,
            max_chars=max_chars,
            long_row_splitter=long_row_splitter,
        )
    if file_extension == ".csv":
        csv_file_to_markdown(
            input_file_path,
            num_header_rows,
            rows_per_markdown,
            base_output_dir,
            csv_encoding,
            csv_delimiter,
            append_header,
            max_chars=max_chars,
            long_row_splitter=long_row_splitter,
        )
        return []
    logger.debug(
        f"Error: Unsupported file type '{file_extension}'Please provide user. Excel (.xlsx, .xls) OR CSV (.csv) files."
    )
    return []


def handler(
    cache_dir,
    file_name: str,
    header_rows: list[int] = [0, 1],
    data_rows: int = 12,
    append_header=True,
):
    """
    The main function that handles file conversions.
    """
    doc_id = uuid4()
    md_file_name = f"{cache_dir}/{doc_id}"

    convert_file_to_markdown(
        input_file_path=file_name,
        base_output_dir=md_file_name,
        num_header_rows=header_rows,
        rows_per_markdown=data_rows,
        append_header=append_header,
    )
    return md_file_name, None, doc_id


if __name__ == "__main__":
    # Define test parameters
    test_cache_dir = "/Users/zhangguoqing/Downloads/tmp"
    test_file_name = "/Users/zhangguoqing/Downloads/战略部数据爬取表.xlsx"
    # Test append_header=True and the index is out of bounds
    test_header_rows = [0, 0]  # start_header_index Out of Scope
    test_data_rows = 2
    test_append_header = True

    # Recall handler Function
    print("--- Test Scenarios: append_header=True, Table header index out of bounds ---")
    handler(
        cache_dir=test_cache_dir,
        file_name=test_file_name,
        header_rows=test_header_rows,
        data_rows=test_data_rows,
        append_header=test_append_header,
    )
