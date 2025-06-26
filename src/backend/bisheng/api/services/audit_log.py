import csv
from datetime import datetime
import json
import time
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Any, List, Optional
from uuid import UUID

import pandas as pd
from celery.schedules import crontab
from langchain_core.language_models import BaseChatModel
from loguru import logger
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from redbeat import RedBeatSchedulerEntry
from sqlalchemy import or_
from sqlmodel import select

from bisheng.api.errcode.base import UnAuthorizedError
from bisheng.api.services.user_service import UserPayload
from bisheng.api.v1.schema.chat_schema import AppChatList
from bisheng.api.v1.schema.workflow import WorkflowEventType
from bisheng.api.v1.schemas import resp_200
from bisheng.database.models.assistant import AssistantDao, Assistant
from bisheng.database.models.audit_log import AuditLog, SystemId, EventType, ObjectType, AuditLogDao
from bisheng.database.models.flow import FlowDao, Flow, FlowType
from bisheng.database.models.group import Group
from bisheng.database.models.group_resource import GroupResourceDao, ResourceTypeEnum
from bisheng.database.models.knowledge import KnowledgeDao, Knowledge
from bisheng.database.models.message import ChatMessageDao, LikedType
from bisheng.database.models.role import Role
from bisheng.database.models.session import MessageSessionDao, SensitiveStatus
from bisheng.database.models.user import UserDao, User
from bisheng.database.models.user_group import UserGroupDao
from bisheng.settings import settings
from bisheng.utils import generate_uuid
from bisheng.api.services.knowledge_imp import extract_code_blocks
from bisheng.api.services.llm import LLMService
from bisheng.api.services.user_service import UserPayload
from bisheng.api.v1.schema.audit import ReviewSessionConfig
from bisheng.api.v1.schema.chat_schema import AppChatList
from bisheng.api.v1.schemas import resp_200
from bisheng.database.base import generate_uuid, session_getter
from bisheng.database.models.assistant import AssistantDao, Assistant
from bisheng.database.models.audit_log import AuditLog, SystemId, EventType, ObjectType, AuditLogDao
from bisheng.database.models.config import ConfigDao, ConfigKeyEnum, Config
from bisheng.database.models.flow import FlowDao, Flow, FlowType
from bisheng.database.models.group import Group, GroupDao
from bisheng.database.models.group_resource import GroupResourceDao, ResourceTypeEnum
from bisheng.database.models.knowledge import KnowledgeDao, Knowledge
from bisheng.database.models.message import MessageDao, ChatMessageDao, ChatMessage
from bisheng.database.models.role import Role
from bisheng.database.models.session import MessageSessionDao, ReviewStatus, MessageSession
from bisheng.database.models.user import UserDao, User
from bisheng.database.models.user_group import UserGroupDao
from bisheng.utils.minio_client import MinioClient

# TODO merge_check 2整个文件都要检查，特别是session查询相关


class AuditLogService:

    @classmethod
    def get_audit_log(cls, login_user: UserPayload, group_ids, operator_ids, start_time, end_time,
                      system_id, event_type, page, limit) -> Any:
        groups = group_ids
        if not login_user.is_admin():
            groups = [str(one.group_id) for one in UserGroupDao.get_user_power_group(login_user.user_id)]
            # 不是任何用戶组的管理员
            if not groups:
                return UnAuthorizedError.return_resp()
            # 将筛选条件的group_id和管理员有权限的groups做交集
            if group_ids:
                groups = list(set(groups) & set(group_ids))
                if not groups:
                    return UnAuthorizedError.return_resp()
        data, total = AuditLogDao.get_audit_logs(groups, operator_ids, start_time, end_time, system_id, event_type,
                                                 page, limit)
        return resp_200(data={'data': data, 'total': total})

    @classmethod
    def get_all_operators(cls, login_user: UserPayload) -> Any:
        groups = []
        if not login_user.is_admin():
            groups = [one.group_id for one in UserGroupDao.get_user_power_group(login_user.user_id)]

        data = AuditLogDao.get_all_operators(groups)
        res = {}
        for one in data:
            if not one[1]:
                continue
            res[one[0]] = {'user_id': one[0], 'user_name': one[1]}
        return resp_200(data=list(res.values()))

    @classmethod
    def _chat_log(cls, user: UserPayload, ip_address: str, event_type: EventType, object_type: ObjectType,
                  object_id: str, object_name: str, resource_type: ResourceTypeEnum):
        # 获取资源所属的分组
        groups = GroupResourceDao.get_resource_group(resource_type, object_id)
        group_ids = [one.group_id for one in groups]
        audit_log = AuditLog(
            operator_id=user.user_id,
            operator_name=user.user_name,
            group_ids=group_ids,
            system_id=SystemId.CHAT.value,
            event_type=event_type.value,
            object_type=object_type.value,
            object_id=object_id,
            object_name=object_name,
            ip_address=ip_address,
        )
        AuditLogDao.insert_audit_logs([audit_log])

    @classmethod
    def create_chat_assistant(cls, user: UserPayload, ip_address: str, assistant_id: str, assistant_info = None):
        """
        新建助手会话的审计日志
        """
        logger.info(f"act=create_chat_assistant user={user.user_name} ip={ip_address} assistant={assistant_id}")
        # 获取助手详情
        obj = UUID(assistant_id)  # DONE merge_check
        assistant_info = AssistantDao.get_one_assistant(str(obj))
        if not assistant_info:
            assistant_info = AssistantDao.get_one_assistant(obj.hex)
        cls._chat_log(user, ip_address, EventType.CREATE_CHAT, ObjectType.ASSISTANT,
                      assistant_id, assistant_info.name, ResourceTypeEnum.ASSISTANT)

    @classmethod
    def create_chat_flow(cls, user: UserPayload, ip_address: str, flow_id: str, flow_info: Flow = None):
        """
        新建技能会话的审计日志
        """
        logger.info(f"act=create_chat_flow user={user.user_name} ip={ip_address} flow={flow_id}")
        if not flow_info:
            flow_info = FlowDao.get_flow_by_id(flow_id)
        cls._chat_log(user, ip_address, EventType.CREATE_CHAT, ObjectType.FLOW,
                      flow_id, flow_info.name, ResourceTypeEnum.FLOW)

    @classmethod
    def create_chat_workflow(cls, user: UserPayload, ip_address: str, flow_id: str, flow_info=None):
        """
        新建工作流会话的审计日志
        """
        logger.info(f"act=create_chat_workflow user={user.user_name} ip={ip_address} flow={flow_id}")
        if not flow_info:
            flow_info = FlowDao.get_flow_by_id(flow_id)
        cls._chat_log(user, ip_address, EventType.CREATE_CHAT, ObjectType.WORK_FLOW,
                      flow_id, flow_info.name, ResourceTypeEnum.WORK_FLOW)

    @classmethod
    def delete_chat_flow(cls, user: UserPayload, ip_address: str, flow_info: Flow):
        """
        删除技能会话的审计日志
        """
        logger.info(f"act=delete_chat_flow user={user.user_name} ip={ip_address} flow={flow_info.id}")
        cls._chat_log(user, ip_address, EventType.DELETE_CHAT, ObjectType.FLOW,
                      flow_info.id, flow_info.name, ResourceTypeEnum.FLOW)

    @classmethod
    def delete_chat_workflow(cls, user: UserPayload, ip_address: str, flow_info: Flow):
        """
        删除技能会话的审计日志
        """
        logger.info(f"act=delete_chat_workflow user={user.user_name} ip={ip_address} flow={flow_info.id}")
        cls._chat_log(user, ip_address, EventType.DELETE_CHAT, ObjectType.WORK_FLOW,
                      flow_info.id, flow_info.name, ResourceTypeEnum.WORK_FLOW)

    @classmethod
    def delete_chat_assistant(cls, user: UserPayload, ip_address: str, assistant_info: Assistant):
        """
        删除助手会话的审计日志
        """
        logger.info(f"act=delete_assistant_flow user={user.user_name} ip={ip_address} assistant={assistant_info.id}")
        cls._chat_log(user, ip_address, EventType.DELETE_CHAT, ObjectType.ASSISTANT,
                      assistant_info.id, assistant_info.name, ResourceTypeEnum.ASSISTANT)

    @classmethod
    def _build_log(cls, user: UserPayload, ip_address: str, event_type: EventType, object_type: ObjectType,
                   object_id: str,
                   object_name: str, resource_type: ResourceTypeEnum):
        """
        构建模块的审计日志
        """
        # 获取资源属于哪些用户组
        groups = GroupResourceDao.get_resource_group(resource_type, object_id)
        group_ids = [one.group_id for one in groups]

        # 插入审计日志
        audit_log = AuditLog(
            operator_id=user.user_id,
            operator_name=user.user_name,
            group_ids=group_ids,
            system_id=SystemId.BUILD.value,
            event_type=event_type.value,
            object_type=object_type.value,
            object_id=object_id,
            object_name=object_name,
            ip_address=ip_address,
        )
        AuditLogDao.insert_audit_logs([audit_log])

    @classmethod
    def create_build_flow(cls, user: UserPayload, ip_address: str, flow_id: str, flow_type: Optional[int] = None):
        """
        新建技能的审计日志
        """
        obj_type = ObjectType.FLOW
        rs_type = ResourceTypeEnum.FLOW
        if flow_type == FlowType.WORKFLOW.value:
            obj_type = ObjectType.WORK_FLOW
            rs_type = ResourceTypeEnum.WORK_FLOW
        logger.info(f"act=create_build_flow user={user.user_name} ip={ip_address} flow={flow_id}")
        flow_info = FlowDao.get_flow_by_id(flow_id)
        cls._build_log(user, ip_address, EventType.CREATE_BUILD, obj_type,
                       flow_info.id, flow_info.name, rs_type)

    @classmethod
    def update_build_flow(cls, user: UserPayload, ip_address: str, flow_id: str, flow_type: Optional[int] = None):
        """
        更新技能的审计日志
        """
        obj_type = ObjectType.FLOW
        rs_type = ResourceTypeEnum.FLOW
        if flow_type == FlowType.WORKFLOW.value:
            obj_type = ObjectType.WORK_FLOW
            rs_type = ResourceTypeEnum.WORK_FLOW
        logger.info(f"act=update_build_flow user={user.user_name} ip={ip_address} flow={flow_id}")
        flow_info = FlowDao.get_flow_by_id(flow_id)
        cls._build_log(user, ip_address, EventType.UPDATE_BUILD, obj_type,
                       flow_info.id, flow_info.name, rs_type)

    @classmethod
    def delete_build_flow(cls, user: UserPayload, ip_address: str, flow_info: Flow, flow_type: Optional[int] = None):
        """
        删除技能的审计日志
        """
        obj_type = ObjectType.FLOW
        rs_type = ResourceTypeEnum.FLOW
        if flow_type == FlowType.WORKFLOW.value:
            obj_type = ObjectType.WORK_FLOW
            rs_type = ResourceTypeEnum.WORK_FLOW
        logger.info(f"act=delete_build_flow user={user.user_name} ip={ip_address} flow={flow_info.id}")
        cls._build_log(user, ip_address, EventType.DELETE_BUILD, obj_type,
                       flow_info.id, flow_info.name, rs_type)

    @classmethod
    def create_build_assistant(cls, user: UserPayload, ip_address: str, assistant_id: str):
        """
        新建助手的审计日志
        """
        logger.info(f"act=create_build_assistant user={user.user_name} ip={ip_address} assistant={assistant_id}")
        assistant_info = AssistantDao.get_one_assistant(assistant_id)
        cls._build_log(user, ip_address, EventType.CREATE_BUILD, ObjectType.ASSISTANT,
                       assistant_info.id, assistant_info.name, ResourceTypeEnum.ASSISTANT)

    @classmethod
    def update_build_assistant(cls, user: UserPayload, ip_address: str, assistant_id: str):
        """
        更新助手的审计日志
        """
        logger.info(f"act=update_build_assistant user={user.user_name} ip={ip_address} assistant={assistant_id}")
        assistant_info = AssistantDao.get_one_assistant(assistant_id)

        cls._build_log(user, ip_address, EventType.UPDATE_BUILD, ObjectType.ASSISTANT,
                       assistant_info.id, assistant_info.name, ResourceTypeEnum.ASSISTANT)

    @classmethod
    def delete_build_assistant(cls, user: UserPayload, ip_address: str, assistant_id: str):
        """
        删除助手的审计日志
        """
        logger.info(f"act=delete_build_assistant user={user.user_name} ip={ip_address} assistant={assistant_id}")
        assistant_info = AssistantDao.get_one_assistant(assistant_id)

        cls._build_log(user, ip_address, EventType.DELETE_BUILD, ObjectType.ASSISTANT,
                       assistant_info.id, assistant_info.name, ResourceTypeEnum.ASSISTANT)

    @classmethod
    def _knowledge_log(cls, user: UserPayload, ip_address: str, event_type: EventType, object_type: ObjectType,
                       object_id: str, object_name: str, resource_type: ResourceTypeEnum, resource_id: str):
        """
        知识库模块的日志
        """
        # 获取资源属于哪些用户组
        groups = GroupResourceDao.get_resource_group(resource_type, resource_id)
        group_ids = [one.group_id for one in groups]

        # 插入审计日志
        audit_log = AuditLog(
            operator_id=user.user_id,
            operator_name=user.user_name,
            group_ids=group_ids,
            system_id=SystemId.KNOWLEDGE.value,
            event_type=event_type.value,
            object_type=object_type.value,
            object_id=object_id,
            object_name=object_name,
            ip_address=ip_address,
        )
        AuditLogDao.insert_audit_logs([audit_log])

    @classmethod
    def create_knowledge(cls, user: UserPayload, ip_address: str, knowledge_id: int):
        """
        新建知识库的审计日志
        """
        logger.info(f"act=create_knowledge user={user.user_name} ip={ip_address} knowledge={knowledge_id}")
        knowledge_info = KnowledgeDao.query_by_id(knowledge_id)
        cls._knowledge_log(user, ip_address, EventType.CREATE_KNOWLEDGE, ObjectType.KNOWLEDGE,
                           str(knowledge_id), knowledge_info.name, ResourceTypeEnum.KNOWLEDGE, str(knowledge_id))

    @classmethod
    def delete_knowledge(cls, user: UserPayload, ip_address: str, knowledge: Knowledge):
        """
        删除知识库的审计日志
        """
        logger.info(f"act=delete_knowledge user={user.user_name} ip={ip_address} knowledge={knowledge.id}")
        cls._knowledge_log(user, ip_address, EventType.DELETE_KNOWLEDGE, ObjectType.KNOWLEDGE,
                           str(knowledge.id), knowledge.name, ResourceTypeEnum.KNOWLEDGE, str(knowledge.id))

    @classmethod
    def upload_knowledge_file(cls, user: UserPayload, ip_address: str, knowledge_id: int, file_name: str):
        """
        知识库上传文件的审计日志
        """
        logger.info(f"act=upload_knowledge_file user={user.user_name} ip={ip_address}"
                    f" knowledge={knowledge_id} file={file_name}")
        cls._knowledge_log(user, ip_address, EventType.UPLOAD_FILE, ObjectType.FILE,
                           str(knowledge_id), file_name, ResourceTypeEnum.KNOWLEDGE, str(knowledge_id))

    @classmethod
    def delete_knowledge_file(cls, user: UserPayload, ip_address: str, knowledge_id: int, file_name: str):
        """
        知识库删除文件的审计日志
        """
        logger.info(f"act=delete_knowledge_file user={user.user_name} ip={ip_address}"
                    f" knowledge={knowledge_id} file={file_name}")
        cls._knowledge_log(user, ip_address, EventType.DELETE_FILE, ObjectType.FILE,
                           str(knowledge_id), file_name, ResourceTypeEnum.KNOWLEDGE, str(knowledge_id))

    @classmethod
    def _system_log(cls, user: UserPayload, ip_address: str, group_ids: List[int], event_type: EventType,
                    object_type: ObjectType, object_id: str, object_name: str, note: str = ''):

        audit_log = AuditLog(
            operator_id=user.user_id,
            operator_name=user.user_name,
            group_ids=group_ids,
            system_id=SystemId.SYSTEM.value,
            event_type=event_type.value,
            object_type=object_type.value,
            object_id=object_id,
            object_name=object_name,
            ip_address=ip_address,
            note=note,
        )
        AuditLogDao.insert_audit_logs([audit_log])

    @classmethod
    def update_user(cls, user: UserPayload, ip_address: str, user_id: int, group_ids: List[int], note: str):
        """
        修改用户的用户组和角色
        """
        logger.info(f"act=update_system_user user={user.user_name} ip={ip_address} user_id={user_id} note={note}")
        user_info = UserDao.get_user(user_id)
        cls._system_log(user, ip_address, group_ids, EventType.UPDATE_USER,
                        ObjectType.USER_CONF, str(user_id), user_info.user_name, note)

    @classmethod
    def forbid_user(cls, user: UserPayload, ip_address: str, user_info: User):
        """
        user: 操作用户
        user_info: 被操作用户
        """
        logger.info(f"act=forbid_user user={user.user_name} ip={ip_address} user_id={user.user_id}")
        # 获取用户所属的分组
        user_group = UserGroupDao.get_user_group(user_info.user_id)
        user_group = [one.group_id for one in user_group]
        cls._system_log(user, ip_address, user_group, EventType.FORBID_USER,
                        ObjectType.USER_CONF, str(user_info.user_id), user_info.user_name)

    @classmethod
    def recover_user(cls, user: UserPayload, ip_address: str, user_info: User):
        logger.info(f"act=recover_user user={user.user_name} ip={ip_address} user_id={user_info.user_id}")
        # 获取用户所属的分组
        user_group = UserGroupDao.get_user_group(user_info.user_id)
        user_group = [one.group_id for one in user_group]
        cls._system_log(user, ip_address, user_group, EventType.RECOVER_USER,
                        ObjectType.USER_CONF, str(user_info.user_id), user_info.user_name)

    @classmethod
    def create_user_group(cls, user: UserPayload, ip_address: str, group_info: Group):
        logger.info(f"act=create_user_group user={user.user_name} ip={ip_address} group_id={group_info.id}")
        cls._system_log(user, ip_address, [group_info.id], EventType.CREATE_USER_GROUP,
                        ObjectType.USER_GROUP_CONF, str(group_info.id), group_info.group_name)

    @classmethod
    def update_user_group(cls, user: UserPayload, ip_address: str, group_info: Group):
        logger.info(f"act=update_user_group user={user.user_name} ip={ip_address} group_id={group_info.id}")
        # 获取用户组信息
        cls._system_log(user, ip_address, [group_info.id], EventType.UPDATE_USER_GROUP,
                        ObjectType.USER_GROUP_CONF, str(group_info.id), group_info.group_name)

    @classmethod
    def delete_user_group(cls, user: UserPayload, ip_address: str, group_info: Group):
        logger.info(f"act=delete_user_group user={user.user_name} ip={ip_address} group_id={group_info.id}")
        # 获取用户组信息
        cls._system_log(user, ip_address, [group_info.id], EventType.DELETE_USER_GROUP,
                        ObjectType.USER_GROUP_CONF, str(group_info.id), group_info.group_name)

    @classmethod
    def create_role(cls, user: UserPayload, ip_address: str, role: Role):
        logger.info(f"act=create_role user={user.user_name} ip={ip_address} role_id={role.id}")

        cls._system_log(user, ip_address, [role.group_id], EventType.CREATE_ROLE,
                        ObjectType.ROLE_CONF, str(role.id), role.role_name)

    @classmethod
    def update_role(cls, user: UserPayload, ip_address: str, role: Role):
        logger.info(f"act=update_role user={user.user_name} ip={ip_address} role_id={role.id}")

        cls._system_log(user, ip_address, [role.group_id], EventType.UPDATE_ROLE,
                        ObjectType.ROLE_CONF, str(role.id), role.role_name)

    @classmethod
    def delete_role(cls, user: UserPayload, ip_address: str, role: Role):
        logger.info(f"act=delete_role user={user.user_name} ip={ip_address} role_id={role.id}")

        cls._system_log(user, ip_address, [role.group_id], EventType.DELETE_ROLE,
                        ObjectType.ROLE_CONF, str(role.id), role.role_name)

    @classmethod
    def user_login(cls, user: UserPayload, ip_address: str):
        logger.info(f"act=user_login user={user.user_name} ip={ip_address} user_id={user.user_id}")
        # 获取用户所属的分组
        user_group = UserGroupDao.get_user_group(user.user_id)
        user_group = [one.group_id for one in user_group]
        cls._system_log(user, ip_address, user_group, EventType.USER_LOGIN,
                        ObjectType.NONE, '', '')

    @classmethod
    def get_session_config(cls) -> ReviewSessionConfig:
        ret = {}
        config = ConfigDao.get_config(ConfigKeyEnum.REVIEW_SESSION_CONFIG)
        if config:
            ret = json.loads(config.value)
        return ReviewSessionConfig(**ret)

    @classmethod
    def update_session_config(cls, user: UserPayload, data: ReviewSessionConfig) -> ReviewSessionConfig:
        from bisheng.worker import bisheng_celery
        if data.flag:
            hour, minute = data.get_hour_minute()
            schedule = {
                'hour': hour,
                'minute': minute
            }
            if day_of_week := data.get_celery_crontab_week() is not None:
                schedule['day_of_week'] = day_of_week

            beat_task = RedBeatSchedulerEntry(name='review_session_message',
                                              task='bisheng.worker.audit.tasks.review_session_message',
                                              schedule=crontab(**schedule),
                                              app=bisheng_celery)
            beat_task.delete()
            beat_task.save()
        ConfigDao.insert_or_update(Config(key=ConfigKeyEnum.REVIEW_SESSION_CONFIG.value, value=json.dumps(data.dict())))
        return data

    @classmethod
    def get_filter_flow_ids(cls, user: UserPayload, flow_ids: List[str], group_ids: List[str]) -> (bool, List):
        """ 通过flow_ids和group_ids获取最终的 技能id筛选条件 """
        flow_ids = [UUID(one).hex for one in flow_ids]  # UUID check done
        group_admins = []
        logger.info(f"flow_ids {flow_ids} | group_ids {group_ids}")
        if not user.is_admin():
            user_groups = UserGroupDao.get_user_power_group(user.user_id)
            # 不是用户组管理员，没有权限
            if not user_groups:
                raise UnAuthorizedError.http_exception()
            group_admins = [one.group_id for one in user_groups]
        # 分组id做交集
        if group_ids:
            if group_admins:
                # 查询了不属于用户管理的用户组，返回为空
                group_admins = list(set(group_admins) & set(group_ids))
                if len(group_admins) == 0:
                    return False, []
            else:
                group_admins = group_ids

        # 获取分组下所有的应用ID
        group_flows = []
        if group_admins:
            group_flows = GroupResourceDao.get_groups_resource(group_admins,
                                                               resource_types=[ResourceTypeEnum.FLOW,
                                                                               ResourceTypeEnum.WORK_FLOW,
                                                                               ResourceTypeEnum.ASSISTANT])
            # 用户管理下的用户组没有资源
            if not group_flows:
                return False, []
            group_flows = [one.third_id for one in group_flows]

        # 获取最终的技能ID限制列表
        filter_flow_ids = []
        if flow_ids and group_flows:
            filter_flow_ids = list(set(group_flows) & set(flow_ids))
            if not filter_flow_ids:
                return False, []
        elif flow_ids:
            filter_flow_ids = flow_ids
        elif group_flows:
            filter_flow_ids = group_flows
        return True, filter_flow_ids

    @classmethod
    def get_session_list(cls, user: UserPayload, flow_ids: List[str], user_ids: List[int], group_ids: List[int],
                         start_date: datetime, end_date: datetime,
                         feedback: str,
                         sensitive_status: List[SensitiveStatus] = None,
                         review_status: List[ReviewStatus] = None,
                         category: List[str] = None,
                         page: int = 1, page_size: int = 10, keyword=None) -> (list, int):
        # flag, filter_flow_ids = cls.get_filter_flow_ids(user, flow_ids, group_ids)
        # if not flag:
        #    return [], 0
        # ---

        filter_flow_ids = flow_ids
        for one in flow_ids:
            if one != one.replace("-", ""):
                filter_flow_ids.append(one.replace("-", ""))
        all_user = UserGroupDao.get_groups_user(group_ids)
        logger.info(f"get_session_list user_ids {user_ids} | all_user {all_user}")
        # user_ids = [str(uid) for uid in user_ids]
        # all_user = [str(one) for one in all_user]
        if len(user_ids) == 0:
            user_ids = all_user
        else:
            user_ids = list(set(user_ids) & set(all_user))
        if len(user_ids) == 0:
            return [], 0
        logger.info(f"get_session_list user_ids {user_ids} | group_ids {group_ids}")
        chat_ids = []
        if start_date or end_date:
            chat_ids = cls.get_filter_chat_ids_by_time(flow_ids, start_date, end_date, category)
        if keyword:
            keyword2 = keyword.encode("unicode_escape").decode().replace("\\u", "%")
            where = select(ChatMessage).where(or_(
                ChatMessage.message.like(f'%{keyword}%'),
                ChatMessage.message.like(f'%{keyword2}%')
            ), ChatMessage.category == 'question')
            if filter_flow_ids:
                where = where.where(ChatMessage.flow_id.in_(filter_flow_ids))
            from sqlalchemy.dialects import mysql
            print("get_session_list Compiled SQL:",
                  where.compile(dialect=mysql.dialect(), compile_kwargs={"literal_binds": True}))
            with session_getter() as session:
                chat_res = session.exec(where).all()
                chat_ids = set(chat_ids) & set([one.chat_id for one in chat_res])
                if len(chat_ids) == 0:
                    chat_ids = [""]
                chat_ids = list(set(chat_ids))
        logger.info(f"get_session_list chat_ids {chat_ids}")
        res = MessageSessionDao.filter_session(chat_ids=chat_ids, sensitive_status=sensitive_status, review_status=review_status, flow_ids=filter_flow_ids,
                                               user_ids=user_ids, start_date=start_date, end_date=end_date,
                                               feedback=feedback, page=page, limit=page_size)
        total = MessageSessionDao.filter_session_count(chat_ids=chat_ids, sensitive_status=sensitive_status, review_status=review_status,
                                                       flow_ids=filter_flow_ids, user_ids=user_ids,
                                                       start_date=start_date, end_date=end_date, feedback=feedback)

        res_users = []
        for one in res:
            res_users.append(one.user_id)
            if start_date or end_date:
                one.like = ChatMessageDao.get_chat_like_num(one.flow_id, one.chat_id, start_date, end_date)
                one.copied = ChatMessageDao.get_chat_copied_num(one.flow_id, one.chat_id, start_date, end_date)
                one.dislike = ChatMessageDao.get_chat_dislike_num(one.flow_id, one.chat_id, start_date, end_date)
        user_list = UserDao.get_user_by_ids(res_users)
        user_map = {user.user_id: user.user_name for user in user_list}

        result = []
        for one in res:
            result.append(AppChatList(chat_id=one.chat_id,
                                      flow_id=one.flow_id,
                                      flow_name=one.flow_name,
                                      flow_type=one.flow_type,
                                      user_id=one.user_id,
                                      user_name=user_map.get(one.user_id, one.user_id),
                                      user_groups=user.get_user_groups(one.user_id),
                                      review_status=one.review_status,
                                      create_time=one.create_time,
                                      update_time=one.update_time,
                                      like_count=one.like,
                                      dislike_count=one.dislike,
                                      copied_count=one.copied))

        return result, total


    @classmethod
    def get_filter_chat_ids_by_time(cls,flow_ids:List[str], start_date: datetime = None, end_date: datetime = None,
                                    category: List[str] = None):
        chat_ids = ChatMessageDao.get_chat_id_by_time(flow_ids, start_date, end_date, category)
        if len(chat_ids) == 0:
            chat_ids = [""]
        return chat_ids

    @classmethod
    def get_session_messages(cls, user: UserPayload, flow_ids: List[str], user_ids: List[int], group_ids: List[int],
                             start_date: datetime, end_date: datetime, feedback: str,
                             sensitive_status: int) -> List[AppChatList]:
        page = 1
        page_size = 50
        res = []
        while True:
            sensitive_status = [SensitiveStatus(sensitive_status)] if sensitive_status else []
            result, total = cls.get_session_list(user=user, flow_ids=flow_ids, user_ids=user_ids, group_ids=group_ids,
                                                 start_date=start_date,
                                                 end_date=end_date, feedback=feedback,
                                                 sensitive_status=sensitive_status, page=page, page_size=page_size)
            if not result:
                break
            page += 1
            res.extend(cls.get_chat_messages(result))
        return res

    @classmethod
    def export_session_messages(cls, user: UserPayload, flow_ids: List[str], user_ids: List[int],
                                group_ids: List[int],
                                start_date: datetime, end_date: datetime,
                                feedback: str, sensitive_status: int) -> str:
        page = 1
        page_size = 30
        excel_data = [
            ['会话ID', '应用名称', '会话创建时间', '用户名称', '消息角色', '消息发送时间', '消息文本内容', '点赞',
             '点踩', '复制']]
        bisheng_pro = settings.get_system_login_method().bisheng_pro
        if bisheng_pro:
            excel_data[0].append('是否命中内容安全审查')

        while True:
            sensitive_status = [SensitiveStatus(sensitive_status)] if sensitive_status else []
            result, total = cls.get_session_list(user=user, flow_ids=flow_ids, user_ids=user_ids, group_ids=group_ids,
                                                 start_date=start_date, end_date=end_date, feedback=feedback,
                                                 sensitive_status=sensitive_status, page=page, page_size=page_size)
            if not result:
                break
            page += 1
            chat_list = cls.get_chat_messages(result)
            for chat in chat_list:
                for message in chat.messages:
                    message_data = [chat.chat_id, chat.flow_name, chat.create_time.strftime('%Y/%m/%d %H:%M:%S'),
                                    chat.user_name,
                                    '用户' if message.category == 'question' else 'AI',
                                    message.create_time.strftime('%Y/%m/%d %H:%M:%S'),
                                    message.message,
                                    '是' if message.liked == LikedType.LIKED.value else '否',
                                    '是' if message.liked == LikedType.DISLIKED.value else '否',
                                    '是' if message.copied else '否']
                    if bisheng_pro:
                        message_data.append(
                            '是' if message.sensitive_status == SensitiveStatus.VIOLATIONS.value else '否')
                    excel_data.append(message_data)

        minio_client = MinioClient()
        tmp_object_name = f'tmp/session/export_{generate_uuid()}.csv'
        with NamedTemporaryFile(mode='w', newline='') as tmp_file:
            csv_writer = csv.writer(tmp_file)
            csv_writer.writerows(excel_data)
            tmp_file.seek(0)
            minio_client.upload_minio(tmp_object_name, tmp_file.name,
                                      'application/text',
                                      minio_client.tmp_bucket)

    @classmethod
    def session_export(cls, all_session: list[AppChatList], export_type: str = "", start_date: datetime=None, end_date: datetime=None):
        excel_data = [["会话ID","应用名称","会话创建时间","用户名称","消息角色","组织架构",
                    "消息发送时间","用户消息文本内容","消息角色", "是否命中安全审查",  # 移除了第一次出现的点赞等列
                    "消息发送时间","用户消息文本内容","消息角色","点赞","点踩","点踩反馈","复制","是否命中安全审查"]]
        for session in all_session:
            flow_id = str(session.flow_id).replace("-", '')
            chat_id = session.chat_id
            where = select(ChatMessage).where(ChatMessage.flow_id == flow_id, ChatMessage.chat_id == chat_id)
            if start_date:
                where = where.where(ChatMessage.update_time >= start_date)
            if end_date:
                where = where.where(ChatMessage.create_time <= end_date)
            with session_getter() as query_session:
                db_message = query_session.exec(where.order_by(ChatMessage.id.asc())).all()
                c_qa = []
                for msg in db_message:
                    if msg.category not in {"question", "stream_msg","output_msg","answer"}:
                        continue
                    if msg.category == "question":
                        if len(c_qa) != 0:
                            while len(c_qa) > len(excel_data[0]):
                                excel_data[0].extend(["消息发送时间","用户消息文本内容","消息角色","点赞","点踩","点踩反馈",
                                                      "复制","是否命中安全审查"])
                            excel_data.append(c_qa)
                        c_qa = [session.chat_id,
                                session.flow_name,
                                session.create_time,
                                session.user_name,
                                "系统",
                                session.user_groups[-1]["name"],
                                msg.create_time,  # 直接添加消息发送时间
                                msg.message,     # 直接添加消息内容
                                "用户" if msg.category == "question" else "AI"]  # 直接添加角色
                         # 添加安全审查状态（第一次）
                        if msg.review_status in {0, 1, 2, 3, 4}:
                            c_qa.append(["", "未审查", "通过", "违规", "审查失败"][msg.review_status])
                        else:
                            c_qa.append("未审查")
                        continue
                    
                    if len(c_qa) == 0:
                        continue
                    # 只从第二次开始添加点赞等字段
                    c_qa.append(msg.create_time) #会话ID
                    c_qa.append(msg.message)
                    c_qa.append("用户" if msg.category == "question" else "AI")
                    c_qa.append("是" if msg.liked == 1 else "否")
                    c_qa.append("是" if msg.liked == 2 else "否")
                    c_qa.append(msg.remark)
                    c_qa.append("是" if msg.copied == 1 else "否")
                    if msg.review_status in {0, 1, 2, 3, 4}:
                        c_qa.append(["", "未审查", "通过", "违规", "审查失败"][msg.review_status])
                    else:
                        c_qa.append("未审查")
                if len(c_qa) != 0:
                    while len(c_qa) > len(excel_data[0]):
                        excel_data[0].extend(
                            ["消息发送时间", "用户消息文本内容", "消息角色", "点赞", "点踩", "点踩反馈", "复制",
                             "是否命中安全审查"])
                    excel_data.append(c_qa)
        wb = Workbook()
        ws = wb.active
        print(excel_data)
        for i in range(len(excel_data)):
            print(excel_data[i])
            for j in range(len(excel_data[i])):
                ws.cell(i + 1, j + 1, excel_data[i][j])

        # 根据 export_type 隐藏列
        if export_type == "audit":
            # 审计模式：隐藏 "点赞"、"点踩"、"点踩反馈"、"复制" 列
            columns_to_hide = ["点赞", "点踩", "点踩反馈", "复制"]
            for col_idx, header in enumerate(excel_data[0], start=1):
                if header in columns_to_hide:
                    ws.column_dimensions[get_column_letter(col_idx)].hidden = True
        elif export_type == "operation":
            # 运营模式：隐藏 "是否命中安全审查" 列
            columns_to_hide = ["是否命中安全审查"]
            for col_idx, header in enumerate(excel_data[0], start=1):
                if header in columns_to_hide:
                    ws.column_dimensions[get_column_letter(col_idx)].hidden = True
                    
        minio_client = MinioClient()
        tmp_object_name = f'tmp/session/export_{generate_uuid()}.docx'
        with NamedTemporaryFile() as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            minio_client.upload_minio(tmp_object_name, tmp_file.name,
                                      'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                                      minio_client.tmp_bucket)

        share_url = minio_client.get_share_link(tmp_object_name, minio_client.tmp_bucket)
        return minio_client.clear_minio_share_host(share_url)

    @classmethod
    def review_session_list(cls, user: UserPayload, flow_ids, user_ids, group_ids, start_date, end_date,
                            feedback, review_status):
        """ 重新审查符合条件的会话 """
        page = 1
        page_size = 10
        while True:
            res, total = cls.get_session_list(user=user, flow_ids=flow_ids, user_ids=user_ids, group_ids=group_ids,
                                              start_date=start_date, end_date=end_date, feedback=feedback,
                                              review_status=review_status, page=page, page_size=page_size)
            if len(res) == 0:
                break
            for one in res:
                cls.review_one_session(one.chat_id, True)
            page += 1

        return cls.get_session_list(user=user, flow_ids=flow_ids, user_ids=user_ids, group_ids=group_ids,
                                    start_date=start_date, end_date=end_date, feedback=feedback, review_status=review_status,
                                    page=1, page_size=10)

    @classmethod
    def review_one_session(cls, chat_id: str, check_all_message: bool = False):
        """ 重新审查一个会话内的消息
        params:
            chat_id: 会话ID
            check_all_message: 是否审查所有消息，默认为False，会过滤掉已审查过的消息
        """
        logger.debug(f"act=review_one_session chat_id={chat_id} all_message={check_all_message}")
        # 审查配置
        review_config = cls.get_session_config()
        if not review_config.flag:
            logger.info(f"act=review flag is close, skip session:{chat_id}")
            return
        # 审查模型
        review_llm = LLMService.get_audit_llm_object()

        all_message = ChatMessageDao.get_msg_by_chat_id(chat_id)
        # 审查通过的消息列表 {id: 1}
        update_pass_messages = []
        # 违规的消息列表 {id: 1, review_reason: []}
        update_violations_messages = []
        old_violations_messages = []
        # 审查失败的消息列表 {id: 1, review_reason: []}
        update_failed_messages = []
        old_failed_messages = []

        message_list = []
        message_content_len = 0
        # 大于多少字符则去请求模型审查
        max_message_len = 500
        session_status = ReviewStatus.PASS.value
        chat_flow_id = None
        chat_flow_type = None
        chat_flow_name = None
        chat_user_id = None
        chat_create_time = None
        for one in all_message:
            if chat_flow_id is None:
                # flow_info = FlowDao.get_flow_by_id(one.flow_id.hex)  # UUID check done
                flow_info = FlowDao.get_flow_by_id(one.flow_id)
                assistant_info = AssistantDao.get_one_assistant(one.flow_id)
                # chat_flow_id = one.flow_id.hex # UUID check done
                chat_flow_id = one.flow_id
                chat_user_id = one.user_id
                chat_create_time = one.create_time
                if flow_info:
                    chat_flow_name = flow_info.name
                    chat_flow_type = flow_info.flow_type
                elif assistant_info:
                    chat_flow_name = assistant_info.name
                    chat_flow_type = FlowType.ASSISTANT.value
                else:
                    logger.debug(f'not found flow info: {one.flow_id}')
                    return
            # 过滤掉工作流的输入事件
            if one.category in ['user_input', 'input']:
                if one.review_status == ReviewStatus.DEFAULT.value:
                    update_pass_messages.append({'id': one.id})
                continue
            if check_all_message or one.review_status == ReviewStatus.DEFAULT.value:
                # 需要审查的消息, 内容为空的消息默认通过审查
                message_content = one.message if one.message else one.intermediate_steps
                if not message_content:
                    update_pass_messages.append({'id': one.id})
                    continue
                message_content_len += message_content.__len__()
                message_list.append({
                    'id': one.id,
                    'message': message_content
                })
                if message_content_len > max_message_len:
                    a, b, c = cls.review_some_message(review_llm, review_config, message_list)
                    update_pass_messages.extend(a)
                    update_violations_messages.extend(b)
                    update_failed_messages.extend(c)
                    message_list = []
                    message_content_len = 0
            else:
                if one.review_status == ReviewStatus.VIOLATIONS.value:
                    old_violations_messages.append(one)
                elif one.review_status == ReviewStatus.FAILED.value:
                    old_failed_messages.append(one)
        if message_list:
            a, b, c = cls.review_some_message(review_llm, review_config, message_list)
            update_pass_messages.extend(a)
            update_violations_messages.extend(b)
            update_failed_messages.extend(c)

        # 更新审查状态
        if update_pass_messages:
            ChatMessageDao.update_review_status([one['id'] for one in update_pass_messages],
                                                session_status, '')
        if update_violations_messages:
            for one in update_violations_messages:
                ChatMessageDao.update_review_status([one['id']], ReviewStatus.VIOLATIONS.value, ','.join(one['reason']))
        if update_failed_messages:
            for one in update_failed_messages:
                ChatMessageDao.update_review_status([one['id']], ReviewStatus.FAILED.value, ','.join(one['reason']))

        # 更新会话的状态
        if update_violations_messages or old_violations_messages:
            session_status = ReviewStatus.VIOLATIONS.value
        elif update_failed_messages or old_failed_messages:
            session_status = ReviewStatus.FAILED.value

        # 更新会话的状态
        message_session = MessageSessionDao.filter_session(chat_ids=[chat_id])
        if message_session.__len__() > 0:
            message_session = message_session[0]
            message_session.review_status = session_status
        else:
            message_session = MessageSession(chat_id=chat_id,
                                             flow_id=chat_flow_id,
                                             flow_type=chat_flow_type,
                                             flow_name=chat_flow_name,
                                             user_id=chat_user_id,
                                             create_time=chat_create_time,
                                             review_status=session_status)
        MessageSessionDao.insert_one(message_session)
        logger.debug(f"act=review_one_session_over chat_id={chat_id} {all_message}")
        return

    @classmethod
    def review_some_message(cls, review_llm: BaseChatModel, review_config: ReviewSessionConfig,
                            message_list: List[dict]) -> (List[int], List[int], List[dict]):
        logger.debug(f'start review_some_message {message_list}')
        try:
            llm_prompt = review_config.prompt
            llm_prompt += f'\n{json.dumps(message_list, ensure_ascii=False, indent=2)}'

            llm_result = review_llm.invoke(llm_prompt)
            logger.debug(f'review message result: {llm_result.content}')

            # 解析模型的输出
            result = extract_code_blocks(llm_result.content)
            if result:
                result = json.loads(result[0])
            else:
                result = json.loads(llm_result.content)

            # 判断有哪些消息审查失败了
            reject_messages = {}
            for one in result.get('messages', []):
                if one.get('message_id') and one.get('violations'):
                    reject_messages[one['message_id']] = {
                        'id': one['message_id'],
                        'reason': one['violations']
                    }
            pass_message = []
            for one in message_list:
                if one['id'] not in reject_messages:
                    pass_message.append({
                        'id': one['id']
                    })
            return pass_message, list(reject_messages.values()), []
        except Exception as e:
            logger.exception(f'review_some_message {message_list} error')
            return [], [], [{'id': one['id'], 'reason': str(e)[-100:]} for one in message_list]

    @classmethod
    def get_session_chart(cls, user: UserPayload, flow_ids: List[str], group_ids: List[str], start_date: datetime,
                          end_date: datetime, order_field: str = None, order_type: str = None, page: int = 0,
                          page_size: int = 0) -> (List, int):

        """ 获取会话聚合数据 """
        # flag, filter_flow_ids = cls.get_filter_flow_ids(user, flow_ids, group_ids)
        # if not flag:
        #     return [], 0

        filter_flow_ids = flow_ids
        for one in flow_ids:
            if one != one.replace('-',''):
                filter_flow_ids.append(one.replace('-',''))
        logger.info(f"get_session_chart: filter_flow_ids={filter_flow_ids} group_ids={group_ids}")
        all_user = UserGroupDao.get_groups_user(group_ids)
        all_user = [str(one) for one in all_user]
        if len(all_user) == 0:
            return [], 0, 0
        logger.info(f"get_session_list all_user {all_user}")
        result = ChatMessageDao.get_chat_info_group(filter_flow_ids, start_date, end_date, order_field,
                                                               order_type, page, page_size,all_user)

        res = result['data']
        total = result['total']
        total_session_num = result['total_session_num']  # 新增的总会话数
        logger.info(f"get_session_list total={total} res={res} total_session_num={total_session_num}")

        if len(res) == 0:
            return res, total, total_session_num
        flow_ids = [one['flow_id'] for one in res]

        # 获取这些应用所属的分组信息
        # app_groups = GroupResourceDao.get_resources_group(None, [one.hex for one in flow_ids])  # UUID check done
        app_groups = GroupResourceDao.get_resources_group(None, flow_ids)  # UUID check done
        app_groups_map = {}
        group_ids = []
        group_info_map = {}
        for one in app_groups:
            group_ids.append(int(one.group_id))
            if one.third_id not in app_groups_map:
                app_groups_map[one.third_id] = []
            app_groups_map[one.third_id].append(int(one.group_id))
        if group_ids:
            group_info = GroupDao.get_group_by_ids(group_ids)
            for one in group_info:
                group_info_map[one.id] = one

        # 获取应用信息
        flow_list = FlowDao.get_flow_by_ids(flow_ids)
        assistant_list = AssistantDao.get_assistants_by_ids(flow_ids)
        flow_map = {flow.id: flow for flow in flow_list}
        assistant_map = {assistant.id: assistant for assistant in assistant_list}

        result = []
        for one in res:
            if flow_map.get(one['flow_id']):
                flow_name = flow_map[one['flow_id']].name
            elif assistant_map.get(one['flow_id']):
                flow_name = assistant_map[one['flow_id']].name
            else:
                continue
            # group_ids = app_groups_map.get(one['flow_id'].hex, [])  # UUID check done
            group_ids = app_groups_map.get(one['flow_id'], [])
            one['name'] = flow_name
            # one['group_info'] = [{
            #     'id': group_id,
            #     'group_name': group_info_map[group_id].group_name if group_info_map.get(group_id) else group_id
            # } for group_id in group_ids]
            one['group_info'] = [ {"id":one['id'],"group_name":one['name']} for one in user.get_user_groups(one['user_id'])]
            result.append(one)

        logger.info(f"get_session_chart result {result} {total} {total_session_num}")
        return result, total, total_session_num

    @classmethod
    def export_session_chart(cls, user: UserPayload, flow_ids: List[str], group_ids: List[str], start_date: datetime,
                             end_date: datetime) -> str:
        """ 导出用户选择的统计数据 """
        result, *_ = cls.get_session_chart(user, flow_ids, group_ids, start_date, end_date, 0, 0)
        excel_data = [['用户组', '应用名称', '会话数', '用户输入消息数', '应用输出消息数', '违规消息数', "好评数1", "好评数2", "差评数"]]
        for one in result:
            excel_data.append([
                ','.join([tmp['group_name'] for tmp in one['group_info']]),
                one['name'],
                one['session_num'],
                one['input_num'],
                one['output_num'],
                one['violations_num'],
                one['likes'],
                one['not_dislikes'],
                one['dislikes'],
            ])

        wb = Workbook()
        ws = wb.active
        for i in range(len(excel_data)):
            for j in range(len(excel_data[i])):
                ws.cell(i + 1, j + 1, excel_data[i][j])

        minio_client = MinioClient()
        tmp_object_name = f'tmp/session/export_{generate_uuid()}.docx'
        with NamedTemporaryFile() as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            minio_client.upload_minio(tmp_object_name, tmp_file.name,
                                      'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                                      minio_client.tmp_bucket)

        share_url = minio_client.get_share_link(tmp_object_name, minio_client.tmp_bucket)
        return minio_client.clear_minio_share_host(share_url)

    @classmethod
    def get_chat_messages(cls, chat_list: List[AppChatList]) -> List[AppChatList]:
        chat_ids = [chat.chat_id for chat in chat_list]

        chat_messages = ChatMessageDao.get_all_message_by_chat_ids(chat_ids)
        chat_messages_map = {}
        for one in chat_messages:
            if one.chat_id not in chat_messages_map:
                chat_messages_map[one.chat_id] = []
            chat_messages_map[one.chat_id].append(one)
        for chat in chat_list:
            chat_messages = chat_messages_map.get(chat.chat_id, [])
            # remove workflow input event, because it's not show in web
            chat.messages = [message for message in chat_messages
                             if message.category != WorkflowEventType.UserInput.value]
        return chat_list

    @classmethod
    def export_audit_session_chart(cls, user: UserPayload, flow_ids: List[str], group_ids: List[str], start_date: datetime,
                             end_date: datetime) -> str:
        """ 导出用户选择的统计数据 """
        result, *_ = cls.get_session_chart(user, flow_ids, group_ids, start_date, end_date, 0, 0)
        excel_data = [
            ['用户组(用户所在部门名称)', '应用名称(用户所使用的应用名称)', '会话数(应用在给定时间区间内产生的会话数)',
             '用户输入消息数(给定时间区间内，某应用所有会话累计的用户输入消息数)', '应用输出消息数(给定时间区间内，某应用所有会话累计的应用输出消息数)',
             '违规消息数(给定时间区间内，某应用所有会话累计的违规消息数)']]
        for one in result:
            excel_data.append([
                ','.join([tmp['group_name'] for tmp in one['group_info']]),
                one['name'],
                one['session_num'],
                one['input_num'],
                one['output_num'],
                one['violations_num']
            ])

        wb = Workbook()
        ws = wb.active
        for i in range(len(excel_data)):
            for j in range(len(excel_data[i])):
                ws.cell(i + 1, j + 1, excel_data[i][j])

        minio_client = MinioClient()
        tmp_object_name = f'tmp/session/export_{generate_uuid()}.docx'
        with NamedTemporaryFile() as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            minio_client.upload_minio(tmp_object_name, tmp_file.name,
                                      'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                                      minio_client.tmp_bucket)

        share_url = minio_client.get_share_link(tmp_object_name, minio_client.tmp_bucket)
        return minio_client.clear_minio_share_host(share_url)


    @classmethod
    def export_operational_session_chart(cls, user: UserPayload, flow_ids: List[str], group_ids: List[str], start_date: datetime,
                             end_date: datetime,like_type) -> str:
        """ 导出用户选择的统计数据 """
        result, *_ = cls.get_session_chart(user, flow_ids, group_ids, start_date, end_date, 0, 0)
        excel_data = [
            ['用户组(用户所在部门名称)', '应用名称(用户所使用的应用名称)', "好评数(用户给予好评的消息数量)","差评数(用户给予差评的消息数量)",
             "应用满意度(好评数/(好评数+差评数) × 100%)","会话数(应用在给定时间区间内产生的会话数)"]]
        for one in result:
            if like_type == 1:
                like = one['likes']
                satisfaction = one['satisfaction']
            else:
                like = one['not_dislikes']
                satisfaction = one['not_nosatisfaction']
            excel_data.append([
                ','.join([tmp['group_name'] for tmp in one['group_info']]),
                one['name'],
                like,
                one['dislikes'],
                f"{satisfaction:.2%}",
                one['session_num']
            ])

        wb = Workbook()
        ws = wb.active
        for i in range(len(excel_data)):
            for j in range(len(excel_data[i])):
                ws.cell(i + 1, j + 1, excel_data[i][j])

        minio_client = MinioClient()
        tmp_object_name = f'tmp/session/export_{generate_uuid()}.docx'
        with NamedTemporaryFile() as tmp_file:
            wb.save(tmp_file.name)
            tmp_file.seek(0)
            minio_client.upload_minio(tmp_object_name, tmp_file.name,
                                      'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                                      minio_client.tmp_bucket)

        share_url = minio_client.get_share_link(tmp_object_name, minio_client.tmp_bucket)
        return minio_client.clear_minio_share_host(share_url)
