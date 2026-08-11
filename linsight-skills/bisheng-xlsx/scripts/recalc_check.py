#!/usr/bin/env python3
"""Recalculate an openpyxl-written .xlsx with LibreOffice, then report errors.

    python skills/bisheng-xlsx/scripts/recalc_check.py output/x.xlsx [--timeout 60]

Why this is mandatory whenever the workbook contains formulas: openpyxl writes
formulas as strings with **no cached value**. Until something evaluates them,
every formula cell reads back as ``None`` to pandas, to
``load_workbook(data_only=True)``, and to most previewers — the file looks empty
where the numbers should be.

LibreOffice evaluates them via a StarBasic macro installed into a throwaway
profile, rewrites the file in place, and this script then re-reads it and names
every cell that came back as an Excel error literal.

BiSheng specifics:

* **Always exits 0.** The executor keeps stdout only on a zero exit; exiting
  non-zero on "errors found" would delete the very report that names them.
* **Private profile per run**, so four concurrent conversions do not collide
  (a fixed profile path is the classic multi-task corruption bug).
* Degrades with an explicit message when LibreOffice has no Calc component —
  some hand-built hosts install only ``libreoffice-writer``, and the resulting
  failure otherwise masquerades as a timeout.
"""

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
import uuid
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print("[FATAL] openpyxl 不可用，无法重算。")
    sys.exit(0)

MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

ERROR_LITERALS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
MAX_LOCATIONS = 50

# '[1]Sheet'!$B$2 — an index into the workbook's external-reference list, i.e. a
# separate file on disk. openpyxl drops the cached value on save; LibreOffice
# then cannot resolve the link and bakes in #NAME?, destroying the data.
EXTERNAL_REF_RE = __import__("re").compile(r"""(?<![\w"\[])'?\[\d+\][^!"\[\]]*'?!""")


def find_soffice() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def has_calc(soffice: str) -> bool:
    """Calc ships as a separate package on RPM hosts; without it xlsx never loads."""
    root = Path(soffice).resolve().parent.parent
    for candidate in (root / "lib" / "libreoffice" / "program", root / "program", Path("/usr/lib/libreoffice/program")):
        if (candidate / "scalc").exists() or (candidate / "scalc.bin").exists():
            return True
    return False


def external_links(path: str) -> list[str]:
    hits = []
    wb = load_workbook(path, data_only=False)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            if not hasattr(ws, "iter_rows"):
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and EXTERNAL_REF_RE.search(cell.value):
                        hits.append(f"{name}!{cell.coordinate}")
                        if len(hits) >= 10:
                            return hits
    finally:
        wb.close()
    return hits


def install_macro(profile: Path, soffice: str, timeout: int) -> str | None:
    """Boot LibreOffice once so it materialises the profile, then drop the macro in."""
    url = profile.as_uri()
    try:
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init", f"-env:UserInstallation={url}"],
            capture_output=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return None
    macro_dir = profile / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        return None
    (macro_dir / "Module1.xba").write_text(MACRO)
    return url


def scan(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    details: dict[str, list[str]] = {e: [] for e in ERROR_LITERALS}
    total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        if not hasattr(ws, "iter_rows"):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for err in ERROR_LITERALS:
                        if err in cell.value:
                            details[err].append(f"{name}!{cell.coordinate}")
                            total += 1
                            break
    wb.close()

    wb_f = load_workbook(path, data_only=False)
    formulas = 0
    uncached = 0
    wb_v = load_workbook(path, data_only=True)
    for name in wb_f.sheetnames:
        ws = wb_f[name]
        if not hasattr(ws, "iter_rows"):
            continue
        vs = wb_v[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    if vs[cell.coordinate].value is None:
                        uncached += 1
    wb_f.close()
    wb_v.close()

    summary = {}
    for err, locations in details.items():
        if locations:
            entry = {"count": len(locations), "locations": locations[:MAX_LOCATIONS]}
            if len(locations) > MAX_LOCATIONS:
                entry["locations_truncated"] = len(locations) - MAX_LOCATIONS
            summary[err] = entry

    return {
        "status": "success" if total == 0 else "errors_found",
        "total_formulas": formulas,
        "total_errors": total,
        "uncached_formulas": uncached,
        "error_summary": summary,
    }


def recalc(path: str, timeout: int, force: bool) -> dict:
    soffice = find_soffice()
    if not soffice:
        return {"error": "环境里没有 soffice（LibreOffice）——无法重算公式。"}
    if not has_calc(soffice):
        return {
            "error": (
                "LibreOffice 装了，但没有 Calc 组件（常见于只装了 libreoffice-writer 的手工部署）。"
                "重算会失败，且报错通常伪装成超时。请让运维补装 libreoffice-calc。"
            )
        }

    links = external_links(path)
    if links and not force:
        return {
            "error": (
                f"这个工作簿引用了外部文件（如 {', '.join(links[:3])}）。openpyxl 保存时已丢掉缓存值，"
                f"重算会让 LibreOffice 解析不到外链、写入 #NAME? 并删掉链接，数据就没了。"
                f"先把这些单元格的值从原件里取出来固化，再重算；确认可以承受损失时加 --force。"
            )
        }

    abs_path = os.path.abspath(path)
    before = os.stat(abs_path)
    before_stamp = (before.st_mtime_ns, before.st_size)

    with tempfile.TemporaryDirectory(prefix=f"lo_recalc_{uuid.uuid4().hex[:8]}_") as tmp:
        profile = Path(tmp) / "profile"
        started = time.monotonic()
        url = install_macro(profile, soffice, timeout)
        if url is None:
            return {"error": f"LibreOffice 没能在 {timeout}s 内建出可用的 profile，公式未重算。"}

        remaining = max(10, int(timeout - (time.monotonic() - started)))
        cmd = [
            soffice,
            "--headless",
            "--norestore",
            "--invisible",
            "-env:SingleAppInstance=false",
            f"-env:UserInstallation={url}",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ]
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=remaining + 15)
        except subprocess.TimeoutExpired:
            return {"error": f"LibreOffice 重算超时（{remaining}s）。加大 --timeout 重试。"}

        if proc.returncode != 0:
            detail = (proc.stderr or "").strip() or f"soffice 退出码 {proc.returncode}"
            hint = ""
            if proc.returncode == 137:
                hint = "（退出码 137 = 被 OOM killer 杀掉，多半是内存配额太小，不是 LibreOffice 坏了）"
            return {"error": f"LibreOffice 重算失败：{detail}{hint}"}

        after = os.stat(abs_path)
        if (after.st_mtime_ns, after.st_size) == before_stamp:
            return {
                "error": (
                    "LibreOffice 正常退出但没有重写文件，等于什么都没算。"
                    "通常是文件被别的进程占用，或宏没被执行。重试一次；仍失败就把文件另存一个新名字再试。"
                )
            }

    return scan(path)


def report(result: dict) -> None:
    print("=" * 72)
    print("公式重算")
    print("=" * 72)

    if "error" in result:
        print(f"[未重算] {result['error']}")
        print()
        print("结论: 不通过 —— 公式没有缓存值，交付出去用户和 pandas 读到的都是空。")
        print("降级方案：如果这台机器确实没有 Calc，就不要用公式，直接在 Python 里算好写入数值，")
        print("并在表里用一列文字说明计算口径（否则用户改了输入也不会重算）。")
        print()
        print("JSON: " + json.dumps(result, ensure_ascii=False))
        return

    print(f"公式总数 : {result['total_formulas']}")
    print(f"错误单元格: {result['total_errors']}")
    if result.get("uncached_formulas"):
        print(f"仍无缓存值: {result['uncached_formulas']}（异常，重算应当把它们都填上）")

    for err, entry in result.get("error_summary", {}).items():
        locations = ", ".join(entry["locations"])
        extra = f"（另有 {entry['locations_truncated']} 处未列出）" if "locations_truncated" in entry else ""
        print(f"  {err} × {entry['count']}: {locations}{extra}")

    print()
    if result["total_errors"]:
        print("结论: 不通过 —— 上面每个错误单元格都要改完再重算。")
    else:
        print("结论: 通过（公式都能算出结果）。")
        print("注意：能算 ≠ 算对。口径错、区间差一行照样是干净的错数字 ——")
        print("先抽查 2–3 个关键公式的结果符不符合预期，再铺开整张表。")
    print()
    print("JSON: " + json.dumps(result, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser(description="用 LibreOffice 重算 xlsx 公式并检查错误")
    parser.add_argument("path")
    parser.add_argument("--timeout", type=int, default=60)
    parser.add_argument("--force", action="store_true", help="即使存在外部链接也强行重算（接受数据丢失）")
    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f"[FATAL] 文件不存在: {args.path}")
        print("提示：用相对路径 `output/x.xlsx`，不要写 `/output/x.xlsx`。")
        return

    try:
        report(recalc(args.path, args.timeout, args.force))
    except Exception:
        print("[FATAL] 重算过程本身出错：")
        traceback.print_exc(file=sys.stdout)


if __name__ == "__main__":
    main()
    sys.exit(0)
