"""An .xlsx openpyxl refuses over its styles must still be parsed, or fail loudly.

Field case (knowledge space 142, file 860): a 28 MB workbook with 18 populated
sheets was listed as "解析成功" while its knowledge space answered 「没有找到相关
内容」 to every question. Nothing had been indexed — ES held 0 documents and the
Milvus collection was never even created.

The cause was three bare ``<fill/>`` elements in ``xl/styles.xml``. openpyxl
rejects those (``TypeError: expected <class 'openpyxl.styles.fills.Fill'>``)
before reading a single sheet, and ``excel_file_to_markdown`` answered that with
``logger.debug(...); return`` — at a level below the log file's sink, so the
reason never even landed on disk. The loader then returned zero documents and the
caller marked the file parsed.

These tests pin both halves: the repair makes such a workbook readable, and a
workbook that stays unreadable raises instead of returning empty.
"""

from __future__ import annotations

import re
import zipfile

import openpyxl
import pytest

from bisheng.common.errcode.knowledge import KnowledgeFileDamagedError
from bisheng.knowledge.rag.pipeline.loader.utils.md_from_excel import convert_file_to_markdown
from bisheng.knowledge.rag.pipeline.loader.utils.xlsx_repair import STYLES_PART, repair_xlsx_styles


def _write_workbook(path, rows=(("名称", "数量"), ("大豆", 12))):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    return path


def _break_styles(path, broken_path, marker="<fill/>"):
    """Inject the field-observed empty ``<fill/>`` into a valid workbook's styles."""
    with zipfile.ZipFile(path) as src:
        styles = src.read(STYLES_PART).decode()
        # Append the empty fill to the existing <fills> list and bump its count.
        broken = re.sub(
            r"(<fills count=\")(\d+)(\")", lambda m: f"{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}", styles
        )
        broken = broken.replace("</fills>", f"{marker}</fills>")
        assert broken != styles
        with zipfile.ZipFile(broken_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                payload = broken.encode() if item.filename == STYLES_PART else src.read(item.filename)
                dst.writestr(item, payload)
    return broken_path


# ---------------------------------------------------------------------------
# repair_xlsx_styles
# ---------------------------------------------------------------------------
def test_openpyxl_really_rejects_an_empty_fill(tmp_path):
    """Guards the premise. If a future openpyxl accepts these, the repair below
    becomes dead code and should be removed rather than left to rot."""
    broken = _break_styles(_write_workbook(tmp_path / "ok.xlsx"), tmp_path / "broken.xlsx")
    with pytest.raises(Exception):
        openpyxl.load_workbook(broken, data_only=True)


def test_repair_makes_the_workbook_loadable_with_its_data_intact(tmp_path):
    broken = _break_styles(_write_workbook(tmp_path / "ok.xlsx"), tmp_path / "broken.xlsx")
    repaired = repair_xlsx_styles(str(broken))
    assert repaired is not None
    wb = openpyxl.load_workbook(repaired, data_only=True)
    # The point of copying every other part verbatim: the cells must survive.
    assert [tuple(r) for r in wb.active.iter_rows(values_only=True)] == [("名称", "数量"), ("大豆", 12)]


def test_the_empty_close_tag_form_is_repaired_too(tmp_path):
    broken = _break_styles(_write_workbook(tmp_path / "ok.xlsx"), tmp_path / "broken.xlsx", marker="<fill></fill>")
    assert repair_xlsx_styles(str(broken)) is not None


def test_a_healthy_workbook_is_left_alone(tmp_path):
    """No repair means no copy — a workbook that opens today must take exactly
    the path it takes today."""
    assert repair_xlsx_styles(str(_write_workbook(tmp_path / "ok.xlsx"))) is None


def test_a_non_zip_file_is_not_this_function_s_problem(tmp_path):
    path = tmp_path / "notreally.xlsx"
    path.write_bytes(b"this is not a zip")
    assert repair_xlsx_styles(str(path)) is None


# ---------------------------------------------------------------------------
# convert_file_to_markdown — the entry point ExcelLoader actually calls
# ---------------------------------------------------------------------------
def test_a_style_broken_workbook_is_converted_via_the_repair(tmp_path):
    broken = _break_styles(_write_workbook(tmp_path / "ok.xlsx"), tmp_path / "broken.xlsx")
    out = tmp_path / "md"
    convert_file_to_markdown(str(broken), [1, 1], 10, str(out), append_header=True)
    produced = sorted(out.glob("*.md"))
    assert produced, "the repair path must yield markdown, not silence"
    assert "大豆" in produced[0].read_text(encoding="utf-8")


def test_an_unreadable_workbook_raises_instead_of_returning_empty(tmp_path):
    """The whole point: a file we cannot open must not look like a file with no
    content. Returning quietly is what let the caller mark it SUCCESS."""
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"PK\x03\x04 not really a workbook")
    with pytest.raises(KnowledgeFileDamagedError):
        convert_file_to_markdown(str(path), [1, 1], 10, str(tmp_path / "md"))
