import io
import json
import uuid
from typing import List, Dict, Literal

import openpyxl
from fastapi import UploadFile
from langchain_core.documents import Document
from langchain_core.embeddings import Embeddings
from langchain_core.language_models import BaseChatModel
from langchain_text_splitters import RecursiveCharacterTextSplitter
from loguru import logger

from bisheng.api.services.knowledge_imp import decide_vectorstores, extract_code_blocks
from bisheng.api.v1.schema.inspiration_schema import SOPManagementSchema, SOPManagementUpdateSchema
from bisheng.api.v1.schema.linsight_schema import SopRecordRead
from bisheng.api.v1.schemas import UnifiedResponseModel, resp_200
from bisheng.common.dependencies.user_deps import UserPayload
from bisheng.common.errcode import BaseErrorCode
from bisheng.common.errcode.http_error import NotFoundError
from bisheng.common.errcode.linsight import (
    LinsightAddSopError, LinsightUpdateSopError, LinsightDeleteSopError,
    LinsightVectorModelError, LinsightDocSearchError, LinsightDocNotFoundError, SopContentOverLimitError
)
from bisheng.common.errcode.server import (
    NoEmbeddingModelError, EmbeddingModelNotExistError, EmbeddingModelTypeError, UploadFileEmptyError
)
from bisheng.common.services.config_service import settings
from bisheng.core.prompts.manager import get_prompt_manager
from bisheng.database.models.linsight_sop import LinsightSOP, LinsightSOPDao, LinsightSOPRecord
from bisheng.interface.embeddings.custom import FakeEmbedding
from bisheng.llm.domain.const import LLMModelType
from bisheng.llm.domain.models import LLMDao
from bisheng.llm.domain.services import LLMService
from bisheng.user.domain.models.user import UserDao
from bisheng.utils import util
from bisheng_langchain.rag.init_retrievers import KeywordRetriever, BaselineVectorRetriever
from bisheng_langchain.retrievers import EnsembleRetriever
from bisheng_langchain.vectorstores import ElasticKeywordsSearch, Milvus


class SOPManageService:
    __doc__ = "InspirationSOPManage Services"

    collection_name = "col_linsight_sop"

    @staticmethod
    async def generate_sop_summary(invoke_user_id: int, sop_content: str, llm: BaseChatModel = None) -> Dict[str, str]:
        """BuatSOPAbstract"""
        default_summary = {"sop_title": "SOP Title", "sop_description": "SOP Description"}

        try:
            if llm is None:
                workbench_conf = await LLMService.get_workbench_llm()
                linsight_conf = settings.get_linsight_conf()
                llm = await LLMService.get_bisheng_linsight_llm(invoke_user_id=invoke_user_id,
                                                                model_id=workbench_conf.task_model.id,
                                                                temperature=linsight_conf.default_temperature)
            prompt_service = await get_prompt_manager()
            prompt_obj = prompt_service.render_prompt(
                namespace="sop",
                prompt_name="gen_sop_summary",
                sop_detail=sop_content
            )

            prompt = [
                ("system", prompt_obj.prompt.system),
                ("user", prompt_obj.prompt.user)
            ]

            response = await llm.ainvoke(prompt)
            if not response.content:
                return default_summary
            code_ret = extract_code_blocks(response.content)
            if code_ret:
                return json.loads(code_ret[0])
            return json.loads(response.content)

        except Exception as e:
            logger.exception(f"BuatSOPSummary failed: {e}")
            return default_summary

    @staticmethod
    async def add_sop_record(sop_record: LinsightSOPRecord) -> LinsightSOPRecord:
        """
        TambahSOPRecord
        """
        if not sop_record.description:
            sop_summary = await SOPManageService.generate_sop_summary(sop_record.user_id, sop_record.content, None)
            sop_record.description = sop_summary["sop_description"]

        return await LinsightSOPDao.create_sop_record(sop_record)

    @staticmethod
    async def get_sop_record(keyword: str = None, sort: str = None, page: int = 1, page_size: int = 10) -> \
            (List[SopRecordRead], int):
        """
        Query by keywordSOPRecord
        """
        user_ids = []
        if keyword:
            # If there are keywords, get the user firstIDVertical
            user_ids = await UserDao.afilter_users(user_ids=[], keyword=keyword)
            user_ids = [one.user_id for one in user_ids]

        res = await LinsightSOPDao.filter_sop_record(keyword, user_ids, page, page_size, sort)
        count = await LinsightSOPDao.count_sop_record(keyword, user_ids)
        if not res:
            return [], 0

        all_users = await UserDao.afilter_users(user_ids=[one.user_id for one in res])
        all_users = {
            one.user_id: one.user_name for one in all_users
        }

        result = []
        for one in res:
            new_one = SopRecordRead.model_validate(one)
            new_one.user_name = all_users.get(one.user_id, str(one.user_id))
            result.append(new_one)
        return result, count

    @staticmethod
    async def update_sop_record_score(session_version_id: str, score: int) -> None:
        await LinsightSOPDao.update_sop_record_score(session_version_id, score)

    @staticmethod
    async def update_sop_record_feedback(session_version_id: str, feedback: str) -> None:
        await LinsightSOPDao.update_sop_record_feedback(session_version_id, feedback)

    @classmethod
    async def sync_sop_record(cls, record_ids: list[int], override: bool = False, save_new: bool = False) \
            -> list[str] | None:
        """
        SynchronousSOPRecord
        :param record_ids: SOPRecordIDVertical
        :param override: Do you want to overwrite the existingSOP
        :param save_new: Do you want to save the newSOP
        :return: If there is a duplicateSOPrecords, return a list of duplicate record names, otherwise returnNone
        """
        sop_records = await LinsightSOPDao.get_sop_record_by_ids(record_ids)

        return await cls._sync_sop_record(sop_records, override, save_new)

    @staticmethod
    async def _sync_sop_record(sop_records: list[LinsightSOPRecord], override: bool = False, save_new: bool = False) \
            -> list[str] | None:

        """
        If there is a duplicateSOPrecords, returning a list of duplicate record names
        """
        records_name_dict = {}
        repeat_names = set()
        name_set = set()
        sop_list = []
        oversize_records = []
        new_records = []
        for one in sop_records:
            if len(one.content) > 50000:
                oversize_records.append(one.name)
                continue
            new_records.append(one)
            if one.name not in name_set:
                records_name_dict[one.name] = one
                name_set.add(one.name)
        sop_records = new_records
        if not sop_records and oversize_records:
            raise SopContentOverLimitError(data={"sop_name": "、".join(oversize_records)})
        if name_set:
            sop_list = await LinsightSOPDao.get_sops_by_names(list(name_set))
            for one in sop_list:
                repeat_names.add(one.name)

        if override:
            # Update existing firstsopGallery
            override_name_dict = {}
            for one in sop_list:
                if one_record := records_name_dict.get(one.name):
                    await SOPManageService.update_sop(SOPManagementUpdateSchema(
                        id=one.id,
                        name=one.name,
                        description=one_record.description,
                        content=one_record.content,
                        rating=one_record.rating,
                        linsight_version_id=one_record.linsight_version_id,
                        showcase=False,
                        user_id=one_record.user_id,
                    ))
                    override_name_dict[one.name] = True
            # Add the restsopRecord
            for one in records_name_dict.values():
                if one.name in override_name_dict:
                    continue
                await SOPManageService.add_sop(SOPManagementSchema(
                    name=one.name,
                    description=one.description,
                    content=one.content,
                    rating=one.rating,
                    linsight_version_id=one.linsight_version_id,
                ), one.user_id)
        elif save_new:
            for one in sop_records:
                new_name = one.name
                if new_name in repeat_names:
                    # Add suffix if there are duplicate records, Limit Length500characters
                    new_name = f"{one.name}Dungeon"
                await SOPManageService.add_sop(SOPManagementSchema(
                    name=new_name,
                    description=one.description,
                    content=one.content,
                    rating=one.rating,
                    linsight_version_id=one.linsight_version_id,
                ), one.user_id)
        else:
            # Explain that there is a duplicate record, which needs to be confirmed by the user
            if sop_list:
                return list(repeat_names)
            # Insert a record into the database
            for one in sop_records:
                await SOPManageService.add_sop(SOPManagementSchema(
                    name=one.name,
                    description=one.description,
                    content=one.content,
                    rating=one.rating,
                    linsight_version_id=one.linsight_version_id,
                ), one.user_id)
        if oversize_records:
            raise SopContentOverLimitError(data={"sop_name": "、".join(oversize_records)})
        return None

    @classmethod
    async def parse_sop_file(cls, file: UploadFile) -> (list, list):
        """
        analyzingSOPDoc.
        :param file: FilePath
        """
        if not file.size:
            raise UploadFileEmptyError()
        error_rows = []
        success_rows = []
        wb = None

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
            sheet = wb.active
            max_rows = sheet.max_row
            for i in range(2, max_rows + 1):
                name = sheet.cell(row=i, column=1).value
                description = sheet.cell(row=i, column=2).value
                content = sheet.cell(row=i, column=3).value
                error_msg = []
                if not name:
                    error_msg.append("name_empty")
                if not content:
                    error_msg.append("description_empty")
                if len(str(name)) >= 500:
                    error_msg.append("name_over_size")
                if len(str(content)) >= 50000:
                    error_msg.append("content_over_size")
                if description and len(str(description)) >= 1000:
                    error_msg.append("description_over_size")
                if error_msg:
                    error_rows.append({
                        "index": i,
                        "error_msg": error_msg
                    })
                else:
                    success_rows.append({
                        "name": str(name),
                        "description": str(description) if description is not None else "",
                        "content": str(content),
                    })
        finally:
            if wb:
                wb.close()
        return success_rows, error_rows

    @classmethod
    async def upload_sop_file(cls, login_user: UserPayload, file: UploadFile, ignore_error: bool, override: bool,
                              save_new: bool) \
            -> (List[Dict], List[Dict], List[str]):
        """
        Upload itSOPDoc.
        :param login_user: Logged in user information
        :param file: FilePath
        :param ignore_error: Do you want to ignore the error
        :param override: Do you want to overwrite the existingSOP
        :param save_new: Do you want to save the newSOP
        :return: Upload results, success_rows, error_rows, repeat_names
        """
        success_rows, error_rows = await cls.parse_sop_file(file)
        if (error_rows or len(success_rows) == 0) and not ignore_error:
            return success_rows, error_rows
        if not success_rows:
            return [], [], []
        records = [LinsightSOPRecord(**one, user_id=login_user.user_id) for one in success_rows]
        repeat_name_list = await cls._sync_sop_record(records, override=override, save_new=save_new)
        return [], [], repeat_name_list

    @classmethod
    async def get_sop_list(cls, keywords: str = None, sort: Literal["asc", "desc"] = "desc", showcase: bool = False,
                           page: int = 1, page_size: int = 10) -> dict:
        """
        DapatkanSOPVertical
        :param keywords: Keyword
        :param sort: Sort By
        :param page: Page
        :param page_size: Items per page
        :param showcase: Whether to show only the selected cases ofSOP
        :return: SOPLists and totals
        """
        sop_pages = await LinsightSOPDao.get_sop_page(keywords=keywords, showcase=showcase, page=page,
                                                      page_size=page_size,
                                                      sort=sort)
        user_ids = list(set([one["user_id"] for one in sop_pages["items"]]))
        user_map = UserDao.aget_user_by_ids(user_ids=user_ids)
        user_map = {one.user_id: one.user_name for one in await user_map}
        for one in sop_pages["items"]:
            one["user_name"] = user_map.get(one["user_id"], str(one["user_id"]))
        return sop_pages

    @staticmethod
    async def add_sop(sop_obj: SOPManagementSchema, user_id: int) -> UnifiedResponseModel | None:
        """
        Add NewSOP
        :param user_id:
        :param sop_obj:
        :return: Added bySOPObjects
        """

        # Get the current global configuration ofembeddingModels
        workbench_conf = await LLMService.get_workbench_llm()
        try:
            emb_model_id = workbench_conf.embedding_model.id
            if not emb_model_id:
                return NoEmbeddingModelError.return_resp()
        except AttributeError:
            return NoEmbeddingModelError.return_resp()

        # CorrectionembeddingModels
        embed_info = LLMDao.get_model_by_id(int(emb_model_id))
        if not embed_info:
            return EmbeddingModelNotExistError.return_resp()
        if embed_info.model_type != LLMModelType.EMBEDDING.value:
            return EmbeddingModelTypeError.return_resp()

        vector_store_id = uuid.uuid4().hex

        embeddings = await LLMService.get_bisheng_linsight_embedding(model_id=embed_info.id,
                                                                     invoke_user_id=user_id)
        try:
            vector_client: Milvus = decide_vectorstores(
                SOPManageService.collection_name, "Milvus", embeddings
            )

            es_client: ElasticKeywordsSearch = decide_vectorstores(
                SOPManageService.collection_name, "ElasticKeywordsSearch", FakeEmbedding()
            )
            metadatas = [{"vector_store_id": vector_store_id}]
            vector_client.add_texts([sop_obj.content[0:10000]], metadatas=metadatas)
            es_client.add_texts([sop_obj.content], ids=[vector_store_id], metadatas=metadatas)
        except Exception as e:
            return LinsightAddSopError.return_resp(data=str(e))

        sop_dict = sop_obj.model_dump(exclude_unset=True)
        sop_dict["vector_store_id"] = vector_store_id  # Set Vector StorageID
        # Here you can add a database operation that willsop_objSave to Database
        sop_model = LinsightSOP(**sop_dict)
        sop_model.user_id = user_id
        sop_model = await LinsightSOPDao.create_sop(sop_model)
        return resp_200(data=sop_model)

    @staticmethod
    async def update_sop(sop_obj: SOPManagementUpdateSchema,
                         update_version_id: bool = True) -> UnifiedResponseModel | None:
        """
        UpdateSOP
        :param sop_obj:
        :param update_version_id: Do you want to update the versionID
        :return: Post UpdateSOPObjects
        """
        # CorrectionSOPpresence or does it
        existing_sop = await LinsightSOPDao.get_sops_by_ids([sop_obj.id])
        if not existing_sop:
            return NotFoundError.return_resp()
        if not update_version_id:
            sop_obj.linsight_version_id = existing_sop[0].linsight_version_id

        if sop_obj.content != existing_sop[0].content:

            # Get the current global configuration ofembeddingModels
            workbench_conf = await LLMService.get_workbench_llm()
            try:
                emb_model_id = workbench_conf.embedding_model.id
                if not emb_model_id:
                    return NoEmbeddingModelError.return_resp()
            except AttributeError:
                return NoEmbeddingModelError.return_resp()

            vector_store_id = existing_sop[0].vector_store_id
            embeddings = await LLMService.get_bisheng_linsight_embedding(invoke_user_id=sop_obj.user_id,
                                                                         model_id=int(emb_model_id))

            # Update Vector Store
            try:
                vector_client: Milvus = decide_vectorstores(
                    SOPManageService.collection_name, "Milvus", embeddings
                )
                es_client: ElasticKeywordsSearch = decide_vectorstores(
                    SOPManageService.collection_name, "ElasticKeywordsSearch", FakeEmbedding()
                )

                vector_client.delete(expr=f"vector_store_id == '{vector_store_id}'")
                es_client.delete([vector_store_id])
                metadatas = [{"vector_store_id": vector_store_id}]
                vector_client.add_texts([sop_obj.content[0:10000]], metadatas=metadatas)
                es_client.add_texts([sop_obj.content], ids=[vector_store_id], metadatas=metadatas)

            except Exception as e:
                return LinsightUpdateSopError.return_resp(data=str(e))

        # Update databaseSOP
        sop_model = await LinsightSOPDao.update_sop(sop_obj)

        return resp_200(data=sop_model)

    @staticmethod
    async def remove_sop(sop_ids: list[int], login_user: UserPayload) -> UnifiedResponseModel | None:
        """
        DeleteSOP
        :param login_user:
        :param sop_ids: SOPUniqueness quantificationIDVertical
        :return: Remove result
        """

        # CorrectionSOPpresence or does it
        existing_sops = await LinsightSOPDao.get_sops_by_ids(sop_ids)
        if not existing_sops:
            return NotFoundError.return_resp()

        # Delete data in vector store
        try:
            vector_store_ids = [sop.vector_store_id for sop in existing_sops]
            vector_client: Milvus = decide_vectorstores(
                SOPManageService.collection_name, "Milvus", FakeEmbedding()
            )
            es_client: ElasticKeywordsSearch = decide_vectorstores(
                SOPManageService.collection_name, "ElasticKeywordsSearch", FakeEmbedding()
            )

            vector_client.delete(expr=f"vector_store_id in {vector_store_ids}")
            es_client.delete(vector_store_ids)

        except Exception as e:
            return LinsightDeleteSopError.return_resp(data=str(e))

        # Delete from databaseSOP
        await LinsightSOPDao.remove_sop(sop_ids=sop_ids)

        return resp_200(data=True)

    @classmethod
    async def get_sop_by_id(cls, sop_id: int) -> LinsightSOP | None:
        """
        accordingIDDapatkanSOP
        :param sop_id: SOPUniqueness quantificationID
        :return: SOPObjects
        """
        sop_models = await LinsightSOPDao.get_sops_by_ids([sop_id])
        if not sop_models:
            return None
        return sop_models[0]

    # sop Library Retrieval
    @classmethod
    async def search_sop(cls, invoke_user_id: int, query: str, k: int = 3) -> (List[Document], BaseErrorCode | None):
        """
        CariSOP
        :param k:
        :param query: Keywords Search
        :return: search results
        """
        # Get the current global configuration ofembeddingModels
        try:
            vector_search = True
            es_search = True
            error_msg = None
            workbench_conf = await LLMService.get_workbench_llm()
            if workbench_conf.embedding_model is None or not workbench_conf.embedding_model.id:
                vector_search = False
                error_msg = LinsightVectorModelError
            else:
                try:
                    emb_model_id = workbench_conf.embedding_model.id
                    embeddings = await LLMService.get_bisheng_linsight_embedding(invoke_user_id=invoke_user_id,
                                                                                 model_id=int(emb_model_id))
                    await embeddings.aembed_query("test")
                except Exception as e:
                    logger.error(f"Vector retrieval model initialization failed: {str(e)}")
                    vector_search = False
                    error_msg = LinsightVectorModelError

            # Create Text Splitter
            text_splitter = RecursiveCharacterTextSplitter()
            retrievers = []
            if vector_search and es_search:
                emb_model_id = workbench_conf.embedding_model.id
                embeddings = await LLMService.get_bisheng_linsight_embedding(invoke_user_id=invoke_user_id,
                                                                             model_id=int(emb_model_id))

                vector_client: Milvus = decide_vectorstores(
                    SOPManageService.collection_name, "Milvus", embeddings
                )

                es_client: ElasticKeywordsSearch = decide_vectorstores(
                    SOPManageService.collection_name, "ElasticKeywordsSearch", FakeEmbedding()
                )

                keyword_retriever = KeywordRetriever(keyword_store=es_client, search_kwargs={"k": 100},
                                                     text_splitter=text_splitter)
                baseline_vector_retriever = BaselineVectorRetriever(vector_store=vector_client,
                                                                    search_kwargs={"k": 100},
                                                                    text_splitter=text_splitter)

                retrievers = [keyword_retriever, baseline_vector_retriever]

            elif es_search and not vector_search:
                # Search with keywords only
                es_client: ElasticKeywordsSearch = decide_vectorstores(
                    SOPManageService.collection_name, "ElasticKeywordsSearch", FakeEmbedding()
                )
                keyword_retriever = KeywordRetriever(keyword_store=es_client, search_kwargs={"k": 100},
                                                     text_splitter=text_splitter)
                retrievers = [keyword_retriever]

            elif vector_search and not es_search:
                # Use vector retrieval only
                emb_model_id = workbench_conf.embedding_model.id
                embeddings = await LLMService.get_bisheng_linsight_embedding(invoke_user_id=invoke_user_id,
                                                                             model_id=int(emb_model_id))

                vector_client: Milvus = decide_vectorstores(
                    SOPManageService.collection_name, "Milvus", embeddings
                )

                baseline_vector_retriever = BaselineVectorRetriever(vector_store=vector_client,
                                                                    search_kwargs={"k": 100},
                                                                    text_splitter=text_splitter)
                retrievers = [baseline_vector_retriever]
            else:
                error_msg = LinsightDocSearchError
                return [], error_msg

            retriever = EnsembleRetriever(retrievers=retrievers, weights=[0.5, 0.5] if len(retrievers) > 1 else [1.0])

            # Perform Retrieval
            results = await retriever.ainvoke(input=query)

            if not results:
                return [], error_msg

            vector_store_ids = [doc.metadata.get("vector_store_id") for doc in results if
                                doc.metadata.get("vector_store_id")]

            # accordingvector_store_idsQuerying repositoriessop
            sop_models = await LinsightSOPDao.get_sop_by_vector_store_ids(vector_store_ids)
            sop_model_vector_store_ids = [sop.vector_store_id for sop in sop_models]

            # Filter results to ensure that only those that exist in the database are returnedSOP
            results = [doc for doc in results if doc.metadata.get("vector_store_id") in sop_model_vector_store_ids]

            # Before filtering and fetchingk result(s) found
            results = results[:k]

            return results, error_msg
        except Exception as e:
            logger.error(f"Failed to search the instruction manual: {str(e)}")
            return [], LinsightDocNotFoundError

    # RebuildSOP VectorStore
    @classmethod
    async def rebuild_sop_vector_store_task(cls, embeddings: Embeddings):
        """
        RebuildSOPVector Storage
        :return: Reconstruction Results
        """
        try:
            # Fetch allSOP
            all_sops = await LinsightSOPDao.get_all_sops()
            if not all_sops:
                logger.info("NoSOPData needs to be reconstructed for vector storage")
                return None

            # Wrapper synchronization function is asynchronous function
            def sync_func(sops, emb):
                """
                Synchronization function for rebuildingSOPVector Storage
                :param emb:
                :param sops:
                :return:
                """

                vector_client: Milvus = decide_vectorstores(
                    SOPManageService.collection_name, "Milvus", emb
                )
                # Delete existing vector storecollection
                if vector_client.col is not None:
                    logger.info("Delete existingSOPVector Storagecollection")
                    vector_client.col.drop()
                    vector_client.col = None
                    vector_client.fields = []

                metadatas = [{"vector_store_id": sop.vector_store_id} for sop in sops]
                contents = [sop.content for sop in sops]

                batch_size = 16
                for i in range(0, len(contents), batch_size):
                    batch_contents = contents[i:i + batch_size]
                    batch_metadatas = metadatas[i:i + batch_size]

                    # Add NewSOPData to vector storage
                    vector_client.add_texts(batch_contents, metadatas=batch_metadatas)

                logger.info("SOPVector store rebuild completed: {}".format(len(sops)))

            # Userun_asyncRun Synchronization Function
            await util.sync_func_to_async(sync_func)(all_sops, embeddings)
            return None


        except Exception as e:
            logger.exception(f"RebuildSOPVector store failed: {str(e)}")
            return None

# if __name__ == '__main__':
#     # test code
#     results, error_msg = asyncio.run(SOPManageService.search_sop(query="Guidelines for the preparation of bidding documents", k=3))
#
#     print(results)
#     print(error_msg)
