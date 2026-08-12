"""openpyxl helpers for building .xlsx inside the BiSheng code interpreter.

Use from a build script:

    import sys
    sys.path.insert(0, "skills/bisheng-xlsx/scripts")
    from xlsx_helpers import autofit_columns, write_table, mark_input, FMT

Everything here is plain openpyxl — no dependency the backend image does not
already ship. The value is in the traps it closes: CJK column width, percentages
stored as whole numbers, sheet names needing quotes, and the audit colour
convention, each of which is a silent wrong-output bug rather than a crash.
"""

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- Audit colour convention -------------------------------------------------
# Blue = a number a human typed. Black = a formula. Green = another sheet.
# Red = another file. Reviewers read the colour before they read the number, so
# getting this wrong makes a model look hand-checked when it was not.
BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
RED = "FF0000"
YELLOW_FILL = "FFFF00"

HEADER_FILL = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
BAND_FILL = "F2F6FA"

# Default face. Keep it to fonts a Chinese user's Excel definitely has —
# rendering happens on their machine, not ours.
FONT_NAME = "微软雅黑"
FONT_NAME_LATIN = "Arial"

FMT = {
    "money": "#,##0;(#,##0);-",
    "money_cny": "¥#,##0;(¥#,##0);-",
    "money_mm": "#,##0.0;(#,##0.0);-",
    "int": "#,##0;(#,##0);-",
    "pct": "0.0%",
    "pct0": "0%",
    "mult": "0.0x",
    "date": "yyyy-mm-dd",
    "text": "@",
}

THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def cjk_width(text) -> int:
    """Display width in half-widths: CJK/full-width glyphs count as 2.

    ``len()`` under-measures Chinese by half, which is exactly why auto-fitted
    columns come out too narrow and numbers render as ``###``.
    """
    if text is None:
        return 0
    width = 0
    for ch in str(text):
        # CJK ideographs, kana, full-width forms, and CJK punctuation.
        if "一" <= ch <= "鿿" or "　" <= ch <= "〿" or "぀" <= ch <= "ヿ" or "＀" <= ch <= "￯":
            width += 2
        else:
            width += 1
    return width


def autofit_columns(ws, min_width: int = 8, max_width: int = 60, padding: int = 3) -> None:
    """Set column widths from the widest cell, counting CJK as double width."""
    widest: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            # A formula's own text says nothing about how wide its result is;
            # reserve a sane default instead of measuring "=SUM(B2:B99)".
            text = "0,000,000" if isinstance(cell.value, str) and cell.value.startswith("=") else cell.value
            widest[cell.column] = max(widest.get(cell.column, 0), cjk_width(text))
    for col, width in widest.items():
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, width + padding))


def style_header(ws, row: int = 1, first_col: int = 1, last_col: int | None = None) -> None:
    """Dark fill + white bold text + centred, the convention reviewers expect."""
    last_col = last_col or ws.max_column
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = 22


def write_table(ws, headers, rows, start_row: int = 1, start_col: int = 1, number_format=None, banded=True):
    """Write a header row + data rows and return the first/last data row numbers.

    ``number_format`` maps a 0-based column offset to a key of :data:`FMT` (or a
    raw format string), e.g. ``{1: "money", 2: "pct"}``.
    """
    number_format = number_format or {}
    for offset, title in enumerate(headers):
        ws.cell(row=start_row, column=start_col + offset, value=title)
    style_header(ws, row=start_row, first_col=start_col, last_col=start_col + len(headers) - 1)

    for r, record in enumerate(rows, start=start_row + 1):
        for offset, value in enumerate(record):
            cell = ws.cell(row=r, column=start_col + offset, value=value)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER_ALL
            key = number_format.get(offset)
            if key:
                cell.number_format = FMT.get(key, key)
            if banded and (r - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=BAND_FILL)
    first_data_row = start_row + 1
    return first_data_row, start_row + len(rows)


def set_pct(cell, fraction, fmt: str = "pct") -> None:
    """Write a percentage. Pass 0.15 for 15%.

    Storing 15 with a percent format renders 1500.0%. This wrapper refuses the
    ambiguous range loudly instead of shipping a wrong-by-100x workbook.
    """
    if fraction is not None and abs(float(fraction)) > 1.5:
        raise ValueError(
            f"set_pct got {fraction!r}: percentages are stored as fractions "
            f"(0.15 == 15%). Pass {float(fraction) / 100!r} instead, or use "
            f"cell.number_format directly if this really is {fraction}x."
        )
    cell.value = fraction
    cell.number_format = FMT.get(fmt, fmt)


def mark_input(cell, note: str | None = None) -> None:
    """A number a human typed: blue. Optionally attach the source as a comment."""
    cell.font = Font(name=FONT_NAME, size=10, color=BLUE)
    if note:
        add_note(cell, note)


def mark_formula(cell) -> None:
    cell.font = Font(name=FONT_NAME, size=10, color=BLACK)


def mark_cross_sheet(cell) -> None:
    cell.font = Font(name=FONT_NAME, size=10, color=GREEN)


def mark_assumption(cell, note: str | None = None) -> None:
    """A key assumption / a cell the user is meant to fill in: yellow fill."""
    cell.fill = PatternFill("solid", fgColor=YELLOW_FILL)
    cell.font = Font(name=FONT_NAME, size=10, color=BLUE)
    if note:
        add_note(cell, note)


def add_note(cell, text: str, author: str = "BiSheng") -> None:
    """Attach a cell comment. This is where an assumption's source belongs."""
    comment = Comment(text, author)
    comment.width = 260
    comment.height = 110
    cell.comment = comment


def sheet_ref(sheet_name: str, coord: str) -> str:
    """Build a cross-sheet reference, quoting the name when it needs quoting.

    ``='Assumptions Inputs'!$B$5`` evaluates; the unquoted form yields #VALUE!.
    """
    needs_quotes = any(ch in sheet_name for ch in " -()&'") or not sheet_name.isascii()
    name = "'" + sheet_name.replace("'", "''") + "'" if needs_quotes else sheet_name
    return f"={name}!{coord}"


def freeze_and_filter(ws, header_row: int = 1, add_filter: bool = True) -> None:
    """Freeze everything above the first data row and turn on autofilter."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if add_filter and ws.max_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def add_legend(ws, cell_coord: str, lines) -> None:
    """Drop a short legend block. A workbook someone must fill in needs one."""
    row = ws[cell_coord].row
    col = ws[cell_coord].column
    ws.cell(row=row, column=col, value="填写说明").font = Font(name=FONT_NAME, bold=True, size=10)
    for offset, line in enumerate(lines, start=1):
        cell = ws.cell(row=row + offset, column=col, value=line)
        cell.font = Font(name=FONT_NAME, size=9, color="595959")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
